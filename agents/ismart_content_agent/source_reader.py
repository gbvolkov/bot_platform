from __future__ import annotations

import re
from html import unescape
from pathlib import Path
from typing import Any

import openpyxl

from .contracts import DonorSource, GenerationRequest


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_COURSE_TRACKER = (
    REPO_ROOT
    / "docs"
    / "ismart"
    / "Материалы для ИИ-агентов"
    / "Копия Python_DOP_Базовый_v4.xlsx"
)
DEFAULT_DONORS = (
    REPO_ROOT
    / "docs"
    / "ismart"
    / "Материалы для ИИ-агентов"
    / "donory_po_zanyatiyam_python.html"
)

AUDIENCE_SHEETS = {
    "8-9": "8-9 классы",
    "10-11": "10-11 классы",
    "СПО": "СПО",
}


def resolve_request_sources(request: GenerationRequest) -> GenerationRequest:
    if request.task_spec or request.lesson_number is None:
        return request

    task_spec = extract_task_spec(
        tracker_path=_resolve_path(request.course_tracker_path, DEFAULT_COURSE_TRACKER),
        audience=request.audience,
        lesson_number=request.lesson_number,
        target_task_level=request.target_task_level,
        target_task_number=request.target_task_number,
    )
    donor_sources = extract_donor_sources(
        donors_path=_resolve_path(request.donors_path, DEFAULT_DONORS),
        lesson_number=request.lesson_number,
    )
    source_refs = [
        f"{task_spec['source_document']}::{task_spec['sheet_name']}::row {task_spec['row_index']}",
        f"{DEFAULT_DONORS.name}::з{request.lesson_number}",
    ]

    return request.model_copy(
        update={
            "module_id": request.module_id or f"module-{task_spec['module_number']}",
            "lesson_id": request.lesson_id or f"lesson-{task_spec['lesson_number']}",
            "topic": request.topic or task_spec["topic"],
            "lesson_title": request.lesson_title or task_spec["lesson_title"],
            "learning_goal": request.learning_goal
            or f"Обучающийся выполняет задание из программы занятия: {task_spec['task_text']}",
            "task_spec": task_spec,
            "donor_sources": donor_sources,
            "source_refs": [*request.source_refs, *source_refs],
        }
    )


def extract_task_spec(
    *,
    tracker_path: Path,
    audience: str,
    lesson_number: int,
    target_task_level: str | None,
    target_task_number: int | None,
) -> dict[str, Any]:
    sheet_name = AUDIENCE_SHEETS.get(audience)
    if not sheet_name:
        raise ValueError(f"Unsupported audience for course tracker: {audience}")

    workbook = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Course tracker sheet not found: {sheet_name}")

    row_index, row_data = _find_lesson_row(workbook[sheet_name], lesson_number)
    tasks = _extract_level_tasks(row_data["audience_content"])
    if not tasks:
        raise ValueError(f"Lesson {lesson_number} does not contain L1/L2/L3 tasks")

    selected = _select_task(
        tasks,
        target_task_level=target_task_level,
        target_task_number=target_task_number,
    )
    module_number = _module_number(row_data["module"])
    return {
        "source_document": str(tracker_path.relative_to(REPO_ROOT)),
        "sheet_name": sheet_name,
        "row_index": row_index,
        "lesson_number": lesson_number,
        "module_number": module_number,
        "module": row_data["module"],
        "topic": row_data["topic"],
        "lesson_title": row_data["lesson_title"],
        "lesson_type": row_data["lesson_type"],
        "hours": row_data["hours"],
        "common_content": row_data["common_content"],
        "audience_content": row_data["audience_content"],
        "task_level": selected["task_level"],
        "task_number": selected["task_number"],
        "task_number_in_level": selected["task_number_in_level"],
        "task_text": selected["task_text"],
        "difficulty_level": _difficulty_from_task_level(selected["task_level"]),
    }


def list_task_specs(
    *,
    tracker_path: Path,
    audience: str,
    limit: int,
) -> list[dict[str, Any]]:
    sheet_name = AUDIENCE_SHEETS.get(audience)
    if not sheet_name:
        raise ValueError(f"Unsupported audience for course tracker: {audience}")
    workbook = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Course tracker sheet not found: {sheet_name}")

    tasks: list[dict[str, Any]] = []
    current_module = ""
    current_topic = ""
    for row_index, row in enumerate(workbook[sheet_name].iter_rows(values_only=True), start=1):
        if len(row) > 1 and row[1]:
            current_module = _clean(row[1])
        if len(row) > 2 and row[2]:
            current_topic = _clean(row[2])
        lesson_number = _int_or_none(row[0] if row else None)
        if lesson_number is None:
            continue
        audience_content = _clean(row[7] if len(row) > 7 else "")
        for selected in _extract_level_tasks(audience_content):
            tasks.append(
                {
                    "source_document": str(tracker_path.relative_to(REPO_ROOT)),
                    "sheet_name": sheet_name,
                    "row_index": row_index,
                    "lesson_number": lesson_number,
                    "module_number": _module_number(current_module),
                    "module": current_module,
                    "topic": current_topic,
                    "lesson_title": _clean(row[3] if len(row) > 3 else ""),
                    "lesson_type": _clean(row[4] if len(row) > 4 else ""),
                    "hours": _clean(row[5] if len(row) > 5 else ""),
                    "common_content": _clean(row[6] if len(row) > 6 else ""),
                    "audience_content": audience_content,
                    "task_level": selected["task_level"],
                    "task_number": selected["task_number"],
                    "task_number_in_level": selected["task_number_in_level"],
                    "task_text": selected["task_text"],
                    "difficulty_level": _difficulty_from_task_level(selected["task_level"]),
                }
            )
            if len(tasks) >= limit:
                return tasks
    if len(tasks) < limit:
        raise ValueError(f"Course tracker contains only {len(tasks)} tasks, requested {limit}")
    return tasks


def resolve_source_path(value: str | None, default: Path) -> Path:
    return _resolve_path(value, default)


def extract_donor_sources(*, donors_path: Path, lesson_number: int) -> list[DonorSource]:
    html = donors_path.read_text(encoding="utf-8")
    row_match = re.search(
        rf"<tr><td[^>]*>\s*<b>з{lesson_number}\.</b>(?P<row>.*?)</tr>",
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not row_match:
        return []

    row_html = row_match.group("row")
    donor_cell = _cell_by_class(row_html, "donor")
    take_cell = _cell_by_class(row_html, "take")
    donors: list[DonorSource] = []

    anchor_pattern = re.compile(
        r'<a href="(?P<href>[^"]+)"[^>]*>(?P<title>.*?)</a>(?P<tail>.*?)(?=<br>|$)',
        flags=re.IGNORECASE | re.DOTALL,
    )
    for index, match in enumerate(anchor_pattern.finditer(donor_cell), start=1):
        tail = _strip_tags(match.group("tail"))
        mode = _donor_mode(tail)
        donors.append(
            DonorSource(
                donor_id=f"lesson-{lesson_number}-donor-{index}",
                donor_mode=mode,
                source_title=_strip_tags(match.group("title")),
                source_uri=unescape(match.group("href")),
                attribution_required=mode == "прямой",
                rewrite_required=mode == "ориентир",
            )
        )

    if donors:
        return donors

    text = _strip_tags(donor_cell)
    if "ваше" in text or "собственная" in text or "платформа" in text:
        return [
            DonorSource(
                donor_id=f"lesson-{lesson_number}-own",
                donor_mode="ваше",
                source_title=text or _strip_tags(take_cell) or "собственная разработка",
            )
        ]
    return []


def _resolve_path(value: str | None, default: Path) -> Path:
    if not value:
        return default
    path = Path(value)
    return path if path.is_absolute() else REPO_ROOT / path


def _find_lesson_row(sheet: Any, lesson_number: int) -> tuple[int, dict[str, str]]:
    current_module = ""
    current_topic = ""
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        number = _int_or_none(row[0] if len(row) > 0 else None)
        if len(row) > 1 and row[1]:
            current_module = str(row[1]).strip()
        if len(row) > 2 and row[2]:
            current_topic = str(row[2]).strip()
        if number != lesson_number:
            continue
        return row_index, {
            "module": current_module,
            "topic": current_topic,
            "lesson_title": _clean(row[3] if len(row) > 3 else ""),
            "lesson_type": _clean(row[4] if len(row) > 4 else ""),
            "hours": _clean(row[5] if len(row) > 5 else ""),
            "common_content": _clean(row[6] if len(row) > 6 else ""),
            "audience_content": _clean(row[7] if len(row) > 7 else ""),
        }
    raise ValueError(f"Lesson number not found in course tracker: {lesson_number}")


def _extract_level_tasks(content: str) -> list[dict[str, Any]]:
    tasks: list[dict[str, Any]] = []
    block_pattern = re.compile(
        r"(?P<level>L[123]):\s*(?P<body>.*?)(?=(?:\n\s*L[123]:)|(?:\n\s*Дополнительно:)|\Z)",
        flags=re.DOTALL,
    )
    for block in block_pattern.finditer(content):
        level = block.group("level")
        for index, task in enumerate(_split_numbered_tasks(block.group("body")), start=1):
            tasks.append(
                {
                    "task_level": level,
                    "task_number": task["task_number"],
                    "task_number_in_level": index,
                    "task_text": task["task_text"],
                }
            )
    return tasks


def _split_numbered_tasks(text: str) -> list[dict[str, Any]]:
    candidates = list(re.finditer(r"(?<!\d)(?P<number>\d+)\.\s+", text))
    matches: list[re.Match[str]] = []
    for candidate in candidates:
        number = int(candidate.group("number"))
        if not matches or number == int(matches[-1].group("number")) + 1:
            matches.append(candidate)
    if not matches:
        cleaned = _clean(text)
        return [{"task_number": 1, "task_text": cleaned}] if cleaned else []

    tasks: list[dict[str, Any]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        task_text = _clean(text[start:end])
        if task_text:
            tasks.append({"task_number": int(match.group("number")), "task_text": task_text})
    return tasks


def _select_task(
    tasks: list[dict[str, Any]],
    *,
    target_task_level: str | None,
    target_task_number: int | None,
) -> dict[str, Any]:
    candidates = tasks
    if target_task_level:
        candidates = [task for task in candidates if task["task_level"] == target_task_level]
    if target_task_number is not None:
        candidates = [task for task in candidates if task["task_number"] == target_task_number]
    if not candidates:
        raise ValueError(
            f"Requested task not found: level={target_task_level!r}, number={target_task_number!r}"
        )
    return candidates[0]


def _difficulty_from_task_level(level: str) -> int:
    return {"L1": 1, "L2": 2, "L3": 3}[level]


def _module_number(module: str) -> int:
    match = re.match(r"\s*(\d+)", module)
    return int(match.group(1)) if match else 0


def _cell_by_class(row_html: str, class_name: str) -> str:
    match = re.search(
        rf'<td class="{class_name}">(?P<cell>.*?)</td>',
        row_html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    return match.group("cell") if match else ""


def _donor_mode(text: str) -> str:
    if "прямой" in text:
        return "прямой"
    if "ориентир" in text:
        return "ориентир"
    return "ваше"


def _strip_tags(value: str) -> str:
    return _clean(re.sub(r"<[^>]+>", " ", unescape(value)))


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[ \t]+", " ", str(value).replace("\xa0", " ")).strip()


def _int_or_none(value: Any) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None
