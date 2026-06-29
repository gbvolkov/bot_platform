from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REFERENCE_FIELDS = (
    "requirements",
    "reference_examples",
    "goals_and_tasks",
    "donor_materials",
    "template_descriptions",
)

LESSON_REFERENCE_COLUMNS = {
    "requirements": 27,
    "reference_examples": 28,
    "goals_and_tasks": 29,
    "donor_materials": 30,
    "template_descriptions": 31,
}


def convert_tracker_to_generation_json(
    workbook_path: Path,
    *,
    sheet_name: str,
    output_path: Path | None = None,
    workspace_dir: Path | None = None,
    references_dir: Path | None = None,
) -> dict[str, Any]:
    workbook_path = workbook_path.resolve()
    workspace_dir = workspace_dir.resolve() if workspace_dir else find_workspace_dir(workbook_path)
    references_dir = references_dir.resolve() if references_dir else find_references_dir(workspace_dir)
    output_path = (output_path or default_output_path(workspace_dir, sheet_name)).resolve()

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet not found: {sheet_name}. Available sheets: {available}")
    sheet = workbook[sheet_name]

    missing_references: list[dict[str, Any]] = []
    ignored_references: list[dict[str, Any]] = []
    parse_warnings: list[dict[str, Any]] = []
    modules: list[dict[str, Any]] = []
    current_module: dict[str, Any] | None = None

    for row in range(5, sheet.max_row + 1):
        first_cell = clean_text(sheet.cell(row, 1).value) or ""
        module_cell = sheet.cell(row, 2).value
        lesson_cell = sheet.cell(row, 3).value
        lesson_type_cell = sheet.cell(row, 4).value

        if lesson_cell is None and module_cell is None and first_cell:
            if "[" in first_cell and "]" in first_cell and "L1=" in first_cell:
                current_module = parse_module_header(first_cell)
                modules.append(current_module)
            elif current_module is not None:
                current_module["totals"] = parse_totals(first_cell)
            continue

        if current_module is None or lesson_cell is None or lesson_type_cell is None:
            continue

        lesson = build_lesson(
            sheet,
            row,
            references_dir=references_dir,
            missing_references=missing_references,
            ignored_references=ignored_references,
        )
        if lesson["lesson_number"] is None:
            parse_warnings.append({"row": row, "issue": "lesson_number_not_found"})

        for level, column in (("l1", 8), ("l2", 9), ("l3", 10)):
            value = clean_text(sheet.cell(row, column).value)
            if value and not lesson["practice_tasks"][level] and re.search(r"\d", value):
                if value.lstrip().startswith("["):
                    continue
                parse_warnings.append(
                    {
                        "row": row,
                        "field": f"practice_tasks.{level}",
                        "issue": "numeric_text_without_parsed_tasks",
                        "text": value[:300],
                    }
                )

        current_module["lessons"].append(lesson)

    payload = {
        "course": {
            "title": clean_text(sheet.cell(1, 1).value),
            "norms": clean_text(sheet.cell(2, 1).value),
            "legend": clean_text(sheet.cell(3, 1).value),
            "source_workbook": as_posix(workbook_path),
            "source_sheet": sheet.title,
        },
        "markdown_references_base": as_posix(references_dir),
        "modules": modules,
        "_conversion": {
            "output": as_posix(output_path),
            "source_workbook": as_posix(workbook_path),
            "source_sheet": sheet.title,
            "workspace_dir": as_posix(workspace_dir),
            "markdown_references_base": as_posix(references_dir),
            "module_count": len(modules),
            "lesson_count": sum(len(module.get("lessons") or []) for module in modules),
            "missing_markdown_references": missing_references,
            "ignored_reference_cells": ignored_references,
            "parse_warnings": parse_warnings,
        },
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def build_lesson(
    sheet: Any,
    row: int,
    *,
    references_dir: Path,
    missing_references: list[dict[str, Any]],
    ignored_references: list[dict[str, Any]],
) -> dict[str, Any]:
    topic, title, title_raw, lesson_number = parse_lesson_title(sheet.cell(row, 3).value)
    if lesson_number is None:
        lesson_number = to_int(sheet.cell(row, 1).value)

    self_work = bool_cell(sheet, row, 18)
    attestation = bool_cell(sheet, row, 20)
    return {
        "lesson_number": lesson_number,
        "tracker_row": row,
        "tracker_index": to_int(sheet.cell(row, 1).value),
        "module": clean_text(sheet.cell(row, 2).value),
        "topic": topic,
        "title": title,
        "title_raw": title_raw,
        "type": clean_text(sheet.cell(row, 4).value),
        "hours": parse_hours(sheet.cell(row, 5).value),
        "content": {
            "general": clean_text(sheet.cell(row, 6).value),
            "audience": sheet.title,
            "audience_specific": clean_text(sheet.cell(row, 7).value),
            "for_grades_8_9": clean_text(sheet.cell(row, 7).value),
        },
        "practice_tasks": {
            "l1": parse_tasks(sheet.cell(row, 8).value),
            "l2": parse_tasks(sheet.cell(row, 9).value),
            "l3": parse_tasks(sheet.cell(row, 10).value),
        },
        "teacher_materials": {
            "theory": clean_text(sheet.cell(row, 11).value),
            "practice": clean_text(sheet.cell(row, 12).value),
        },
        "content_flags": {
            "theory": bool_cell(sheet, row, 13),
            "practice": bool_cell(sheet, row, 14),
            "case_task": bool_cell(sheet, row, 15),
            "project": bool_cell(sheet, row, 16),
            "template": bool(self_work or attestation),
            "group_work_excluded": bool_cell(sheet, row, 17),
            "self_work": self_work,
            "current_control": bool_cell(sheet, row, 19),
            "attestation": attestation,
        },
        "difficulty": {
            "l1": {
                "count": to_int(sheet.cell(row, 21).value),
                "percent": percent_value(sheet.cell(row, 22).value),
            },
            "l2": {
                "count": to_int(sheet.cell(row, 23).value),
                "percent": percent_value(sheet.cell(row, 24).value),
            },
            "l3": {
                "count": to_int(sheet.cell(row, 25).value) or 0,
                "raw": clean_text(sheet.cell(row, 25).value),
            },
            "violation": clean_text(sheet.cell(row, 26).value),
        },
        "materials_md": build_references(
            sheet,
            row,
            references_dir=references_dir,
            missing_references=missing_references,
            ignored_references=ignored_references,
        ),
    }


def build_references(
    sheet: Any,
    row: int,
    *,
    references_dir: Path,
    missing_references: list[dict[str, Any]],
    ignored_references: list[dict[str, Any]],
) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {field: [] for field in REFERENCE_FIELDS}
    seen: set[str] = set()

    for field, column in LESSON_REFERENCE_COLUMNS.items():
        for raw in split_reference_cell(sheet.cell(row, column).value):
            resolved = reference_to_markdown_path(raw, references_dir)
            if resolved is None:
                item = {"row": row, "field": field, "source": raw}
                if is_service_reference_cell(raw):
                    ignored_references.append(item)
                else:
                    missing_references.append(item)
                continue
            path_value = as_posix(resolved)
            if path_value not in seen:
                result[field].append(path_value)
                seen.add(path_value)

    return result


def parse_module_header(value: Any) -> dict[str, Any]:
    text = clean_text(value) or ""
    numbers = [int(item) for item in re.findall(r"\d+", text)]
    hours = numbers[1] if len(numbers) > 1 else None
    return {
        "title": text,
        "hours": hours,
        "l1": parse_count_percent(text, "L1"),
        "l2": parse_count_percent(text, "L2"),
        "l3": parse_count_percent(text, "L3"),
        "lessons": [],
    }


def parse_totals(value: Any) -> dict[str, Any]:
    text = clean_text(value) or ""
    numbers = [int(item) for item in re.findall(r"\d+", text)]
    total_tasks = numbers[1] if len(numbers) > 1 else None
    l3 = parse_count_percent(text, "L3")
    return {
        "raw": text,
        "total_tasks": total_tasks,
        "l1": parse_count_percent(text, "L1"),
        "l2": parse_count_percent(text, "L2"),
        "l3_tasks": l3.get("tasks"),
        "l3": l3,
    }


def parse_count_percent(text: str, key: str) -> dict[str, Any]:
    match = re.search(rf"{re.escape(key)}\s*=\s*(\d+)(?:\(([^)]*)\))?", text)
    if not match:
        return {"tasks": None, "percent": None}
    return {"tasks": int(match.group(1)), "percent": match.group(2)}


def parse_lesson_title(value: Any) -> tuple[str | None, str | None, str | None, int | None]:
    text = clean_text(value) or ""
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    topic = lines[0] if lines else None
    title_raw = lines[1] if len(lines) > 1 else (lines[0] if lines else None)
    title_without_icon = re.sub(r"^\D*?(\d+\.)", r"\1", title_raw or "").strip()
    match = re.match(r"^(\d+)\.\s*(.+)$", title_without_icon)
    if match:
        return topic, match.group(2).strip(), title_raw, int(match.group(1))
    return topic, title_without_icon or title_raw, title_raw, None


def parse_hours(raw: Any) -> dict[str, Any]:
    text = clean_text(raw) or ""
    result: dict[str, Any] = {
        "raw": text,
        "theory": 0,
        "practice": 0,
        "self_study": 0,
        "assessment": 0,
    }
    mapping = {
        "\u0422": "theory",
        "\u041f": "practice",
        "\u0421\u0420": "self_study",
        "\u041a": "assessment",
    }
    for marker, value in re.findall(
        r"(\u0421\u0420|\u0422|\u041f|\u041a)\s*:\s*([0-9]+(?:[,.][0-9]+)?)\s*\u0447",
        text,
    ):
        number = float(value.replace(",", "."))
        result[mapping[marker]] = int(number) if number.is_integer() else number
    return result


def parse_tasks(value: Any) -> list[dict[str, Any]]:
    text = clean_text(value)
    if not text:
        return []

    bonus_marker = re.search(r"(?m)^\s*\[[^\]]+\]", text)
    if bonus_marker:
        text = text[: bonus_marker.start()].strip()

    matches = list(re.finditer(r"(?<!\d)(\d{1,2})\.\s+", text))
    tasks: list[dict[str, Any]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        task_text = text[start:end].strip()
        task_text = re.sub(r"^L[123]\s*:\s*", "", task_text).strip()
        if task_text:
            tasks.append({"number": int(match.group(1)), "text": task_text})
    return tasks


def reference_to_markdown_path(raw: str, references_dir: Path) -> Path | None:
    item = raw.strip()
    suffix = Path(item).suffix.lower()
    if suffix == ".md":
        candidate = Path(item).name
    elif suffix in {".docx", ".pptx", ".pdf", ".html"}:
        candidate = Path(item).with_suffix(".md").name
    else:
        candidate = Path(f"{item}.md").name

    path = references_dir / candidate
    return path.resolve() if path.exists() else None


def split_reference_cell(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [
        line.strip().strip(";").strip()
        for line in text.split("\n")
        if line.strip().strip(";").strip()
    ]


def is_service_reference_cell(value: str) -> bool:
    text = value.strip()
    return Path(text).suffix == "" and not any(char.isdigit() for char in text)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text or None


def to_int(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value) if int(value) == value else None
    text = str(value).strip()
    if text in {"", "-", "\u2014"}:
        return None
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return None


def percent_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return None if text in {"", "-", "\u2014"} else text
    if isinstance(value, (int, float)):
        number = float(value)
        if 0 <= number <= 1:
            number *= 100
        rounded = round(number, 1)
        return f"{int(rounded)}%" if rounded.is_integer() else f"{rounded}%"
    return str(value).strip() or None


def bool_cell(sheet: Any, row: int, column: int) -> bool:
    return bool(sheet.cell(row, column).value)


def find_workspace_dir(workbook_path: Path) -> Path:
    for candidate in workbook_path.parent.iterdir():
        if candidate.is_dir() and (candidate / "prompts_skills").is_dir():
            return candidate.resolve()
    raise FileNotFoundError(
        "Could not locate workspace dir with prompts_skills next to the tracker. "
        "Pass --workspace-dir explicitly."
    )


def find_references_dir(workspace_dir: Path) -> Path:
    for candidate in workspace_dir.iterdir():
        if candidate.is_dir() and candidate.name != "prompts_skills":
            if any(child.suffix.lower() == ".md" for child in candidate.iterdir()):
                return candidate.resolve()
    raise FileNotFoundError(
        "Could not locate Markdown references dir under workspace dir. "
        "Pass --references-dir explicitly."
    )


def default_output_path(workspace_dir: Path, sheet_name: str) -> Path:
    slug = re.sub(r"\W+", "_", sheet_name.lower(), flags=re.UNICODE).strip("_")
    slug = slug or "sheet"
    return workspace_dir / f"generation_input_{slug}_from_tracker.json"


def as_posix(path: Path) -> str:
    return str(path).replace("\\", "/")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert an iSMART course tracker .xlsx sheet into iSMART generator JSON."
    )
    parser.add_argument("--input", required=True, help="Path to tracker .xlsx.")
    parser.add_argument("--sheet", default="8-9 классы", help="Worksheet name to convert.")
    parser.add_argument("--output", help="Path to write generation JSON.")
    parser.add_argument(
        "--workspace-dir",
        help="Workspace dir containing prompts_skills and Markdown references. Auto-detected by default.",
    )
    parser.add_argument(
        "--references-dir",
        help="Markdown references directory. Auto-detected from workspace dir by default.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    args = build_arg_parser().parse_args(argv)
    payload = convert_tracker_to_generation_json(
        Path(args.input),
        sheet_name=args.sheet,
        output_path=Path(args.output) if args.output else None,
        workspace_dir=Path(args.workspace_dir) if args.workspace_dir else None,
        references_dir=Path(args.references_dir) if args.references_dir else None,
    )
    conversion = payload["_conversion"]
    print(
        json.dumps(
            {
                "output": conversion["output"],
                "modules": conversion["module_count"],
                "lessons": conversion["lesson_count"],
                "missing_markdown_references": len(conversion["missing_markdown_references"]),
                "ignored_reference_cells": len(conversion["ignored_reference_cells"]),
                "parse_warnings": len(conversion["parse_warnings"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
