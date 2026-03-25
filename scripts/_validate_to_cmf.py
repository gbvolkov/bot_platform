from __future__ import annotations

import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "gaz-pricing"
DB_PATH = DATA_DIR / "gaz_pricing.sqlite"


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").replace("\r", " ").replace("\n", " ")
    text = " ".join(text.split()).strip()
    return text or None


def parse_number(value: object) -> float | None:
    text = clean_text(value)
    if text is None:
        return None
    compact = text.replace(" ", "").replace(",", ".")
    try:
        return float(compact)
    except ValueError:
        return None


def parse_service_index(value: object) -> int | None:
    text = clean_text(value)
    if text is None:
        return None
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def build_expected_rows(workbook_path: Path, sheet_name: str) -> tuple[list[tuple[object, ...]], int]:
    workbook = load_workbook(workbook_path, data_only=True)
    ws = workbook[sheet_name]
    rows: list[tuple[object, ...]] = []
    raw_count = 0

    blocks = [
        {
            "group_name": clean_text(ws["D2"].value),
            "row_start": 5,
            "row_end": 14,
            "service_col": 2,
            "mileage_col": 3,
            "diesel_col": 4,
            "diesel_promo_col": None,
            "gasoline_col": None,
            "gasoline_promo_col": None,
        },
        {
            "group_name": clean_text(ws["E2"].value),
            "row_start": 5,
            "row_end": 14,
            "service_col": 2,
            "mileage_col": 3,
            "diesel_col": 5,
            "diesel_promo_col": None,
            "gasoline_col": 6,
            "gasoline_promo_col": None,
        },
        {
            "group_name": clean_text(ws["I2"].value),
            "row_start": 4,
            "row_end": 14,
            "service_col": 7,
            "mileage_col": 8,
            "diesel_col": 9,
            "diesel_promo_col": 10,
            "gasoline_col": None,
            "gasoline_promo_col": None,
        },
        {
            "group_name": clean_text(ws["K2"].value),
            "row_start": 4,
            "row_end": 14,
            "service_col": 7,
            "mileage_col": 8,
            "diesel_col": 11,
            "diesel_promo_col": 12,
            "gasoline_col": None,
            "gasoline_promo_col": None,
        },
        {
            "group_name": clean_text(ws["M2"].value),
            "row_start": 4,
            "row_end": 16,
            "service_col": 13,
            "mileage_col": 14,
            "diesel_col": 15,
            "diesel_promo_col": None,
            "gasoline_col": None,
            "gasoline_promo_col": None,
        },
    ]

    for block in blocks:
        cols = [
            block["service_col"],
            block["mileage_col"],
            block["diesel_col"],
            block["diesel_promo_col"],
            block["gasoline_col"],
            block["gasoline_promo_col"],
        ]
        for row_index in range(block["row_start"], block["row_end"] + 1):
            service_label = clean_text(ws.cell(row_index, block["service_col"]).value)
            if not service_label:
                continue
            rows.append(
                (
                    block["group_name"],
                    service_label,
                    parse_service_index(service_label),
                    parse_number(ws.cell(row_index, block["mileage_col"]).value),
                    parse_number(ws.cell(row_index, block["diesel_col"]).value) if block["diesel_col"] else None,
                    parse_number(ws.cell(row_index, block["diesel_promo_col"]).value) if block["diesel_promo_col"] else None,
                    parse_number(ws.cell(row_index, block["gasoline_col"]).value) if block["gasoline_col"] else None,
                    parse_number(ws.cell(row_index, block["gasoline_promo_col"]).value) if block["gasoline_promo_col"] else None,
                )
            )
            for column_index in cols:
                if column_index is None:
                    continue
                if clean_text(ws.cell(row_index, column_index).value) is not None:
                    raw_count += 1

    workbook.close()
    return rows, raw_count


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    problems: list[tuple[object, ...]] = []

    workbook_paths = []
    for workbook_path in sorted(DATA_DIR.glob("*.xlsx")):
        workbook = load_workbook(workbook_path, data_only=True)
        if "ТО ЦМФ" in workbook.sheetnames:
            workbook_paths.append(workbook_path)
        workbook.close()

    for workbook_path in workbook_paths:
        expected_rows, expected_raw_count = build_expected_rows(workbook_path, "ТО ЦМФ")
        actual_rows = conn.execute(
            """
            SELECT group_name, service_label, service_index, mileage_km,
                   cost_diesel_rub, cost_diesel_promo_rub, cost_gasoline_rub, cost_gasoline_promo_rub
            FROM comparison_service_groups
            WHERE source_file = ? AND source_sheet = ?
            ORDER BY group_name, service_index
            """,
            (workbook_path.name, "ТО ЦМФ"),
        ).fetchall()
        actual_tuples = [tuple(row) for row in actual_rows]
        expected_sorted = sorted(expected_rows, key=lambda row: (row[0] or "", row[2] or 0))
        if actual_tuples != expected_sorted:
            problems.append((workbook_path.name, "service_rows_mismatch", len(expected_rows), len(actual_tuples)))

        actual_raw_count = conn.execute(
            """
            SELECT COUNT(*)
            FROM comparisons_raw_params
            WHERE source_file = ? AND source_sheet = ? AND record_scope = 'service_group'
            """,
            (workbook_path.name, "ТО ЦМФ"),
        ).fetchone()[0]
        if actual_raw_count != expected_raw_count:
            problems.append((workbook_path.name, "raw_count", expected_raw_count, actual_raw_count))

        non_service_raw = conn.execute(
            """
            SELECT COUNT(*)
            FROM comparisons_raw_params
            WHERE source_file = ? AND source_sheet = ? AND record_scope <> 'service_group'
            """,
            (workbook_path.name, "ТО ЦМФ"),
        ).fetchone()[0]
        if non_service_raw != 0:
            problems.append((workbook_path.name, "non_service_raw", 0, non_service_raw))

    print(f"workbooks={len(workbook_paths)}")
    print(f"problems={len(problems)}")
    for problem in problems[:50]:
        print(problem)


if __name__ == "__main__":
    main()
