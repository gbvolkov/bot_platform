from __future__ import annotations

import re
import sqlite3
import sys
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = REPO_ROOT / "data" / "gaz-pricing"
DB_PATH = SOURCE_DIR / "gaz_pricing_norm.sqlite"

NORMALIZED_COLUMNS = [
    "id",
    "source_file",
    "source_sheet",
    "sheet_type",
    "base_model",
    "comp_full_name",
    "comp_brand",
    "comp_model",
    "price_rub_min",
    "price_rub_max",
    "price_comment",
    "vehicle_type",
    "body_type",
    "passenger_capacity",
    "seat_count",
    "cab_seat_count",
    "engine_fuel_type",
    "engine_power_hp",
    "engine_power_kw",
    "engine_volume_l",
    "transmission_type",
    "transmission_gears",
    "drive_type",
    "wheel_formula",
    "gross_weight_kg",
    "gross_weight_kg_min",
    "gross_weight_kg_max",
    "curb_weight_kg",
    "curb_weight_kg_min",
    "curb_weight_kg_max",
    "payload_kg",
    "payload_kg_min",
    "payload_kg_max",
    "length_mm",
    "width_mm",
    "height_mm",
    "wheelbase_mm",
    "ground_clearance_mm",
    "fuel_tank_l",
    "fuel_consumption_l100km",
    "service_interval_km",
    "service_interval_months",
    "ownership_cost_rub_km",
    "warranty_months",
    "warranty_km",
    "notes",
]

NORMALIZED_TEXT_COLUMNS = {
    "source_file",
    "source_sheet",
    "sheet_type",
    "base_model",
    "comp_full_name",
    "comp_brand",
    "comp_model",
    "price_comment",
    "vehicle_type",
    "body_type",
    "engine_fuel_type",
    "transmission_type",
    "drive_type",
    "wheel_formula",
    "notes",
}

SERVICE_COLUMNS = [
    "source_file",
    "source_sheet",
    "base_model",
    "group_name",
    "service_label",
    "service_index",
    "mileage_km",
    "cost_diesel_rub",
    "cost_diesel_promo_rub",
    "cost_gasoline_rub",
    "cost_gasoline_promo_rub",
    "notes",
]

SERVICE_TEXT_COLUMNS = {
    "source_file",
    "source_sheet",
    "base_model",
    "group_name",
    "service_label",
    "notes",
}

OPTION_COLUMNS = [
    "source_file",
    "source_sheet",
    "base_model",
    "comp_full_name",
    "comp_brand",
    "option_group",
    "option_name",
    "option_status_raw",
    "option_status_norm",
    "option_price_raw",
    "option_price_rub",
    "notes",
    "row_order",
    "column_order",
]

OPTION_TEXT_COLUMNS = {
    "source_file",
    "source_sheet",
    "base_model",
    "comp_full_name",
    "comp_brand",
    "option_group",
    "option_name",
    "option_status_raw",
    "option_status_norm",
    "option_price_raw",
    "notes",
}

RAW_COLUMNS = [
    "source_file",
    "source_sheet",
    "sheet_type",
    "base_model",
    "record_scope",
    "comp_full_name",
    "comp_brand",
    "group_name",
    "param_name_raw",
    "param_name_norm",
    "value_raw",
    "value_num",
    "value_text",
    "unit_raw",
    "unit_norm",
    "parse_status",
    "parse_comment",
    "cell_address",
    "row_order",
    "column_order",
]

RAW_TEXT_COLUMNS = {
    "source_file",
    "source_sheet",
    "sheet_type",
    "base_model",
    "record_scope",
    "comp_full_name",
    "comp_brand",
    "group_name",
    "param_name_raw",
    "param_name_norm",
    "value_raw",
    "value_text",
    "unit_raw",
    "unit_norm",
    "parse_status",
    "parse_comment",
    "cell_address",
}

SHEET_SVESY_BAZY = "свесы-базы"
SHEET_LIST2 = "Лист2"
SHEET_PAZ = "ПАЗ"
SHEET_PAZ_EXPANDED = "ПАЗ (расширенные)"
SHEET_BORT_TTX = "БОРТ_ТТХ и состав"
SHEET_CMF_TTX = "ЦМФ_ТТХ и состав"
SHEET_LDT_TTX = "LDT_ТТХ и состав"
SHEET_MINIVEN = "Минивен"
SHEET_MINIVEN_2 = "Минивен (2)"
SHEET_BUS_TTX = "АВТОБУС_ТТХ и состав"
SHEET_9_10T = "9-10т"
SHEET_BORT_SHORT = "БОРТ_кратко"
SHEET_CMF_SHORT = "ЦМФ_кратко"
SHEET_BUS_SHORT = "Автобус_кратко"
SHEET_LDT_SHORT = "LDT_ТТХ и состав_кратко"
SHEET_MINIVEN_SHORT = "Минивен_кратко"
SHEET_COMPASS = "Компас-5 и Газель Некст"
SHEET_SF5 = "SF5 и Ко"
SHEET_BORT_STVLAD = "БОРТ_СтВлад"
SHEET_CMF_STVLAD = "ЦМФ_СтВлад"
SHEET_LDT_STVLAD = "LDT_СтВлад"
SHEET_TO_BORT = "ТО Борт"
SHEET_TO_CMF = "ТО ЦМФ"
SHEET_COMPLECTATION = "Комплектация"
SHEET_ADVANTAGES = "Преимущества"
SHEET_ADVANTAGES_MINIVAN = "Преимущества минивэн"
SHEET_REVIEWS = "Отзывы"
SHEET_REVIEWS_LOWER = "отзывы"
SHEET_CLIENT_REVIEWS_NN = "Отзывы_клиенты_НН"
SHEET_CLIENT_REVIEWS_ATLANT = "Отзывы_клиенты_Атлант"
SHEET_NEGATIVE_POSITIVE_ORM = "Негатив-Позитив ORM"
SHEET_QUALITY = "качество"
SHEET_DPO_STATS = "Статистика ДПО"

ID_RULE_MODEL_ONLY = "model_only"
ID_RULE_BRAND_MODEL = "brand_model"
ID_RULE_MANUFACTURER_MODEL = "manufacturer_model"

SOURCE_ALLOWLIST: dict[str, dict[str, str]] = {
    "Вектор Некст_Сравнение с конкурентами ПАЗ.xlsx": {SHEET_PAZ_EXPANDED: ID_RULE_MODEL_ONLY},
    "Газель NEXT_Сравнение с конкурентами.xlsx": {
        SHEET_BORT_TTX: ID_RULE_BRAND_MODEL,
        SHEET_CMF_TTX: ID_RULE_BRAND_MODEL,
        SHEET_BUS_TTX: ID_RULE_MANUFACTURER_MODEL,
    },
    "Газель NN_Сравнение с конкурентами.xlsx": {
        SHEET_BORT_TTX: ID_RULE_BRAND_MODEL,
        SHEET_CMF_TTX: ID_RULE_BRAND_MODEL,
    },
    "Газель Сити_Сравнение с конкурентами.xlsx": {
        SHEET_BUS_TTX: ID_RULE_MANUFACTURER_MODEL,
    },
    "ПАЗ 3205_Сравнение с конкурентами ПАЗ.xlsx": {SHEET_PAZ_EXPANDED: ID_RULE_MODEL_ONLY},
    "ПАЗ 4234_Сравнение с конкурентами ПАЗ.xlsx": {SHEET_PAZ_EXPANDED: ID_RULE_MODEL_ONLY},
    "Садко 9_сравнение с конкурентами.xlsx": {SHEET_LDT_TTX: ID_RULE_BRAND_MODEL},
    "Садко NEXT_сравнение с конкурентами.xlsx": {SHEET_LDT_TTX: ID_RULE_BRAND_MODEL},
    "Ситимакс 8_Сравнение с конкурентами ПАЗ.xlsx": {SHEET_PAZ_EXPANDED: ID_RULE_MODEL_ONLY},
    "Ситимакс 9_Сравнение с конкурентами ПАЗ.xlsx": {SHEET_PAZ_EXPANDED: ID_RULE_MODEL_ONLY},
    "Соболь NN_Минивэн_Сравнение с конкурентами.xlsx": {SHEET_MINIVEN: ID_RULE_BRAND_MODEL},
}

INTEGER_FIELDS = {
    "passenger_capacity",
    "seat_count",
    "cab_seat_count",
    "transmission_gears",
    "service_interval_months",
    "warranty_months",
}

REAL_FIELDS = {
    "price_rub_min",
    "price_rub_max",
    "engine_power_hp",
    "engine_power_kw",
    "engine_volume_l",
    "gross_weight_kg",
    "gross_weight_kg_min",
    "gross_weight_kg_max",
    "curb_weight_kg",
    "curb_weight_kg_min",
    "curb_weight_kg_max",
    "payload_kg",
    "payload_kg_min",
    "payload_kg_max",
    "length_mm",
    "width_mm",
    "height_mm",
    "wheelbase_mm",
    "ground_clearance_mm",
    "fuel_tank_l",
    "fuel_consumption_l100km",
    "service_interval_km",
    "ownership_cost_rub_km",
    "warranty_km",
}

Parser = Callable[[dict[str, Any]], None]


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def lowercase_text(value: Any) -> str | None:
    text = clean_text(value)
    return text.lower() if text else None


def with_nocase_columns(columns: list[str], text_columns: set[str]) -> list[str]:
    expanded: list[str] = []
    for column in columns:
        expanded.append(column)
        if column in text_columns:
            expanded.append(f"{column}_nocase")
    return expanded


NORMALIZED_COLUMNS = with_nocase_columns(NORMALIZED_COLUMNS, NORMALIZED_TEXT_COLUMNS)
SERVICE_COLUMNS = with_nocase_columns(SERVICE_COLUMNS, SERVICE_TEXT_COLUMNS)
OPTION_COLUMNS = with_nocase_columns(OPTION_COLUMNS, OPTION_TEXT_COLUMNS)
RAW_COLUMNS = with_nocase_columns(RAW_COLUMNS, RAW_TEXT_COLUMNS)
NORMALIZED_BASE_COLUMNS = [column for column in NORMALIZED_COLUMNS if not column.endswith("_nocase")]

BOUNDED_NUMERIC_FIELDS = (
    "gross_weight_kg",
    "curb_weight_kg",
    "payload_kg",
)


def normalize_label(value: Any) -> str:
    return (clean_text(value) or "").lower().replace("ё", "е")


TRANSLITERATION_MAP = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "i",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}


def transliterate_to_ascii(value: str) -> str:
    return "".join(TRANSLITERATION_MAP.get(char, char) for char in value)


def normalize_id_part(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    normalized = transliterate_to_ascii(text.lower())
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    if not normalized or normalized == "0":
        return None
    return normalized


def build_id(*parts: Any) -> str | None:
    normalized_parts: list[str] = []
    for part in parts:
        normalized = normalize_id_part(part)
        if not normalized:
            continue
        if normalized in normalized_parts:
            continue
        normalized_parts.append(normalized)
    return "_".join(normalized_parts) if normalized_parts else None


def canonicalize_identity_parts(brand_value: Any, model_value: Any) -> tuple[str | None, str | None, str | None, str | None]:
    brand = clean_text(brand_value)
    model = clean_text(model_value)
    if brand and not model:
        model = brand
    elif model and not brand:
        brand = model
    full_name = join_parts(brand, model) or brand or model
    return brand, model, full_name, build_id(brand, model)


def normalize_unit(label: str) -> tuple[str | None, str | None]:
    norm = normalize_label(label)
    if not norm:
        return None, None
    if "затраты на 1 км" in norm or "руб/км" in norm:
        return "руб/км", "rub_km"
    if "л/100" in norm:
        return "л/100 км", "l100km"
    if "куб" in norm or "см3" in norm or "см³" in norm:
        return "куб.см", "cc"
    if "руб" in norm or "цена" in norm or "стоим" in norm:
        return "руб.", "rub"
    if re.search(r"\bкг\b", norm) or "масса" in norm or "грузопод" in norm:
        return "кг", "kg"
    if re.search(r"\bмм\b", norm) or "дхшхв" in norm:
        return "мм", "mm"
    if re.search(r"\bл\.?\s*с\.?\b", norm):
        return "л.с.", "hp"
    if re.search(r"\bквт\b", norm):
        return "кВт", "kw"
    if re.search(r"\bкм\b", norm):
        return "км", "km"
    if re.search(r"(^|\W)л($|\W)", norm):
        return "л", "l"
    if "%" in norm:
        return "%", "pct"
    if "мест" in norm or "пассаж" in norm:
        return "мест", "seats"
    return None, None


def extract_base_model(file_name: str) -> str:
    stem = Path(file_name).stem
    return re.split(r"_[Сс]равнение", stem, maxsplit=1)[0].strip()


def number_tokens(text: str | None) -> list[float]:
    raw = clean_text(text)
    if not raw:
        return []
    raw = raw.replace("…", " ").replace("–", "-").replace("—", "-")
    tokens = re.findall(r"\d[\d\s.,]*", raw)
    numbers: list[float] = []
    for token in tokens:
        compact = re.sub(r"\s+", "", token)
        if not compact:
            continue
        if "," in compact and "." not in compact:
            compact = compact.replace(",", ".")
        else:
            compact = compact.replace(",", "")
        try:
            numbers.append(float(compact))
        except ValueError:
            continue
    return numbers


def first_number(text: str | None) -> float | None:
    numbers = number_tokens(text)
    return numbers[0] if len(numbers) == 1 else None


def parse_additive_count(text: str | None) -> int | None:
    raw = clean_text(text)
    if not raw:
        return None
    if any(marker in raw for marker in ("/", "…", "-", ",")):
        return None
    numbers = number_tokens(raw)
    if not numbers:
        return None
    if "+" in raw:
        return int(sum(numbers))
    return int(numbers[0]) if len(numbers) == 1 else None


def parse_price_values(raw: str | None, label_context: str = "") -> tuple[str | None, float | None, float | None, str | None]:
    text = clean_text(raw)
    if not text:
        return None, None, None, None
    numbers = number_tokens(text)
    if not numbers:
        return text, None, None, None
    scale = 1000.0 if "тыс" in f"{normalize_label(label_context)} {normalize_label(text)}" else 1.0
    scaled = [number * scale for number in numbers]
    comment = None
    label_norm = normalize_label(label_context)
    if "без учета скид" in label_norm:
        comment = "без учета скидок и акций"
    elif "по прайсу" in label_norm:
        comment = "по прайсу"
    elif "с ндс" in normalize_label(text):
        comment = "с НДС"
    return text, min(scaled), max(scaled), comment


def parse_option_cell(raw: str | None) -> tuple[str | None, str | None, str | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None, None, None
    low = normalize_label(text)
    normalized = None
    if low == "listed":
        return text, "listed", None, None
    if "недоступ" in low:
        normalized = "unavailable"
    elif low in {"-", "нет"}:
        normalized = "not_present"
    elif low == "+":
        normalized = "included"
    elif "база" in low:
        normalized = "base"
    elif "опция" in low and "?" not in low:
        normalized = "optional"
    elif "??" in low or "?" in low:
        normalized = "unknown"

    price_raw, price_min, _, _ = parse_price_values(text)
    status_raw = text
    if price_min is not None and re.search(r"[A-Za-zА-Яа-я+\-?]", text):
        match = re.search(r"\d[\d\s.,]*", text)
        if match:
            numeric = clean_text(match.group(0))
            price_raw = numeric
            status_raw = clean_text(text.replace(match.group(0), " ").strip(" /")) or status_raw
            if not normalized and status_raw != text:
                normalized = parse_option_cell(status_raw)[1]
    elif price_min is not None and not normalized:
        status_raw = None
    return status_raw, normalized, price_raw, price_min


def join_parts(*parts: str | None) -> str | None:
    items: list[str] = []
    for part in parts:
        cleaned = clean_text(part)
        if not cleaned:
            continue
        if not items:
            items.append(cleaned)
            continue
        last = items[-1]
        last_norm = normalize_label(last)
        cleaned_norm = normalize_label(cleaned)
        if cleaned_norm == last_norm:
            continue
        if cleaned_norm.startswith(last_norm):
            items[-1] = cleaned
            continue
        if last_norm.startswith(cleaned_norm):
            continue
        if last_norm.endswith(f" {cleaned_norm}"):
            continue
        items.append(cleaned)
    return " ".join(items) if items else None


def parse_dimensions(raw: str | None) -> tuple[float | None, float | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None, None
    numbers = number_tokens(text.split("(")[0])
    if len(numbers) == 3:
        return numbers[0], numbers[1], numbers[2]
    return None, None, None


def parse_mass_pair(raw: str | None) -> tuple[float | None, float | None]:
    text = clean_text(raw)
    if not text or "/" not in text:
        return None, None
    parts = [part.strip() for part in text.split("/")]
    if len(parts) < 2:
        return None, None
    return first_number(parts[0]), first_number(parts[1])


def parse_numeric_bounds(raw: str | None, *, scale_small_values: bool = False) -> tuple[float | None, float | None]:
    values = number_tokens(raw)
    if not values:
        return None, None
    if scale_small_values:
        values = [value * 1000.0 if value < 50 else value for value in values]
    return min(values), max(values)


def parse_primary_numeric_bounds(raw: str | None, *, scale_small_values: bool = False) -> tuple[float | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    head = clean_text(text.split("(", 1)[0])
    if head and head != text:
        values = number_tokens(head)
        if values:
            if scale_small_values:
                values = [value * 1000.0 if value < 50 else value for value in values]
            return min(values), max(values)
    return parse_numeric_bounds(text, scale_small_values=scale_small_values)


def parse_slash_pair_bounds(raw: str | None) -> tuple[tuple[float | None, float | None], tuple[float | None, float | None]]:
    text = clean_text(raw)
    if not text or "/" not in text:
        return (None, None), (None, None)
    left, right = [clean_text(part) for part in text.split("/", 1)]
    return parse_numeric_bounds(left), parse_numeric_bounds(right)


def assign_numeric_range(record: dict[str, Any], field: str, lower: float | None, upper: float | None, *, overwrite: bool = False) -> bool:
    if lower is None or upper is None:
        return False
    min_key = f"{field}_min"
    max_key = f"{field}_max"
    if overwrite or record.get(min_key) is None:
        record[min_key] = lower
    if overwrite or record.get(max_key) is None:
        record[max_key] = upper
    if overwrite or record.get(field) is None:
        record[field] = lower
    return True


def assign_numeric_range_from_raw(
    record: dict[str, Any],
    field: str,
    raw: str | None,
    *,
    overwrite: bool = False,
    prefer_existing: bool = False,
    scale_small_values: bool = False,
) -> bool:
    min_key = f"{field}_min"
    max_key = f"{field}_max"
    if prefer_existing and any(record.get(key) is not None for key in (field, min_key, max_key)):
        return False
    lower, upper = parse_numeric_bounds(raw, scale_small_values=scale_small_values)
    return assign_numeric_range(record, field, lower, upper, overwrite=overwrite)


def assign_primary_numeric_range_from_raw(
    record: dict[str, Any],
    field: str,
    raw: str | None,
    *,
    overwrite: bool = False,
    prefer_existing: bool = False,
    scale_small_values: bool = False,
) -> bool:
    min_key = f"{field}_min"
    max_key = f"{field}_max"
    if prefer_existing and any(record.get(key) is not None for key in (field, min_key, max_key)):
        return False
    lower, upper = parse_primary_numeric_bounds(raw, scale_small_values=scale_small_values)
    return assign_numeric_range(record, field, lower, upper, overwrite=overwrite)


def assign_numeric_lower_from_raw(
    record: dict[str, Any],
    field: str,
    raw: str | None,
    *,
    overwrite: bool = False,
    scale_small_values: bool = False,
) -> bool:
    lower, _ = parse_numeric_bounds(raw, scale_small_values=scale_small_values)
    if lower is None:
        return False
    min_key = f"{field}_min"
    max_key = f"{field}_max"
    if overwrite or record.get(min_key) is None:
        record[min_key] = lower
    if overwrite or record.get(max_key) is None:
        record[max_key] = lower
    if overwrite or record.get(field) is None:
        record[field] = lower
    return True


def assign_numeric_upper_from_raw(
    record: dict[str, Any],
    field: str,
    raw: str | None,
    *,
    overwrite: bool = False,
    scale_small_values: bool = False,
) -> bool:
    _, upper = parse_numeric_bounds(raw, scale_small_values=scale_small_values)
    if upper is None:
        return False
    min_key = f"{field}_min"
    max_key = f"{field}_max"
    if overwrite or record.get(max_key) is None:
        record[max_key] = upper
    if overwrite or record.get(min_key) is None:
        record[min_key] = upper
    if overwrite or record.get(field) is None:
        record[field] = record.get(min_key) or upper
    return True


def normalize_bounded_numeric_fields(row: dict[str, Any]) -> None:
    for field in BOUNDED_NUMERIC_FIELDS:
        min_key = f"{field}_min"
        max_key = f"{field}_max"
        lower = row.get(min_key)
        upper = row.get(max_key)
        value = row.get(field)
        if lower is None and value is not None:
            lower = value
        if upper is None and value is not None:
            upper = value
        if lower is None and upper is not None:
            lower = upper
        if upper is None and lower is not None:
            upper = lower
        if lower is not None:
            row[field] = lower
            row[min_key] = lower
            row[max_key] = upper


def parse_power(raw: str | None, label: str) -> tuple[float | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    low = normalize_label(text)
    hp_match = re.search(r"(\d[\d\s.,]*)\s*(?:л\.?\s*с\.?|hp)", low)
    kw_match = re.search(r"(\d[\d\s.,]*)\s*квт", low)
    hp = number_tokens(hp_match.group(1))[0] if hp_match else None
    kw = number_tokens(kw_match.group(1))[0] if kw_match else None
    if hp is None and kw is None:
        single = first_number(text)
        if single is not None:
            if "квт" in normalize_label(label):
                kw = single
            else:
                hp = single
    if hp is None and kw is not None:
        hp = round(kw * 1.35962, 2)
    if kw is None and hp is not None and "квт" in normalize_label(label):
        kw = round(hp / 1.35962, 2)
    return hp, kw


def parse_engine_volume(raw: str | None, label: str) -> float | None:
    value = first_number(raw)
    if value is None:
        return None
    low = f"{normalize_label(label)} {normalize_label(raw)}"
    if "куб" in low or "см3" in low or "cc" in low:
        return round(value / 1000.0, 3)
    return round(value, 3)


def parse_transmission(raw: str | None) -> tuple[str | None, int | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    low = normalize_label(text)
    transmission_type = None
    if "мкп" in low or re.search(r"\bmt\b", low):
        transmission_type = "MT"
    elif "акп" in low or re.search(r"\bat\b", low):
        transmission_type = "AT"
    elif "робот" in low or "amt" in low:
        transmission_type = "AMT"
    elif "вариатор" in low or "cvt" in low:
        transmission_type = "CVT"
    gears = None
    for number in number_tokens(text):
        integer = int(number)
        if 3 <= integer <= 12:
            gears = integer
            break
    return transmission_type, gears


def parse_drive(raw: str | None) -> tuple[str | None, str | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    low = normalize_label(text).replace("х", "x")
    formula_match = re.search(r"(\d)\s*[x*]\s*(\d)", low)
    if formula_match:
        formula = f"{formula_match.group(1)}x{formula_match.group(2)}"
        return formula, formula
    if "перед" in low:
        return "fwd", None
    if "зад" in low:
        return "rwd", None
    if "полн" in low:
        return "awd", None
    return None, None


def parse_warranty(raw: str | None) -> tuple[int | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    parts = [part.strip() for part in text.split("/")]
    months = None
    km = None
    if parts:
        left = parts[0]
        values = number_tokens(left)
        if values:
            months = int(values[0] * 12) if "лет" in normalize_label(left) else int(values[0])
    if len(parts) > 1:
        right = normalize_label(parts[1])
        if "без огр" not in right:
            values = number_tokens(parts[1])
            if values:
                km = values[0] * 1000.0 if "тыс" in right else values[0]
    return months, km


def single_numeric_value(raw: str | None) -> float | None:
    text = clean_text(raw)
    if not text:
        return None
    values = number_tokens(text)
    if len(values) != 1:
        return None
    return values[0]


def primary_numeric_value(raw: str | None) -> float | None:
    text = clean_text(raw)
    if not text:
        return None
    head = clean_text(text.split("(", 1)[0])
    if head and head != text:
        values = number_tokens(head)
        if len(values) == 1:
            return values[0]
    return single_numeric_value(text)


def parse_thousand_km_value(raw: str | None) -> float | None:
    text = clean_text(raw)
    if not text or normalize_label(text) == "без ограничения":
        return None
    value = single_numeric_value(text)
    if value is None:
        return None
    return value * 1000.0


def parse_mixed_engine_volume_l(raw: str | None) -> float | None:
    value = single_numeric_value(raw)
    if value is None:
        return None
    if value >= 10:
        return round(value / 1000.0, 3)
    return round(value, 3)


def parse_interval_km_value(raw: str | None) -> float | None:
    text = clean_text(raw)
    if not text:
        return None
    values = number_tokens(text)
    if not values:
        return None
    return max(values)


def parse_combined_warranty_value(raw: str | None) -> tuple[int | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    norm = normalize_label(text)
    if "/" in text:
        left, right = [part.strip() for part in text.split("/", 1)]
        months = None
        km = None
        left_values = number_tokens(left)
        if left_values:
            months = int(left_values[0] * 12) if "год" in normalize_label(left) else int(left_values[0])
        if "без огр" not in normalize_label(right):
            right_values = number_tokens(right)
            if right_values:
                km = right_values[0] * 1000.0 if right_values[0] < 1000 else right_values[0]
        return months, km
    values = number_tokens(text)
    if not values:
        return None, None
    months = None
    km = None
    if "год" in norm:
        months = int(values[0] * 12)
        if len(values) > 1:
            km = values[1] * 1000.0 if values[1] < 1000 else values[1]
    else:
        months = int(values[0])
        if len(values) > 1:
            km = values[1] * 1000.0 if values[1] < 1000 else values[1]
    return months, km


def parse_minivan_power(raw: str | None) -> tuple[float | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    values = number_tokens(text)
    if len(values) == 1:
        kw = values[0]
        return round(kw * 1.35962, 2), kw
    if len(values) >= 2:
        first = values[0]
        second = values[1]
        if first <= 120 < second:
            return second, first
        if second <= 120 < first:
            return first, second
        return second, first
    return None, None


def has_range_marker(raw: str | None) -> bool:
    text = clean_text(raw)
    if not text:
        return False
    return "…" in text or "..." in text


def parse_number_literal(token: str) -> float | None:
    compact = re.sub(r"\s+", "", token.strip()).replace("−", "-")
    compact = compact.rstrip(".,;/")
    if not compact:
        return None
    if "," in compact and "." not in compact:
        compact = compact.replace(",", ".")
    else:
        compact = compact.replace(",", "")
    try:
        return float(compact)
    except ValueError:
        return None


def strict_number_matches(raw: str | None, *, allow_negative: bool = False) -> list[str]:
    text = clean_text(raw)
    if not text or has_range_marker(text):
        return []
    pattern = r"(?<![\d/])-?\d[\d\s.,]*" if allow_negative else r"(?<![\d/])\d[\d\s.,]*"
    return re.findall(pattern, text)


def strict_single_number(raw: str | None, *, allow_negative: bool = False) -> float | None:
    matches = strict_number_matches(raw, allow_negative=allow_negative)
    if len(matches) != 1:
        return None
    return parse_number_literal(matches[0])


def parse_paz_mass_pair(raw: str | None) -> tuple[float | None, float | None]:
    text = clean_text(raw)
    if not text or "/" not in text:
        return None, None
    left, right = [clean_text(part) for part in text.split("/", 1)]
    return strict_single_number(left), strict_single_number(right)


def parse_paz_dimensions(raw: str | None) -> tuple[float | None, float | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None, None
    head = clean_text(text.split("(", 1)[0])
    if not head or has_range_marker(head):
        return None, None, None
    parts = [clean_text(part) for part in head.split("/")]
    if len(parts) != 3:
        return None, None, None
    numbers = [strict_single_number(part) for part in parts]
    if any(number is None for number in numbers):
        return None, None, None
    return numbers[0], numbers[1], numbers[2]


def parse_paz_seat_count(raw: str | None) -> int | None:
    text = clean_text(raw)
    if not text or has_range_marker(text):
        return None
    matches = strict_number_matches(text)
    if not matches:
        return None
    values = [parse_number_literal(match) for match in matches]
    if any(value is None or int(value) != value for value in values):
        return None
    return int(sum(value for value in values if value is not None))


def parse_paz_fuel_type(raw: str | None) -> str | None:
    text = clean_text(raw)
    if not text or "," not in text:
        return None
    _, fuel_part = text.split(",", 1)
    return clean_text(fuel_part)


def parse_paz_warranty(raw: str | None) -> tuple[int | None, float | None]:
    text = clean_text(raw)
    if not text or "/" not in text:
        return None, None
    left, right = [clean_text(part) for part in text.split("/", 1)]
    months_value = strict_single_number(left)
    if months_value is None:
        return None, None
    left_norm = normalize_label(left)
    if "год" in left_norm or "лет" in left_norm:
        months = int(round(months_value * 12))
    else:
        months = int(round(months_value))
    if "без огранич" in normalize_label(right):
        return months, None
    km_value = strict_single_number(right)
    if km_value is None:
        return months, None
    if "тыс" in normalize_label(right):
        return months, km_value * 1000.0
    return months, km_value


def parse_paz_price(raw: str | None) -> tuple[str | None, float | None, float | None, str | None]:
    text = clean_text(raw)
    if not text:
        return None, None, None, None
    low = normalize_label(text)
    comment_parts = ["по прайсу", "с НДС"]
    if "расчетная цена" in low:
        comment_parts.append("расчетная цена")
    if "не завозится" in low:
        comment_parts.append("не завозится")
    if "снят с производства" in low:
        comment_parts.append("снят с производства")
    matches = strict_number_matches(text)
    if len(matches) != 1:
        return text, None, None, "; ".join(comment_parts) if comment_parts else None
    value = parse_number_literal(matches[0])
    if value is None:
        return text, None, None, "; ".join(comment_parts) if comment_parts else None
    if "млн" in low:
        scaled = value * 1_000_000.0
    else:
        scaled = value * 1000.0
    return text, scaled, scaled, "; ".join(comment_parts)


def parse_paz_option_value(raw: str | None) -> tuple[str | None, str | None, str | None, float | None]:
    text = clean_text(raw)
    if not text:
        return None, None, None, None
    if "/" in text:
        left, right = [clean_text(part) for part in text.split("/", 1)]
        left_number = strict_single_number(left, allow_negative=True)
        right_number = strict_single_number(right, allow_negative=True)
        if left_number is not None and right_number is None:
            return right, parse_option_cell(right)[1], left, left_number
        if right_number is not None and left_number is None:
            return left, parse_option_cell(left)[1], right, right_number
    signed_value = strict_single_number(text, allow_negative=True)
    if signed_value is not None:
        return None, None, text, signed_value
    status_raw, status_norm, price_raw, price_rub = parse_option_cell(text)
    if price_rub is None and status_norm is not None:
        price_raw = None
    return status_raw, status_norm, price_raw, price_rub


def parse_ldt_dimensions(raw: str | None) -> tuple[float | None, float | None, float | None]:
    text = clean_text(raw)
    if not text or "|" in text or has_range_marker(text):
        return None, None, None
    head = clean_text(text.split("(", 1)[0])
    if not head:
        return None, None, None
    parts = [clean_text(part) for part in re.split(r"[хx×]", head)]
    if len(parts) != 3:
        return None, None, None
    numbers = [strict_single_number(part) for part in parts]
    if any(number is None for number in numbers):
        return None, None, None
    return numbers[0], numbers[1], numbers[2]


class GazPricingImporter:
    def __init__(self, source_dir: Path, db_path: Path) -> None:
        self.source_dir = source_dir
        self.db_path = db_path
        self.conn: sqlite3.Connection | None = None
        self.staged_rows: list[dict[str, Any]] = []
        self.warnings: list[str] = []
        self.stats = {
            "workbooks": 0,
            "selected_sheets": 0,
            "staged_rows": 0,
            "merged_rows": 0,
            "warnings": 0,
        }
        self.parsers: dict[str, Parser] = {
            SHEET_SVESY_BAZY: self.parse_sheet_svesy_bazy,
            SHEET_LIST2: self.parse_sheet_list2,
            SHEET_PAZ: self.parse_sheet_paz,
            SHEET_PAZ_EXPANDED: self.parse_sheet_paz_expanded,
            SHEET_BORT_TTX: self.parse_sheet_bort_ttx,
            SHEET_CMF_TTX: self.parse_sheet_cmf_ttx,
            SHEET_LDT_TTX: self.parse_sheet_ldt_ttx,
            SHEET_MINIVEN: self.parse_sheet_miniven,
            SHEET_MINIVEN_2: self.parse_sheet_miniven_2,
            SHEET_BUS_TTX: self.parse_sheet_bus_ttx,
            SHEET_9_10T: self.parse_sheet_9_10t,
            SHEET_BORT_SHORT: self.parse_sheet_bort_short,
            SHEET_CMF_SHORT: self.parse_sheet_cmf_short,
            SHEET_BUS_SHORT: self.parse_sheet_bus_short,
            SHEET_LDT_SHORT: self.parse_sheet_ldt_short,
            SHEET_MINIVEN_SHORT: self.parse_sheet_miniven_short,
            SHEET_COMPASS: self.parse_sheet_compass,
            SHEET_SF5: self.parse_sheet_sf5,
            SHEET_BORT_STVLAD: self.parse_sheet_bort_stvlad,
            SHEET_CMF_STVLAD: self.parse_sheet_cmf_stvlad,
            SHEET_LDT_STVLAD: self.parse_sheet_ldt_stvlad,
            SHEET_TO_BORT: self.parse_sheet_to_bort,
            SHEET_TO_CMF: self.parse_sheet_to_cmf,
            SHEET_COMPLECTATION: self.parse_sheet_complectation,
            SHEET_ADVANTAGES: self.parse_sheet_advantages,
            SHEET_ADVANTAGES_MINIVAN: self.parse_sheet_advantages_minivan,
            SHEET_REVIEWS: self.parse_sheet_reviews,
            SHEET_REVIEWS_LOWER: self.parse_sheet_reviews_lower,
            SHEET_CLIENT_REVIEWS_NN: self.parse_sheet_client_reviews_nn,
            SHEET_CLIENT_REVIEWS_ATLANT: self.parse_sheet_client_reviews_atlant,
            SHEET_NEGATIVE_POSITIVE_ORM: self.parse_sheet_negative_positive_orm,
            SHEET_QUALITY: self.parse_sheet_quality,
            SHEET_DPO_STATS: self.parse_sheet_dpo_stats,
        }

    @property
    def db(self) -> sqlite3.Connection:
        if self.conn is None:
            raise RuntimeError("Database connection is not initialized")
        return self.conn

    def run(self) -> None:
        if self.db_path.exists():
            self.db_path.unlink()
        self.conn = sqlite3.connect(self.db_path)
        self.db.row_factory = sqlite3.Row
        self.create_schema()
        try:
            for source_file, sheet_map in SOURCE_ALLOWLIST.items():
                workbook_path = self.source_dir / source_file
                if not workbook_path.exists():
                    self.warn(f"missing source workbook: {source_file}")
                    continue
                self.stats["workbooks"] += 1
                workbook = load_workbook(workbook_path, read_only=False, data_only=True)
                base_model = extract_base_model(workbook_path.name)
                sheets_by_name = {ws.title: ws for ws in workbook.worksheets}
                for source_sheet, id_rule in sheet_map.items():
                    ws = sheets_by_name.get(source_sheet)
                    if ws is None:
                        self.warn(f"missing source sheet: {source_file}/{source_sheet}")
                        continue
                    parser = self.parsers.get(source_sheet)
                    if parser is None:
                        raise ValueError(f"Unknown allowlisted sheet name: {source_sheet!r} in {workbook_path.name}")
                    self.stats["selected_sheets"] += 1
                    ctx = {
                        "workbook_path": workbook_path,
                        "source_file": workbook_path.name,
                        "source_sheet": source_sheet,
                        "base_model": base_model,
                        "ws": ws,
                        "id_rule": id_rule,
                    }
                    parser(ctx)
                workbook.close()
            for row in self.merge_staged_rows():
                self.insert_row("comparisons_normalized", NORMALIZED_COLUMNS, row)
            self.db.commit()
        except Exception:
            self.db.rollback()
            raise
        finally:
            self.db.close()
            self.conn = None

    def sqlite_type_for_column(self, column: str) -> str:
        base_column = column.removesuffix("_nocase") if column.endswith("_nocase") else column
        if column.endswith("_nocase") or base_column in NORMALIZED_TEXT_COLUMNS or base_column == "id":
            return "TEXT"
        if base_column in INTEGER_FIELDS:
            return "INTEGER"
        if base_column in REAL_FIELDS:
            return "REAL"
        return "TEXT"

    def create_schema(self) -> None:
        column_defs: list[str] = []
        for column in NORMALIZED_COLUMNS:
            not_null = " NOT NULL" if column in {"id", "source_file", "source_sheet", "sheet_type", "base_model"} else ""
            column_defs.append(f"{column} {self.sqlite_type_for_column(column)}{not_null}")
        self.db.executescript(
            f"""
            CREATE TABLE comparisons_normalized (
                {", ".join(column_defs)}
            );

            CREATE INDEX idx_norm_id ON comparisons_normalized(id);
            CREATE INDEX idx_norm_source ON comparisons_normalized(source_file, source_sheet);
            CREATE INDEX idx_norm_base_model_nocase ON comparisons_normalized(base_model_nocase);
            CREATE INDEX idx_norm_comp_full_name_nocase ON comparisons_normalized(comp_full_name_nocase);
            CREATE INDEX idx_norm_comp_brand_nocase ON comparisons_normalized(comp_brand_nocase);
            CREATE INDEX idx_norm_comp_model_nocase ON comparisons_normalized(comp_model_nocase);
            """
        )

    def warn(self, message: str) -> None:
        self.warnings.append(message)
        self.stats["warnings"] = len(self.warnings)
        sys.stderr.write(f"WARNING: {message}\n")

    def merge_staged_rows(self) -> list[dict[str, Any]]:
        merged_by_id: dict[str, dict[str, Any]] = {}
        for row in self.staged_rows:
            row_id = row["id"]
            existing = merged_by_id.get(row_id)
            if existing is None:
                merged_by_id[row_id] = dict(row)
                continue
            for column in NORMALIZED_BASE_COLUMNS:
                incoming = row.get(column)
                if incoming is None:
                    continue
                if existing.get(column) is None:
                    existing[column] = incoming
            existing["comp_brand"], existing["comp_model"], existing["comp_full_name"], existing["id"] = canonicalize_identity_parts(
                existing.get("comp_brand"),
                existing.get("comp_model"),
            )
            existing["base_model"] = existing.get("comp_brand")
            normalize_bounded_numeric_fields(existing)
        merged_rows = [merged_by_id[row_id] for row_id in sorted(merged_by_id)]
        self.stats["merged_rows"] = len(merged_rows)
        return merged_rows

    def require_contains(self, ws: Worksheet, coord: str, expected: str) -> None:
        value = clean_text(ws[coord].value) or ""
        if normalize_label(expected) not in normalize_label(value):
            raise ValueError(f"{ws.title}: expected {coord} to contain {expected!r}, got {value!r}")

    def require_exact(self, ws: Worksheet, coord: str, expected: str) -> None:
        value = clean_text(ws[coord].value)
        if value != expected:
            raise ValueError(f"{ws.title}: expected {coord} to equal {expected!r}, got {value!r}")

    def insert_row(self, table: str, columns: list[str], row: dict[str, Any]) -> None:
        payload: dict[str, Any] = {}
        for column in columns:
            if column.endswith("_nocase"):
                payload[column] = lowercase_text(row.get(column.removesuffix("_nocase")))
            else:
                payload[column] = row.get(column)
        placeholders = ", ".join("?" for _ in columns)
        sql = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})"
        self.db.execute(sql, [payload[column] for column in columns])

    def insert_normalized(self, row: dict[str, Any]) -> None:
        staged_row = {column: row.get(column) for column in NORMALIZED_BASE_COLUMNS}
        normalize_bounded_numeric_fields(staged_row)
        staged_row["comp_brand"], staged_row["comp_model"], staged_row["comp_full_name"], staged_row["id"] = canonicalize_identity_parts(
            staged_row.get("comp_brand"),
            staged_row.get("comp_model"),
        )
        staged_row["base_model"] = staged_row.get("comp_brand")
        if staged_row["id"] is None:
            location = f"{staged_row.get('source_file')}/{staged_row.get('source_sheet')}:{row.get('column_index')}"
            self.warn(f"skipping row without normalized id: {location}")
            return
        self.staged_rows.append(staged_row)
        self.stats["staged_rows"] += 1

    def insert_service(self, row: dict[str, Any]) -> None:
        return

    def insert_option(self, row: dict[str, Any]) -> None:
        return

    def insert_raw(self, row: dict[str, Any]) -> None:
        return

    def append_note(self, record: dict[str, Any], note: str | None) -> None:
        text = clean_text(note)
        if not text:
            return
        current = clean_text(record.get("notes"))
        if current:
            if text not in current:
                record["notes"] = f"{current}; {text}"
        else:
            record["notes"] = text

    def is_real_model_header(self, value: Any) -> bool:
        text = clean_text(value)
        if not text:
            return False
        norm = normalize_label(text)
        if norm.startswith("ср.знач"):
            return False
        if "плюсы" in norm:
            return False
        if "примечание" in norm:
            return False
        return True

    def new_model_record(
        self,
        ctx: dict[str, Any],
        sheet_type: str,
        column_index: int,
        comp_full_name: str | None,
        comp_brand: str | None = None,
        comp_model: str | None = None,
        vehicle_type: str | None = None,
        body_type: str | None = None,
    ) -> dict[str, Any]:
        identity_brand, identity_model, full_name, row_id = canonicalize_identity_parts(
            comp_brand,
            comp_model if comp_model is not None else comp_full_name,
        )
        row = {column: None for column in NORMALIZED_BASE_COLUMNS}
        row.update(
            {
                "id": row_id,
                "source_file": ctx["source_file"],
                "source_sheet": ctx["source_sheet"],
                "sheet_type": sheet_type,
                "base_model": identity_brand,
                "comp_full_name": full_name,
                "comp_brand": identity_brand,
                "comp_model": identity_model,
                "vehicle_type": vehicle_type,
                "body_type": body_type,
                "column_index": column_index,
            }
        )
        return row

    def header_parts(self, ws: Worksheet, column_index: int, rows: list[int] | None) -> list[str]:
        if not rows:
            return []
        parts: list[str] = []
        for row_index in rows:
            value = clean_text(ws.cell(row_index, column_index).value)
            if value:
                parts.append(value)
        return parts

    def models_from_brand_model_rows(
        self,
        ctx: dict[str, Any],
        brand_row: int,
        model_row: int,
        *,
        start_col: int = 2,
        allowed_cols: list[int] | None = None,
        prefix_rows: list[int] | None = None,
        sheet_type: str,
        vehicle_type: str | None,
        body_type: str | None,
    ) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        columns = allowed_cols or list(range(start_col, ws.max_column + 1))
        for column_index in columns:
            brand = clean_text(ws.cell(brand_row, column_index).value)
            model = clean_text(ws.cell(model_row, column_index).value)
            full_name = join_parts(brand, model)
            if not self.is_real_model_header(full_name):
                continue
            records.append(
                self.new_model_record(
                    ctx,
                    sheet_type,
                    column_index,
                    full_name,
                    comp_brand=brand,
                    comp_model=model,
                    vehicle_type=vehicle_type,
                    body_type=body_type,
                )
            )
        return records

    def models_from_full_name_row(
        self,
        ctx: dict[str, Any],
        header_row: int,
        *,
        start_col: int = 2,
        allowed_cols: list[int] | None = None,
        prefix_rows: list[int] | None = None,
        sheet_type: str,
        vehicle_type: str | None,
        body_type: str | None,
    ) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        columns = allowed_cols or list(range(start_col, ws.max_column + 1))
        for column_index in columns:
            header_value = clean_text(ws.cell(header_row, column_index).value)
            full_name = join_parts(*self.header_parts(ws, column_index, prefix_rows), header_value)
            if not self.is_real_model_header(full_name):
                continue
            records.append(
                self.new_model_record(
                    ctx,
                    sheet_type,
                    column_index,
                    full_name,
                    comp_brand=header_value,
                    vehicle_type=vehicle_type,
                    body_type=body_type,
                )
            )
        return records

    def models_from_combined_name_rows(
        self,
        ctx: dict[str, Any],
        name_row: int,
        detail_row: int,
        *,
        start_col: int = 2,
        allowed_cols: list[int] | None = None,
        prefix_rows: list[int] | None = None,
        sheet_type: str,
        vehicle_type: str | None,
        body_type: str | None,
    ) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        columns = allowed_cols or list(range(start_col, ws.max_column + 1))
        for column_index in columns:
            name = clean_text(ws.cell(name_row, column_index).value)
            detail = clean_text(ws.cell(detail_row, column_index).value)
            full_name = join_parts(*self.header_parts(ws, column_index, prefix_rows), name, detail)
            if not self.is_real_model_header(full_name):
                continue
            records.append(
                self.new_model_record(
                    ctx,
                    sheet_type,
                    column_index,
                    full_name,
                    comp_brand=name,
                    vehicle_type=vehicle_type,
                    body_type=body_type,
                )
            )
        return records

    def build_minivan_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in [2, 3, 4, 5, 6, 7]:
            name = clean_text(ws.cell(2, column_index).value)
            detail = clean_text(ws.cell(3, column_index).value)
            if not name or not detail:
                raise ValueError(f"{ws.title}: expected model header in {get_column_letter(column_index)}2/{get_column_letter(column_index)}3")
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    join_parts(name, detail) or name,
                    comp_brand=name,
                    comp_model=detail,
                    vehicle_type="minivan",
                    body_type="minivan",
                )
            )
        return records

    def build_paz_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        expected_headers = {
            3: "Вектор NEXT 7.6",
            4: "CITYMAX 8",
            5: "ПАЗ 4234",
            6: "Промтех Руслайнер",
            8: "Вектор NEXT 8.8",
            9: "КАМАZ-4280-F5 (Vega)",
            10: "SIMAZ 2258-538",
            11: "SIMAZ 2258-538",
            13: "МАЗ-206",
            14: "KAMAZ-4290-30-5M",
            15: "ЛиАЗ-4292",
            16: "Volgabus 4298",
            17: "CITYMAX 9",
            18: "Yutong ZK 6890 HGQ",
            19: "SAZ LE60",
            20: "Yutong ZK6852",
        }
        brand_by_column = {
            3: "Вектор NEXT",
            4: "CITYMAX",
            5: "ПАЗ",
            6: "Промтех",
            8: "Вектор NEXT",
            9: "КАМАZ",
            10: "SIMAZ",
            11: "SIMAZ",
            13: "МАЗ",
            14: "KAMAZ",
            15: "ЛиАЗ",
            16: "Volgabus",
            17: "CITYMAX",
            18: "Yutong",
            19: "SAZ",
            20: "Yutong",
        }
        records: list[dict[str, Any]] = []
        for column_index, expected_name in expected_headers.items():
            actual_name = clean_text(ws.cell(2, column_index).value)
            if actual_name != expected_name:
                raise ValueError(
                    f"{ws.title}: expected {get_column_letter(column_index)}2 = {expected_name!r}, got {actual_name!r}"
                )
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    actual_name,
                    comp_brand=actual_name,
                    comp_model=actual_name,
                    vehicle_type="bus",
                    body_type="bus",
                )
            )
        return records

    def apply_paz_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None, float | None]:
        label_text = clean_text(label)
        mapped = False
        value_num = None
        parse_comment = None

        if label_text == "Старт производства":
            self.append_note(record, f"{label_text}: {raw_value}")
            value_num = strict_single_number(raw_value)
        elif label_text == "Класс":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Колесная формула":
            if raw_value == "4x2":
                record["wheel_formula"] = raw_value
                mapped = True
        elif label_text == "Полная / снаряженная массы, кг":
            gross_bounds, curb_bounds = parse_slash_pair_bounds(raw_value)
            if assign_numeric_range(record, "gross_weight_kg", *gross_bounds):
                mapped = True
                value_num = record.get("gross_weight_kg")
            if assign_numeric_range(record, "curb_weight_kg", *curb_bounds):
                mapped = True
                if value_num is None:
                    value_num = record.get("curb_weight_kg")
        elif label_text == "Кол-во служебных дверей":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Пассажировместимость общая":
            value = strict_single_number(raw_value)
            if value is not None:
                record["passenger_capacity"] = int(value)
                mapped = True
                value_num = value
        elif label_text == "Мест для сидения":
            value = parse_paz_seat_count(raw_value)
            if value is not None:
                record["seat_count"] = value
                mapped = True
                value_num = float(value)
        elif label_text == "Уровень пола":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Габариты ДxШxВ, мм":
            length, width, height = parse_paz_dimensions(raw_value)
            if length is not None:
                record["length_mm"] = length
                mapped = True
            if width is not None:
                record["width_mm"] = width
                mapped = True
            if height is not None:
                record["height_mm"] = height
                mapped = True
        elif label_text == "Колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Дорожный просвет, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Двигатель":
            fuel_type = parse_paz_fuel_type(raw_value)
            if fuel_type is not None:
                record["engine_fuel_type"] = fuel_type
                mapped = True
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Мощность двигателя, л.с.":
            value = strict_single_number(raw_value)
            if value is not None:
                record["engine_power_hp"] = value
                mapped = True
                value_num = value
        elif label_text == "КПП":
            transmission_map = {
                "4АТ": ("AT", 4),
                "5АТ": ("AT", 5),
                "5МТ": ("MT", 5),
                "6АТ": ("AT", 6),
                "6МТ": ("MT", 6),
                "6MT": ("MT", 6),
            }
            transmission = transmission_map.get(raw_value)
            if transmission is not None:
                record["transmission_type"] = transmission[0]
                record["transmission_gears"] = transmission[1]
                mapped = True
                value_num = float(transmission[1])
        elif label_text == "Емкость топливного бака, л":
            value = strict_single_number(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
                value_num = value
        elif label_text == "Гарантия":
            months, km = parse_paz_warranty(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
        elif label_text == "Цена по прайсу, тыс. руб. с НДС":
            _, price_min, price_max, price_comment = parse_paz_price(raw_value)
            if price_min is not None and price_max is not None:
                record["price_rub_min"] = price_min
                record["price_rub_max"] = price_max
                record["price_comment"] = price_comment
                mapped = True
                value_num = price_min
            else:
                parse_comment = price_comment
            if re.search(r"[A-Za-zА-Яа-я]", raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if strict_number_matches(raw_value, allow_negative=True) else "text")
        return status, parse_comment, value_num

    def insert_paz_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        option_group = clean_text(ws["B29"].value)
        if option_group != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected B29 = 'Комплектации и опции', got {option_group!r}")
        records_by_column = {record["column_index"]: record for record in records}
        for row_index in range(30, 80):
            option_name = clean_text(ws.cell(row_index, 2).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in B{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                status_raw, status_norm, price_raw, price_rub = parse_paz_option_value(raw_value)
                if status_raw is None and price_raw is not None:
                    status_raw = "опция"
                    status_norm = "optional"
                self.insert_option(
                    {
                        "source_file": ctx["source_file"],
                        "source_sheet": ctx["source_sheet"],
                        "base_model": ctx["base_model"],
                        "comp_full_name": record.get("comp_full_name"),
                        "comp_brand": record.get("comp_brand"),
                        "option_group": option_group,
                        "option_name": option_name,
                        "option_status_raw": status_raw,
                        "option_status_norm": status_norm,
                        "option_price_raw": price_raw,
                        "option_price_rub": price_rub,
                        "notes": None,
                        "row_order": row_index,
                        "column_order": column_index,
                    }
                )
                raw_joined = clean_text(" / ".join(part for part in [price_raw, status_raw] if clean_text(part))) or raw_value
                parse_status = "parsed" if status_norm or price_rub is not None else ("partial" if strict_number_matches(raw_joined, allow_negative=True) else "text")
                option_raw = self.raw_row(
                    ctx,
                    sheet_type="options",
                    record_scope="option",
                    param_name_raw=option_name,
                    value_raw=raw_joined,
                    row_order=row_index,
                    column_order=column_index,
                    comp_record=record,
                    parse_status=parse_status,
                )
                option_raw["value_num"] = price_rub
                self.insert_raw(option_raw)

    def build_paz_expanded_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        expected_headers = {
            3: "Вектор NEXT 7.1",
            4: "Вектор NEXT 7.6",
            5: "CITYMAX 8",
            6: "ПАЗ 4234",
            7: "Промтех Руслайнер",
            9: "Вектор NEXT 8.8",
            10: "Вектор NEXT 8.8 (CNG)",
            11: "КАМАZ-4280-F5 (Vega)",
            12: "АВИОР",
            13: "SIMAZ 2258-538 (шасси Isuzu)",
            14: "SIMAZ 2258-538 (шасси DongFeng)",
            16: "МАЗ-206047",
            17: "МАЗ-206947 CNG",
            18: "KAMAZ-4290-30-5M",
            19: "KAMAZ-4290-30-5N CNG",
            20: "ЛиАЗ-429260",
            21: "ЛиАЗ-4292667 CNG",
            22: "SIMAZ 4282 CNG",
            23: "Volgabus 4298",
            24: "CITYMAX 9",
            25: "CITYMAX 9 CNG",
            26: "Yutong ZK 6890 HGQ",
            27: "SAZ LE60",
            28: "Yutong ZK6852",
        }
        brand_by_column = {
            3: "Вектор NEXT",
            4: "Вектор NEXT",
            5: "CITYMAX",
            6: "ПАЗ",
            7: "Промтех",
            9: "Вектор NEXT",
            10: "Вектор NEXT",
            11: "КАМАZ",
            12: "АВИОР",
            13: "SIMAZ",
            14: "SIMAZ",
            16: "МАЗ",
            17: "МАЗ",
            18: "KAMAZ",
            19: "KAMAZ",
            20: "ЛиАЗ",
            21: "ЛиАЗ",
            22: "SIMAZ",
            23: "Volgabus",
            24: "CITYMAX",
            25: "CITYMAX",
            26: "Yutong",
            27: "SAZ",
            28: "Yutong",
        }
        records: list[dict[str, Any]] = []
        for column_index, expected_name in expected_headers.items():
            actual_name = clean_text(ws.cell(2, column_index).value)
            if actual_name != expected_name:
                raise ValueError(
                    f"{ws.title}: expected {get_column_letter(column_index)}2 = {expected_name!r}, got {actual_name!r}"
                )
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    actual_name,
                    comp_brand=actual_name,
                    comp_model=actual_name,
                    vehicle_type="bus",
                    body_type="bus",
                )
            )
        return records

    def apply_paz_expanded_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None, float | None]:
        label_text = clean_text(label)
        mapped = False
        value_num = None
        parse_comment = None

        if label_text == "Старт производства":
            self.append_note(record, f"{label_text}: {raw_value}")
            value_num = strict_single_number(raw_value)
        elif label_text == "Класс":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Колесная формула":
            if raw_value == "4x2":
                record["wheel_formula"] = raw_value
                mapped = True
        elif label_text == "Полная / снаряженная массы, кг":
            gross_bounds, curb_bounds = parse_slash_pair_bounds(raw_value)
            if assign_numeric_range(record, "gross_weight_kg", *gross_bounds):
                mapped = True
                value_num = record.get("gross_weight_kg")
            if assign_numeric_range(record, "curb_weight_kg", *curb_bounds):
                mapped = True
                if value_num is None:
                    value_num = record.get("curb_weight_kg")
        elif label_text == "Кол-во служебных дверей":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Пассажировместимость общая":
            value = strict_single_number(raw_value)
            if value is not None:
                record["passenger_capacity"] = int(value)
                mapped = True
                value_num = value
        elif label_text == "Мест для сидения":
            value = parse_paz_seat_count(raw_value)
            if value is not None:
                record["seat_count"] = value
                mapped = True
                value_num = float(value)
        elif label_text == "Уровень пола":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Габариты ДxШxВ, мм":
            length, width, height = parse_paz_dimensions(raw_value)
            if length is not None:
                record["length_mm"] = length
                mapped = True
            if width is not None:
                record["width_mm"] = width
                mapped = True
            if height is not None:
                record["height_mm"] = height
                mapped = True
        elif label_text == "Колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Дорожный просвет, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Двигатель":
            fuel_type = parse_paz_fuel_type(raw_value)
            if fuel_type is not None:
                record["engine_fuel_type"] = fuel_type
                mapped = True
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Объем двигателя":
            value = strict_single_number(raw_value)
            if value is not None:
                record["engine_volume_l"] = round(value, 3)
                mapped = True
                value_num = value
        elif label_text == "Мощность двигателя, л.с.":
            value = strict_single_number(raw_value)
            if value is not None:
                record["engine_power_hp"] = value
                mapped = True
                value_num = value
        elif label_text == "КПП":
            transmission_map = {
                "6MT": ("MT", 6),
                "6MT (Fast Gear 6J76Т)": ("MT", 6),
                "6MT (ISUZU MYY6S)": ("MT", 6),
                "6MT (Fast Gear)": ("MT", 6),
                "6АТ (Fast Gear F6A145RB)": ("AT", 6),
                "6АТ": ("AT", 6),
                "5АТ": ("AT", 5),
                "АКПП Voith 854.6S 4ст.": ("AT", 4),
            }
            transmission = transmission_map.get(raw_value)
            if transmission is not None:
                record["transmission_type"] = transmission[0]
                record["transmission_gears"] = transmission[1]
                mapped = True
                value_num = float(transmission[1])
        elif label_text == "Емкость топливного бака, л":
            value = strict_single_number(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
                value_num = value
        elif label_text == "Гарантия":
            months, km = parse_paz_warranty(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
        elif label_text == "Межсервисный интервал":
            value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
                value_num = value
        elif label_text == "Цена по прайсу, тыс. руб. с НДС":
            _, price_min, price_max, price_comment = parse_paz_price(raw_value)
            if price_min is not None and price_max is not None:
                record["price_rub_min"] = price_min
                record["price_rub_max"] = price_max
                record["price_comment"] = price_comment
                mapped = True
                value_num = price_min
            else:
                parse_comment = price_comment
            if re.search(r"[A-Za-zА-Яа-я]", raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text in {
            "Нагрузка на переднюю/заднюю ось",
            "Объём багажника, м3",
            "Высота салона",
            "Радиус разворота",
            "Крутящий момент, Н*м",
            "Максимальная скорость, км/ч",
            "Экология",
            "Расход топлива на л/м3/100 км",
            "Емкость газового баллона, м³",
            "Шины",
            "Передняя подвеска",
            "Задняя подвеска",
            "Тормозная системв",
            "Тормозные механизмы передние / задние",
            "Тормозные механизмы  передние / задние",
            "Вентиляция",
            "Система отопления",
            "Гарантия на кузов",
        }:
            if label_text in {"Экология"}:
                self.append_note(record, f"{label_text}: {raw_value}")
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if strict_number_matches(raw_value, allow_negative=True) else "text")
        return status, parse_comment, value_num

    def insert_paz_expanded_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        option_group = clean_text(ws["B41"].value)
        if option_group != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected B41 = 'Комплектации и опции', got {option_group!r}")
        records_by_column = {record["column_index"]: record for record in records}
        for row_index in range(42, 92):
            option_name = clean_text(ws.cell(row_index, 2).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in B{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                status_raw, status_norm, price_raw, price_rub = parse_paz_option_value(raw_value)
                if status_raw is None and price_raw is not None:
                    status_raw = "опция"
                    status_norm = "optional"
                self.insert_option(
                    {
                        "source_file": ctx["source_file"],
                        "source_sheet": ctx["source_sheet"],
                        "base_model": ctx["base_model"],
                        "comp_full_name": record.get("comp_full_name"),
                        "comp_brand": record.get("comp_brand"),
                        "option_group": option_group,
                        "option_name": option_name,
                        "option_status_raw": status_raw,
                        "option_status_norm": status_norm,
                        "option_price_raw": price_raw,
                        "option_price_rub": price_rub,
                        "notes": None,
                        "row_order": row_index,
                        "column_order": column_index,
                    }
                )
                raw_joined = clean_text(" / ".join(part for part in [price_raw, status_raw] if clean_text(part))) or raw_value
                parse_status = "parsed" if status_norm or price_rub is not None else ("partial" if strict_number_matches(raw_joined, allow_negative=True) else "text")
                option_raw = self.raw_row(
                    ctx,
                    sheet_type="options",
                    record_scope="option",
                    param_name_raw=option_name,
                    value_raw=raw_joined,
                    row_order=row_index,
                    column_order=column_index,
                    comp_record=record,
                    parse_status=parse_status,
                )
                option_raw["value_num"] = price_rub
                self.insert_raw(option_raw)

    def build_ldt_ttx_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        expected_names = {
            2: "Валдай NEXT",
            3: "Валдай 8",
            4: "Газон Next 8,7 т",
            5: "САДКО NEXT",
            6: "САДКО 9",
            7: "Isuzu ELF 8.0",
            8: "Foton S 85",
            9: "DonFeng C80",
            10: "DonFeng Z80",
            11: "JAC N90",
            12: 'Камаз "Компас-9"',
            13: "Газон Next 10 т",
        }
        records: list[dict[str, Any]] = []
        for column_index, expected_brand in expected_names.items():
            brand = clean_text(ws.cell(1, column_index).value)
            model = clean_text(ws.cell(2, column_index).value)
            if brand != expected_brand:
                raise ValueError(
                    f"{ws.title}: expected {get_column_letter(column_index)}1 = {expected_brand!r}, got {brand!r}"
                )
            full_name = join_parts(brand, model) or brand
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    full_name,
                    comp_brand=brand,
                    comp_model=model,
                    vehicle_type="truck",
                    body_type="chassis",
                )
            )
        return records

    def get_bus_ttx_layout(self, ws: Worksheet) -> dict[str, Any]:
        prefixes = {
            2: "L2",
            3: "L3",
            4: "L3",
            5: "L3",
            7: "L3",
            8: "L3",
            9: "L4",
        }
        optional_prefix = {6: "L3"}
        if clean_text(ws["F2"].value) == "SF5":
            prefixes.update(optional_prefix)
            model_columns = [2, 3, 4, 5, 6, 8, 9]
            tail_mode = "listed_dual"
            if clean_text(ws["F3"].value) != "MCA-V1FB":
                raise ValueError(f"{ws.title}: expected F3 = 'MCA-V1FB', got {clean_text(ws['F3'].value)!r}")
            if ws["K120"].value != 250000:
                raise ValueError(f"{ws.title}: expected K120 = 250000, got {ws['K120'].value!r}")
        else:
            model_columns = [2, 3, 4, 5, 8, 9]
            tail_mode = "atlant_price_list"
            if clean_text(ws["F2"].value) is not None:
                raise ValueError(f"{ws.title}: expected F2 to be empty, got {clean_text(ws['F2'].value)!r}")
            if clean_text(ws["F120"].value) != "Примечание":
                raise ValueError(f"{ws.title}: expected F120 = 'Примечание', got {clean_text(ws['F120'].value)!r}")
        for column_index, prefix in prefixes.items():
            actual_prefix = clean_text(ws.cell(1, column_index).value)
            if actual_prefix != prefix:
                raise ValueError(
                    f"{ws.title}: expected {get_column_letter(column_index)}1 = {prefix!r}, got {actual_prefix!r}"
                )
        for coord in ("G2", "G3", "J1", "J2", "J3", "K1", "K2", "K3"):
            if clean_text(ws[coord].value) is not None:
                raise ValueError(f"{ws.title}: expected {coord} to be empty, got {clean_text(ws[coord].value)!r}")
        return {
            "model_columns": model_columns,
            "tail_mode": tail_mode,
        }

    def build_bus_ttx_records(self, ctx: dict[str, Any], model_columns: list[int]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in model_columns:
            brand = clean_text(ws.cell(2, column_index).value)
            model = clean_text(ws.cell(3, column_index).value)
            if not brand or not model:
                raise ValueError(
                    f"{ws.title}: expected complete header in {get_column_letter(column_index)}2:{get_column_letter(column_index)}3"
                )
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    join_parts(brand, model) or brand,
                    comp_brand=brand,
                    comp_model=model,
                    vehicle_type="bus",
                    body_type="bus",
                )
            )
        return records

    def parse_bus_power_exact(self, raw_value: str) -> tuple[float | None, float | None]:
        match = re.search(r"([\d\s.,]+)\s*\(([\d\s.,]+)\)", raw_value)
        if not match:
            return None, None
        kw = parse_number_literal(match.group(1))
        hp = parse_number_literal(match.group(2))
        return hp, kw

    def apply_bus_ttx_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None, float | None]:
        label_text = clean_text(label)
        mapped = False
        value_num = None

        if label_text == "Класс автобуса":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Кол-во мест для сидения (включая водителя), чел.":
            value = strict_single_number(raw_value)
            if value is not None:
                record["seat_count"] = int(value)
                mapped = True
                value_num = value
        elif label_text == "Пассажировместимость (без учета водителя), чел.":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["passenger_capacity"] = value
                mapped = True
                value_num = float(value)
        elif label_text == "Полная масса":
            if assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value):
                mapped = True
                value_num = record.get("gross_weight_kg")
        elif label_text == "Снаряженная масса min, кг":
            if assign_numeric_lower_from_raw(record, "curb_weight_kg", raw_value):
                mapped = True
                value_num = record.get("curb_weight_kg_min")
        elif label_text == "Снаряженная масса max, кг":
            if assign_numeric_upper_from_raw(record, "curb_weight_kg", raw_value):
                mapped = True
                value_num = record.get("curb_weight_kg_max")
        elif label_text == "Грузоподьемность min, кг":
            if assign_numeric_lower_from_raw(record, "payload_kg", raw_value):
                mapped = True
                value_num = record.get("payload_kg_min")
        elif label_text == "Грузоподьемность max, кг":
            if assign_numeric_upper_from_raw(record, "payload_kg", raw_value):
                mapped = True
                value_num = record.get("payload_kg_max")
        elif label_text == "Длина, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["length_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "без задн.подножки, мм":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
                value_num = value
        elif label_text in {"передний свес, мм", "задний свес, мм"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Ширина, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["width_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Высота (по люку), мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["height_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Высота с доп. кондиционером, мм":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text in {
            "Ширина проема боковой двери, мм",
            "Высота проема боковой двери, мм",
            "Высота проема боковой двери от ступеньки, мм",
            "Высота проема задних дверей, мм",
            "Ширина проема задних дверей, мм",
            "Высота потолка в пассажирском салоне, мм",
            "Ширина между входными поручнями",
            "Мин. радиус разворота, м",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Дорожный просвет, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливо":
            record["engine_fuel_type"] = raw_value
            mapped = True
        elif label_text == "Объем двигателя, куб.см":
            value = parse_mixed_engine_volume_l(raw_value)
            if value is not None:
                record["engine_volume_l"] = value
                mapped = True
                value_num = value
        elif label_text == "Мощность двигателя, кВт (л.с.)":
            hp, kw = self.parse_bus_power_exact(raw_value)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
                value_num = kw
        elif label_text in {"при об. мин.", "Крутящий момент", "Экологический класс"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "КПП":
            record["transmission_type"] = raw_value
            mapped = True
            gears_map = {
                "5МКПП": 5,
                "6МКПП": 6,
            }
            gears = gears_map.get(raw_value)
            if gears is not None:
                record["transmission_gears"] = gears
                value_num = float(gears)
        elif label_text in {
            "Задний блокируемый дифференциал",
            "Тормозные механизмы",
            "Электронная тормозная система",
            "Стояночный тормоз",
            "Блокируемый дифференциал",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливный бак, л":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
                value_num = value
        elif label_text == "Колесная формула":
            normalized = raw_value.replace("х", "x")
            formula_match = re.match(r"(\d)x(\d)\s*\(([^)]+)\)", normalized)
            if formula_match:
                record["wheel_formula"] = f"{formula_match.group(1)}x{formula_match.group(2)}"
                record["drive_type"] = clean_text(formula_match.group(3))
                mapped = True
        elif label_text in {"Ошиновка", "Размер шин передних", "задних", "Диски"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
        elif label_text == "Гарантия от сквозной коррозии":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
                value_num = value
        elif label_text == "Цена":
            if self.set_price(record, raw_value, label_text):
                mapped = True
                value_num = strict_single_number(raw_value)
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if strict_number_matches(raw_value, allow_negative=True) else "text")
        return status, None, value_num

    def insert_bus_ttx_main_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A51"].value) != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected A51 = 'Комплектации и опции'")
        group_rows = {
            52: "Функциональность",
            65: "Безопасность",
            73: "Экстерьер",
            79: "Кабина",
            98: "Пассажирский салон",
        }
        for row_index, group_name in group_rows.items():
            if clean_text(ws.cell(row_index, 1).value) != group_name:
                raise ValueError(f"{ws.title}: expected A{row_index} = {group_name!r}")
        groups_by_range = [
            (53, 64, "Функциональность"),
            (66, 72, "Безопасность"),
            (74, 78, "Экстерьер"),
            (80, 97, "Кабина"),
            (99, 115, "Пассажирский салон"),
        ]
        records_by_column = {record["column_index"]: record for record in records}
        for start_row, end_row, group_name in groups_by_range:
            for row_index in range(start_row, end_row + 1):
                option_name = clean_text(ws.cell(row_index, 1).value)
                if not option_name:
                    raise ValueError(f"{ws.title}: expected option name in A{row_index}")
                for column_index, record in records_by_column.items():
                    raw_value = clean_text(ws.cell(row_index, column_index).value)
                    if raw_value is None:
                        continue
                    self.insert_option_row(
                        ctx,
                        record=record,
                        option_group=group_name,
                        option_name=option_name,
                        row_order=row_index,
                        status_override=raw_value,
                        status_only=True,
                    )

    def insert_bus_ttx_tail_option_rows(self, ctx: dict[str, Any], records_by_column: dict[int, dict[str, Any]], tail_mode: str) -> None:
        ws = ctx["ws"]
        if tail_mode == "atlant_price_list":
            atlant_record = records_by_column[5]
            for row_index in range(121, 144):
                option_name = clean_text(ws.cell(row_index, 5).value)
                option_price = clean_text(ws.cell(row_index, 6).value)
                if not option_name or option_price is None:
                    continue
                self.insert_option_row(
                    ctx,
                    record=atlant_record,
                    option_group="Доп. опции",
                    option_name=option_name,
                    row_order=row_index,
                    price_override=option_price,
                    status_override="опция",
                )
            return
        if tail_mode == "listed_dual":
            atlant_record = records_by_column[5]
            sf5_record = records_by_column[6]
            for row_index in range(121, 144):
                atlant_option = clean_text(ws.cell(row_index, 5).value)
                sf5_option = clean_text(ws.cell(row_index, 6).value)
                if atlant_option and sf5_option and atlant_option != sf5_option:
                    raise ValueError(f"{ws.title}: expected E{row_index} and F{row_index} to match, got {atlant_option!r} / {sf5_option!r}")
                if atlant_option:
                    self.insert_option_row(
                        ctx,
                        record=atlant_record,
                        option_group="Доп. опции",
                        option_name=atlant_option,
                        row_order=row_index,
                        status_override="listed",
                    )
                if sf5_option:
                    self.insert_option_row(
                        ctx,
                        record=sf5_record,
                        option_group="Доп. опции",
                        option_name=sf5_option,
                        row_order=row_index,
                        status_override="listed",
                    )
            return
        raise ValueError(f"{ws.title}: unsupported tail mode {tail_mode!r}")

    def apply_ldt_ttx_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None, float | None]:
        label_text = clean_text(label)
        mapped = False
        value_num = None

        if label_text == "Колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
                value_num = float(value)
        elif label_text == 'Модификация "спальник" или "двухрядка"':
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Полная масса, кг":
            if assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value):
                mapped = True
                value_num = record.get("gross_weight_kg")
        elif label_text == "Снаряженная масса шасси, кг":
            if assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value):
                mapped = True
                value_num = record.get("curb_weight_kg")
        elif label_text == "Грузоподъемность шасси, кг":
            if assign_numeric_range_from_raw(record, "payload_kg", raw_value):
                mapped = True
                value_num = record.get("payload_kg")
        elif label_text == "Габаритные размеры шасси, мм (ДхШхВ)":
            length, width, height = parse_ldt_dimensions(raw_value)
            if length is not None:
                record["length_mm"] = length
                mapped = True
            if width is not None:
                record["width_mm"] = width
                mapped = True
            if height is not None:
                record["height_mm"] = height
                mapped = True
        elif label_text == "Дорожный просвет, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
                value_num = value
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Экологический класс":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Объем двигателя, л":
            value = strict_single_number(raw_value)
            if value is not None:
                record["engine_volume_l"] = round(value, 3)
                mapped = True
                value_num = value
        elif label_text == "Max мощность, кВт (л.с.)":
            hp = None
            kw = None
            match = re.search(r"([\d\s.,]+)\s*\(([\d\s.,]+)\)", raw_value)
            if match:
                kw = parse_number_literal(match.group(1))
                hp = parse_number_literal(match.group(2))
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
            if kw is not None:
                value_num = kw
        elif label_text == "Коробка передач":
            transmission_map = {
                "6МКПП": ("MT", 6),
                "5МКПП": ("MT", 5),
                "8МКПП": ("MT", 8),
            }
            transmission = transmission_map.get(raw_value)
            if transmission is not None:
                record["transmission_type"] = transmission[0]
                record["transmission_gears"] = transmission[1]
                mapped = True
                value_num = float(transmission[1])
        elif label_text == "Топливный бак, л":
            value = strict_single_number(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
                value_num = value
        elif label_text == "Гарантия, месяцы / пробег тыс. км":
            text = clean_text(raw_value)
            if text and "/" in text:
                left, right = [clean_text(part) for part in text.split("/", 1)]
                months = strict_single_number(left)
                if months is not None:
                    record["warranty_months"] = int(months)
                    mapped = True
                if right and "без огр" not in normalize_label(right):
                    km = strict_single_number(right)
                    if km is not None:
                        record["warranty_km"] = km * 1000.0
                        mapped = True
        elif label_text == "Межсервисный интервал, км":
            value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
                value_num = value
        elif label_text == "Средний расход топлива фактический, л/100 км":
            value = strict_single_number(raw_value)
            if value is not None:
                record["fuel_consumption_l100km"] = value
                mapped = True
                value_num = value
        elif label_text == "Цена а/м":
            if self.set_price(record, raw_value, label_text):
                mapped = True
                value_num = strict_single_number(raw_value)
        elif label_text == "Затраты на 1 км пробега":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ownership_cost_rub_km"] = value
                mapped = True
                value_num = value
        elif label_text in {
            "Мин. радиус разворота, м",
            "Колея передних колес, мм",
            "Погрузочная высота шасси, мм",
            "при об/мин",
            "Max крутящий момент, Н*м",
            "Система нейтрализации",
            "Аккумулятор",
            "Генератор",
            "Напряжение бортовой сети",
            "Тормозная система",
            "Тормозные механизмы",
            "Электронная тормозная система",
            "Стояночный тормоз",
            "Горный тормоз",
            "От сквозной коррозии, лет",
            "Первое ТО",
            "Эксклюзивные опции",
            "Конструктивные особенности",
            "Страховка+налоги+лизинг за период",
            "Cтоимость ТО за период",
            "Стоимость ТО и б/и запчастей",
        }:
            pass
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if strict_number_matches(raw_value, allow_negative=True) else "text")
        return status, None, value_num

    def insert_ldt_ttx_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A32"].value) != "Комплектация и опции":
            raise ValueError(f"{ws.title}: expected A32 = 'Комплектация и опции'")
        records_by_column = {record["column_index"]: record for record in records}
        for row_index in range(33, 72):
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group="Комплектация и опции",
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def raw_row(
        self,
        ctx: dict[str, Any],
        *,
        sheet_type: str,
        record_scope: str,
        param_name_raw: str | None,
        value_raw: str | None,
        row_order: int,
        column_order: int,
        comp_record: dict[str, Any] | None = None,
        group_name: str | None = None,
        parse_status: str | None = None,
        parse_comment: str | None = None,
    ) -> dict[str, Any]:
        unit_raw, unit_norm = normalize_unit(param_name_raw or "")
        value_num = first_number(value_raw)
        if parse_status is None:
            numbers = number_tokens(value_raw)
            if value_num is not None:
                parse_status = "parsed"
            elif numbers:
                parse_status = "partial"
            else:
                parse_status = "text"
        return {
            "source_file": ctx["source_file"],
            "source_sheet": ctx["source_sheet"],
            "sheet_type": sheet_type,
            "base_model": ctx["base_model"],
            "record_scope": record_scope,
            "comp_full_name": None if comp_record is None else comp_record.get("comp_full_name"),
            "comp_brand": None if comp_record is None else comp_record.get("comp_brand"),
            "group_name": group_name,
            "param_name_raw": param_name_raw,
            "param_name_norm": normalize_label(param_name_raw),
            "value_raw": value_raw,
            "value_num": value_num,
            "value_text": value_raw,
            "unit_raw": unit_raw,
            "unit_norm": unit_norm,
            "parse_status": parse_status,
            "parse_comment": parse_comment,
            "cell_address": f"{get_column_letter(column_order)}{row_order}",
            "row_order": row_order,
            "column_order": column_order,
        }

    def set_price(self, record: dict[str, Any], raw_value: str | None, label: str) -> bool:
        _, price_min, price_max, price_comment = parse_price_values(raw_value, label)
        if price_min is None or price_max is None:
            return False
        record["price_rub_min"] = price_min
        record["price_rub_max"] = price_max
        if price_comment:
            record["price_comment"] = price_comment
        return True

    def normalize_fuel_type(self, raw: str | None) -> str | None:
        text = normalize_label(raw)
        if not text:
            return None
        if "элект" in text:
            return "electric"
        if "диз" in text:
            return "diesel"
        if "бенз" in text:
            return "gasoline"
        if "метан" in text or "cng" in text:
            return "cng"
        if "газ" in text and "бенз" in text:
            return "bifuel"
        return clean_text(raw)

    def apply_common_vehicle_fields(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[bool, str | None]:
        mapped = False
        norm = normalize_label(label)

        def assign(field: str, value: Any, *, overwrite: bool = False) -> None:
            nonlocal mapped
            if value is None:
                return
            if overwrite or record.get(field) is None:
                record[field] = value
                mapped = True

        if "цен" in norm or "стоимость приобрет" in norm:
            mapped = self.set_price(record, raw_value, label) or mapped

        if "полная" in norm and "снаряж" in norm:
            gross_bounds, curb_bounds = parse_slash_pair_bounds(raw_value)
            if assign_numeric_range(record, "gross_weight_kg", *gross_bounds):
                mapped = True
            if assign_numeric_range(record, "curb_weight_kg", *curb_bounds):
                mapped = True
            return mapped, None

        if "габарит" in norm or "дхшхв" in norm:
            length, width, height = parse_dimensions(raw_value)
            assign("length_mm", length)
            assign("width_mm", width)
            assign("height_mm", height)
            return mapped, None

        if "колесная база" in norm:
            assign("wheelbase_mm", first_number(raw_value))
        elif norm.startswith("длина") and "отсек" not in norm and "проем" not in norm:
            assign("length_mm", first_number(raw_value))
        elif norm.startswith("ширина") and "отсек" not in norm and "проем" not in norm:
            assign("width_mm", first_number(raw_value))
        elif norm.startswith("высота") and "отсек" not in norm and "проем" not in norm:
            assign("height_mm", first_number(raw_value))
        elif "дорожный просвет" in norm:
            assign("ground_clearance_mm", first_number(raw_value))
        elif "кабина" in norm and "мест" in norm:
            assign("cab_seat_count", parse_additive_count(raw_value))
        elif "мест для сидения" in norm:
            assign("seat_count", parse_additive_count(raw_value))
        elif "пассажировмест" in norm:
            assign("passenger_capacity", parse_additive_count(raw_value) or (int(first_number(raw_value)) if first_number(raw_value) is not None else None))
        elif norm.startswith("полная масса"):
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif "снаряженная масса" in norm:
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif "грузопод" in norm:
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif "колесная формула" in norm:
            drive_type, wheel_formula = parse_drive(raw_value)
            assign("drive_type", drive_type)
            assign("wheel_formula", wheel_formula)
        elif "привод" in norm:
            drive_type, _ = parse_drive(raw_value)
            assign("drive_type", drive_type)
        elif "кпп" in norm or "коробка передач" in norm or "трансмис" in norm:
            transmission_type, gears = parse_transmission(raw_value)
            assign("transmission_type", transmission_type)
            assign("transmission_gears", gears)
        elif "топливный бак" in norm:
            assign("fuel_tank_l", first_number(raw_value))
        elif "расход топлива" in norm:
            assign("fuel_consumption_l100km", first_number(raw_value))
        elif "топливо" == norm or norm.startswith("топливо"):
            assign("engine_fuel_type", self.normalize_fuel_type(raw_value))
        elif "объем двигателя" in norm:
            assign("engine_volume_l", parse_engine_volume(raw_value, label))
        elif "мощность" in norm:
            hp, kw = parse_power(raw_value, label)
            assign("engine_power_hp", hp)
            assign("engine_power_kw", kw)
        elif "гарантия общ" in norm:
            months, km = parse_warranty(raw_value)
            assign("warranty_months", months)
            assign("warranty_km", km)
        elif "межсервис" in norm:
            assign("service_interval_km", first_number(raw_value))
        elif "класс" in norm or "уровень пола" in norm or norm == "двигатель":
            self.append_note(record, f"{clean_text(label)}: {clean_text(raw_value)}")

        return mapped, None

    def apply_paz_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped, comment = self.apply_common_vehicle_fields(record, label, raw_value)
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, comment

    def apply_bort_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped, comment = self.apply_common_vehicle_fields(record, label, raw_value)
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, comment

    def apply_cmf_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped, comment = self.apply_common_vehicle_fields(record, label, raw_value)
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, comment

    def apply_ldt_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped, comment = self.apply_common_vehicle_fields(record, label, raw_value)
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, comment

    def apply_miniven_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped, comment = self.apply_common_vehicle_fields(record, label, raw_value)
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, comment

    def apply_bus_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped, comment = self.apply_common_vehicle_fields(record, label, raw_value)
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, comment

    def apply_ownership_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        norm = normalize_label(label)
        if "цен" in norm:
            mapped = self.set_price(record, raw_value, label) or mapped
        elif "грузопод" in norm:
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value, scale_small_values=True) or mapped
        elif "мощность" in norm:
            hp, kw = parse_power(raw_value, label)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
        elif "расход топлива" in norm:
            value = first_number(raw_value)
            if value is not None:
                record["fuel_consumption_l100km"] = value
                mapped = True
        elif "стоимость владения" in norm and "1 км" in norm:
            value = first_number(raw_value)
            if value is not None:
                record["ownership_cost_rub_km"] = value
                mapped = True
        else:
            mapped, _ = self.apply_common_vehicle_fields(record, label, raw_value)
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def apply_bort_stvlad_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        label_text = clean_text(label)
        if label_text == "Цена":
            mapped = self.set_price(record, raw_value, label_text) or mapped
        elif label_text == "Мощность двигателя, л.с.":
            hp, kw = parse_power(raw_value, label_text)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
        elif label_text == "Грузоподъёмность":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value, scale_small_values=True) or mapped
        elif label_text == "Средний расход топлива фактический, л/100 км":
            value = first_number(raw_value)
            if value is not None:
                record["fuel_consumption_l100km"] = value
                mapped = True
        elif label_text == "Затраты на 1 км пробега":
            value = first_number(raw_value)
            if value is not None:
                record["ownership_cost_rub_km"] = value
                mapped = True
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def apply_cmf_stvlad_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        if clean_text(label) == "Цена":
            mapped = self.set_price(record, raw_value, label) or mapped
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def apply_ldt_stvlad_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        label_text = clean_text(label)
        if label_text == "Цена приобретения":
            mapped = self.set_price(record, raw_value, label_text) or mapped
        elif label_text == "Средний расход топлива (фактический, в реальных условиях эксплуатации), л/100 км":
            value = first_number(raw_value)
            if value is not None:
                record["fuel_consumption_l100km"] = value
                mapped = True
        elif label_text == "Затраты на 1 км пробега":
            value = first_number(raw_value)
            if value is not None:
                record["ownership_cost_rub_km"] = value
                mapped = True
        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def get_bort_ttx_layout(self, ws: Worksheet) -> dict[str, Any]:
        if clean_text(ws["A2"].value) != "Марка" or clean_text(ws["A3"].value) != "Модель":
            raise RuntimeError("БОРТ_ТТХ и состав: unexpected header rows")
        if clean_text(ws["A6"].value) == "Двухрядная кабина":
            return {
                "allowed_cols": list(range(2, 24)),
                "fuel_row": 28,
            }
        if clean_text(ws["A6"].value) == "Полная масса а/м":
            return {
                "allowed_cols": list(range(2, 25)),
                "fuel_row": 27,
            }
        raise RuntimeError("БОРТ_ТТХ и состав: unsupported layout")

    def disambiguate_bort_ttx_records(self, ws: Worksheet, records: list[dict[str, Any]], fuel_row: int) -> None:
        groups: dict[str, list[dict[str, Any]]] = {}
        for record in records:
            full_name = clean_text(record.get("comp_full_name"))
            if not full_name:
                continue
            groups.setdefault(full_name, []).append(record)

        for full_name, group in groups.items():
            if len(group) == 1:
                continue

            fuel_values = [clean_text(ws.cell(fuel_row, record["column_index"]).value) for record in group]
            if all(fuel_values) and len(set(fuel_values)) == len(group):
                for record, fuel_value in zip(group, fuel_values, strict=True):
                    record["comp_full_name"] = join_parts(record["comp_full_name"], fuel_value)
                continue

            wheelbase_values = [clean_text(ws.cell(4, record["column_index"]).value) for record in group]
            if all(wheelbase_values) and len(set(wheelbase_values)) == len(group):
                for record, wheelbase_value in zip(group, wheelbase_values, strict=True):
                    record["comp_full_name"] = join_parts(record["comp_full_name"], f"{wheelbase_value} мм")
                continue

            raise RuntimeError(f"БОРТ_ТТХ и состав: duplicate model header cannot be disambiguated: {full_name}")

    def apply_bort_ttx_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        label_text = clean_text(label)

        if label_text == "колесная база, мм":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
        elif label_text == "Полная масса а/м":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса а/м, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text in {"Грузоподьемность а/м, кг", "грузоподьемность шасси, кг"}:
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value, prefer_existing=True) or mapped
        elif label_text == "Габариты а/м (ДхШхВ), мм":
            length, width, height = parse_dimensions(raw_value)
            if length is not None:
                record["length_mm"] = length
                mapped = True
            if width is not None:
                record["width_mm"] = width
                mapped = True
            if height is not None:
                record["height_mm"] = height
                mapped = True
        elif label_text == "Дорожный просвет (клиренс), мм":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "топливо":
            value = clean_text(raw_value)
            if value:
                record["engine_fuel_type"] = value
                mapped = True
        elif label_text == "Объем двигателя, куб.см":
            value = parse_engine_volume(raw_value, label_text)
            if value is not None:
                record["engine_volume_l"] = value
                mapped = True
        elif label_text == "Мощность двигателя, л.с.":
            hp, kw = parse_power(raw_value, label_text)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = parse_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text == "Топливный бак, л":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула (привод)":
            drive_type, wheel_formula = parse_drive(raw_value)
            if drive_type is not None:
                record["drive_type"] = drive_type
                mapped = True
            if wheel_formula is not None:
                record["wheel_formula"] = wheel_formula
                mapped = True
        elif label_text == "гарантия, мес":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["warranty_months"] = int(value)
                mapped = True
        elif label_text == "пробег тыс. км":
            value = parse_thousand_km_value(raw_value)
            if value is not None:
                record["warranty_km"] = value
                mapped = True
        elif label_text == "Межсервисный интервал, км":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        elif label_text == "Средний расход топлива фактический, л/100 км":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["fuel_consumption_l100km"] = value
                mapped = True
        elif label_text == "Цена а/м":
            mapped = self.set_price(record, raw_value, label_text) or mapped
        elif label_text == "Затраты на 1 км пробега":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["ownership_cost_rub_km"] = value
                mapped = True

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def get_cmf_ttx_layout(self, ws: Worksheet) -> dict[str, Any]:
        if clean_text(ws["A2"].value) != "Марка" or clean_text(ws["A3"].value) != "Модель":
            raise RuntimeError("ЦМФ_ТТХ и состав: unexpected header rows")
        if clean_text(ws["A6"].value) == "Двухрядная кабина":
            return {
                "allowed_cols": list(range(2, 20)),
            }
        if clean_text(ws["A6"].value) == "Полная масса":
            return {
                "allowed_cols": list(range(2, 18)),
            }
        raise RuntimeError("ЦМФ_ТТХ и состав: unsupported layout")

    def apply_cmf_ttx_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        label_text = clean_text(label)

        if label_text == "колесная база, мм":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
        elif label_text == "Полная масса":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподьемность, кг":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text == "Длина, мм":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["length_mm"] = value
                mapped = True
        elif label_text == "Ширина, мм":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["width_mm"] = value
                mapped = True
        elif label_text == "Высота, мм":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["height_mm"] = value
                mapped = True
        elif label_text == "Дорожный просвет, мм":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливо":
            value = clean_text(raw_value)
            if value:
                record["engine_fuel_type"] = value
                mapped = True
        elif label_text == "Объем двигателя, куб.см":
            value = parse_mixed_engine_volume_l(raw_value)
            if value is not None:
                record["engine_volume_l"] = value
                mapped = True
        elif label_text == "Мощность двигателя, л.с.":
            hp, kw = parse_power(raw_value, label_text)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = parse_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text == "Топливный бак, л":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула (привод)":
            drive_type, wheel_formula = parse_drive(raw_value)
            if drive_type is not None:
                record["drive_type"] = drive_type
                mapped = True
            if wheel_formula is not None:
                record["wheel_formula"] = wheel_formula
                mapped = True
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
        elif label_text == "Межсервисный интервал, км":
            value = parse_interval_km_value(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        elif label_text == "Средний расход топлива фактический, л/100 км":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["fuel_consumption_l100km"] = value
                mapped = True

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def apply_miniven_ttx_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        label_text = clean_text(label)

        if label_text == "Цена без учета скидок и акций, руб.":
            mapped = self.set_price(record, raw_value, label_text) or mapped
        elif label_text == "Колесная база, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кол-во мест для сидения (включая водителя), чел.":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["seat_count"] = value
                mapped = True
        elif label_text == "Пассажировместимость (без учета водителя), чел.":
            value = single_numeric_value(raw_value)
            if value is not None:
                record["passenger_capacity"] = int(value)
                mapped = True
        elif label_text == "Полная масса":
            mapped = assign_primary_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса, кг":
            mapped = assign_primary_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподьемность, кг":
            mapped = assign_primary_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text == "Длина, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["length_mm"] = value
                mapped = True
        elif label_text == "Ширина, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["width_mm"] = value
                mapped = True
        elif label_text == "Высота, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["height_mm"] = value
                mapped = True
        elif label_text == "Дорожный просвет, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливо":
            value = clean_text(raw_value)
            if value:
                record["engine_fuel_type"] = value
                mapped = True
        elif label_text == "Объем двигателя, куб.см":
            value = parse_mixed_engine_volume_l(raw_value)
            if value is not None:
                record["engine_volume_l"] = value
                mapped = True
        elif label_text == "Мощность двигателя, кВт (л.с.)":
            hp, kw = parse_minivan_power(raw_value)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = parse_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text == "Топливный бак, л":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула":
            drive_type, wheel_formula = parse_drive(raw_value)
            if drive_type is not None:
                record["drive_type"] = drive_type
                mapped = True
            if wheel_formula is not None:
                record["wheel_formula"] = wheel_formula
                mapped = True
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огранич" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = primary_numeric_value(raw_value) or parse_interval_km_value(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        elif label_text == "Средний расход топлива фактический, л/100 км":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["fuel_consumption_l100km"] = value
                mapped = True
        elif label_text == "Затраты на 1 км пробега":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["ownership_cost_rub_km"] = value
                mapped = True

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def apply_miniven_short_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        mapped = False
        label_text = clean_text(label)

        if label_text == "Цена без учета скидок и акций, руб.":
            mapped = self.set_price(record, raw_value, label_text) or mapped
        elif label_text == "Колесная база, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кол-во мест для сидения (включая водителя), чел.":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["seat_count"] = value
                mapped = True
        elif label_text == "Пассажировместимость (без учета водителя), чел.":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["passenger_capacity"] = int(value)
                mapped = True
        elif label_text == "Полная масса":
            mapped = assign_primary_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса, кг":
            mapped = assign_primary_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподьемность, кг":
            mapped = assign_primary_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text == "Длина, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["length_mm"] = value
                mapped = True
        elif label_text == "Ширина, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["width_mm"] = value
                mapped = True
        elif label_text == "Высота, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["height_mm"] = value
                mapped = True
        elif label_text == "Дорожный просвет, мм":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливо":
            value = clean_text(raw_value)
            if value:
                record["engine_fuel_type"] = value
                mapped = True
        elif label_text == "Объем двигателя, куб.см":
            value = parse_mixed_engine_volume_l(raw_value)
            if value is not None:
                record["engine_volume_l"] = value
                mapped = True
        elif label_text == "Мощность двигателя, кВт (л.с.)":
            hp, kw = parse_minivan_power(raw_value)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = parse_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text == "Топливный бак, л":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула":
            drive_type, wheel_formula = parse_drive(raw_value)
            if drive_type is not None:
                record["drive_type"] = drive_type
                mapped = True
            if wheel_formula is not None:
                record["wheel_formula"] = wheel_formula
                mapped = True
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огранич" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Гарантия от сквозной коррозии":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = primary_numeric_value(raw_value) or parse_interval_km_value(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        elif label_text == "Экологический класс":
            self.append_note(record, f"{label_text}: {raw_value}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def ingest_model_matrix(
        self,
        ctx: dict[str, Any],
        *,
        sheet_type: str,
        records: list[dict[str, Any]],
        label_col: int,
        row_start: int,
        row_end: int,
        apply_label: Callable[[dict[str, Any], str, str], tuple[str, str | None]],
    ) -> None:
        ws = ctx["ws"]
        for row_index in range(row_start, row_end + 1):
            label = clean_text(ws.cell(row_index, label_col).value)
            if not label:
                continue
            for record in records:
                column_index = record["column_index"]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = apply_label(record, label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type=sheet_type,
                        record_scope="model",
                        param_name_raw=label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

    def insert_option_row(
        self,
        ctx: dict[str, Any],
        *,
        record: dict[str, Any],
        option_group: str | None,
        option_name: str,
        row_order: int,
        raw_value: str | None = None,
        price_override: str | None = None,
        status_override: str | None = None,
        status_only: bool = False,
    ) -> None:
        if clean_text(raw_value) is None and clean_text(price_override) is None and clean_text(status_override) is None:
            return
        if status_only:
            status_raw = clean_text(status_override if status_override is not None else raw_value)
            status_norm = parse_option_cell(status_raw)[1]
            price_raw = None
            price_rub = None
        else:
            status_raw, status_norm, price_raw, price_rub = parse_option_cell(raw_value)
            if status_override is not None:
                status_raw = clean_text(status_override)
                status_norm = parse_option_cell(status_override)[1]
            if price_override is not None:
                price_raw = clean_text(price_override)
                _, price_min, _, _ = parse_price_values(price_raw, option_name)
                price_rub = price_min
        if status_raw is None and price_raw is not None:
            status_raw = "опция"
            status_norm = "optional"
        self.insert_option(
            {
                "source_file": ctx["source_file"],
                "source_sheet": ctx["source_sheet"],
                "base_model": ctx["base_model"],
                "comp_full_name": record.get("comp_full_name"),
                "comp_brand": record.get("comp_brand"),
                "option_group": option_group,
                "option_name": option_name,
                "option_status_raw": status_raw,
                "option_status_norm": status_norm,
                "option_price_raw": price_raw,
                "option_price_rub": price_rub,
                "notes": None,
                "row_order": row_order,
                "column_order": record["column_index"],
            }
        )
        raw_value_joined = clean_text(" / ".join(part for part in [price_raw, status_raw] if clean_text(part)))
        parse_status = "parsed" if status_norm or price_rub is not None else ("partial" if number_tokens(raw_value_joined) else "text")
        self.insert_raw(
            self.raw_row(
                ctx,
                sheet_type="options",
                record_scope="option",
                param_name_raw=option_name,
                value_raw=raw_value_joined,
                row_order=row_order,
                column_order=record["column_index"],
                comp_record=record,
                parse_status=parse_status,
                parse_comment=None,
            )
        )

    def insert_minivan_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        records_by_col = {record["column_index"]: record for record in records}
        option_rows = [
            (59, "Функциональность / Зимний пакет"),
            (60, "Функциональность / Зимний пакет"),
            (61, "Функциональность / Зимний пакет"),
            (62, "Функциональность / Зимний пакет"),
            (63, "Функциональность / Зимний пакет"),
            (66, "Функциональность / Пакет Pro"),
            (67, "Функциональность / Пакет Pro"),
            (68, "Функциональность / Пакет Pro"),
            (70, "Функциональность / Пакет Pro"),
            (71, "Функциональность / Пакет Pro"),
            (72, "Функциональность / Пакет Pro"),
            (73, "Функциональность / Пакет Pro"),
            (74, "Функциональность / Пакет Pro"),
            (75, "Функциональность / Пакет Pro"),
            (76, "Функциональность / Пакет Pro"),
            (77, "Функциональность / Пакет Pro"),
            (78, "Функциональность / Пакет Pro"),
            (79, "Функциональность / Пакет Pro"),
            (80, "Функциональность / Пакет Pro"),
            (81, "Функциональность / Пакет Pro"),
            (84, "Безопасность"),
            (85, "Безопасность"),
            (86, "Безопасность"),
            (87, "Безопасность"),
            (88, "Безопасность"),
            (89, "Безопасность"),
            (90, "Безопасность"),
            (93, "Экстерьер"),
            (94, "Экстерьер"),
            (95, "Экстерьер"),
            (96, "Экстерьер"),
            (97, "Экстерьер"),
            (98, "Экстерьер"),
            (99, "Экстерьер"),
            (100, "Экстерьер"),
            (103, "Кабина"),
            (104, "Кабина"),
            (105, "Кабина"),
            (106, "Кабина"),
            (107, "Кабина"),
            (108, "Кабина"),
            (109, "Кабина"),
            (110, "Кабина"),
            (111, "Кабина"),
            (112, "Кабина"),
            (113, "Кабина"),
            (114, "Кабина"),
            (115, "Кабина"),
            (116, "Кабина"),
            (117, "Кабина"),
            (118, "Кабина"),
            (120, "Пассажирский салон"),
            (121, "Пассажирский салон"),
            (122, "Пассажирский салон"),
            (123, "Пассажирский салон"),
            (124, "Пассажирский салон"),
            (125, "Пассажирский салон"),
            (126, "Пассажирский салон"),
            (127, "Пассажирский салон"),
            (128, "Пассажирский салон"),
            (129, "Пассажирский салон"),
            (130, "Пассажирский салон"),
            (131, "Пассажирский салон"),
            (132, "Пассажирский салон"),
            (133, "Пассажирский салон"),
            (134, "Пассажирский салон"),
            (143, "Дополнительно / Эксклюзивные опции"),
        ]
        for row_index, option_group in option_rows:
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_col.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group=option_group,
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def insert_minivan_short_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        records_by_col = {record["column_index"]: record for record in records}
        option_rows = [
            (59, "Функциональность / Зимний пакет"),
            (60, "Функциональность / Зимний пакет"),
            (61, "Функциональность / Зимний пакет"),
            (62, "Функциональность / Зимний пакет"),
            (65, "Функциональность / Пакет Pro"),
            (66, "Функциональность / Пакет Pro"),
            (67, "Функциональность / Пакет Pro"),
            (69, "Функциональность / Пакет Pro"),
            (70, "Функциональность / Пакет Pro"),
            (71, "Функциональность / Пакет Pro"),
            (72, "Функциональность / Пакет Pro"),
            (73, "Функциональность / Пакет Pro"),
            (74, "Функциональность / Пакет Pro"),
            (75, "Функциональность / Пакет Pro"),
            (76, "Функциональность / Пакет Pro"),
            (77, "Функциональность / Пакет Pro"),
            (78, "Функциональность / Пакет Pro"),
            (79, "Функциональность / Пакет Pro"),
            (80, "Функциональность / Пакет Pro"),
            (83, "Безопасность"),
            (84, "Безопасность"),
            (85, "Безопасность"),
            (86, "Безопасность"),
            (87, "Безопасность"),
            (88, "Безопасность"),
            (89, "Безопасность"),
            (92, "Экстерьер"),
            (93, "Экстерьер"),
            (94, "Экстерьер"),
            (95, "Экстерьер"),
            (96, "Экстерьер"),
            (97, "Экстерьер"),
            (98, "Экстерьер"),
            (99, "Экстерьер"),
            (102, "Кабина"),
            (103, "Кабина"),
            (104, "Кабина"),
            (105, "Кабина"),
            (106, "Кабина"),
            (107, "Кабина"),
            (108, "Кабина"),
            (109, "Кабина"),
            (110, "Кабина"),
            (111, "Кабина"),
            (112, "Кабина"),
            (113, "Кабина"),
            (114, "Кабина"),
            (115, "Кабина"),
            (116, "Кабина"),
            (117, "Кабина"),
            (119, "Пассажирский салон"),
            (120, "Пассажирский салон"),
            (121, "Пассажирский салон"),
            (122, "Пассажирский салон"),
            (123, "Пассажирский салон"),
            (124, "Пассажирский салон"),
            (125, "Пассажирский салон"),
            (126, "Пассажирский салон"),
            (127, "Пассажирский салон"),
            (128, "Пассажирский салон"),
            (129, "Пассажирский салон"),
            (130, "Пассажирский салон"),
            (131, "Пассажирский салон"),
            (132, "Пассажирский салон"),
            (133, "Пассажирский салон"),
        ]
        for row_index, option_group in option_rows:
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_col.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group=option_group,
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def parse_raw_grid_sheet(self, ctx: dict[str, Any], *, sheet_type: str) -> None:
        ws = ctx["ws"]
        top_headers = {column: clean_text(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)}
        for row_index in range(1, ws.max_row + 1):
            row_label = clean_text(ws.cell(row_index, 1).value)
            for column_index in range(1, ws.max_column + 1):
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                if row_index == 1 and column_index == 1:
                    continue
                if column_index == 1:
                    param_name = top_headers.get(2) or row_label or f"row_{row_index}"
                else:
                    param_name = row_label or top_headers.get(column_index) or f"col_{column_index}"
                group_name = None
                header_value = top_headers.get(column_index)
                if column_index > 1 and header_value and header_value != param_name:
                    group_name = header_value
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type=sheet_type,
                        record_scope="qualitative",
                        param_name_raw=param_name,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=group_name,
                    )
                )

    def insert_advantages_text_block(
        self,
        ctx: dict[str, Any],
        *,
        label_col: int,
        left_value_col: int,
        right_value_col: int,
        row_start: int,
        row_end: int,
    ) -> None:
        ws = ctx["ws"]
        current_label: str | None = None
        left_group = clean_text(ws.cell(1, left_value_col).value)
        right_group = clean_text(ws.cell(1, right_value_col).value)
        for row_index in range(row_start, row_end + 1):
            label = clean_text(ws.cell(row_index, label_col).value)
            if label:
                current_label = label
            if not current_label:
                continue
            for column_index, group_name in (
                (left_value_col, left_group),
                (right_value_col, right_group),
            ):
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_label,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=group_name,
                    )
                )

    def insert_advantages_summary_block(
        self,
        ctx: dict[str, Any],
        *,
        label_col: int,
        value_cols: list[int],
        row_start: int,
        row_end: int,
    ) -> None:
        ws = ctx["ws"]
        current_label: str | None = None
        group_names = {column_index: clean_text(ws.cell(1, column_index).value) for column_index in value_cols}
        for row_index in range(row_start, row_end + 1):
            label = clean_text(ws.cell(row_index, label_col).value)
            if label:
                current_label = label
            if not current_label:
                continue
            for column_index in value_cols:
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_label,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=group_names[column_index],
                    )
                )

    def insert_reviews_block(
        self,
        ctx: dict[str, Any],
        *,
        text_col: int,
        count_col: int,
        header_row: int,
        body_row_start: int,
        body_row_end: int,
        category_rows: list[int],
    ) -> None:
        ws = ctx["ws"]
        group_name = clean_text(ws.cell(header_row, text_col).value)
        header_count = clean_text(ws.cell(header_row, count_col).value)
        if header_count is not None:
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="qualitative",
                    record_scope="qualitative",
                    param_name_raw=None,
                    value_raw=header_count,
                    row_order=header_row,
                    column_order=count_col,
                    group_name=group_name,
                )
            )

        current_category: str | None = None
        category_row_set = set(category_rows)
        for row_index in range(body_row_start, body_row_end + 1):
            label = clean_text(ws.cell(row_index, text_col).value)
            count_value = clean_text(ws.cell(row_index, count_col).value)
            if row_index in category_row_set:
                current_category = label
                if count_value is not None:
                    self.insert_raw(
                        self.raw_row(
                            ctx,
                            sheet_type="qualitative",
                            record_scope="qualitative",
                            param_name_raw=current_category,
                            value_raw=count_value,
                            row_order=row_index,
                            column_order=count_col,
                            group_name=group_name,
                        )
                    )
                continue
            if label is None:
                continue
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="qualitative",
                    record_scope="qualitative",
                    param_name_raw=current_category,
                    value_raw=label,
                    row_order=row_index,
                    column_order=text_col,
                    group_name=group_name,
                )
            )

    def parse_service_group(
        self,
        ctx: dict[str, Any],
        *,
        group_name: str,
        row_start: int,
        row_end: int,
        service_col: int,
        mileage_col: int,
        diesel_col: int | None = None,
        diesel_promo_col: int | None = None,
        gasoline_col: int | None = None,
        gasoline_promo_col: int | None = None,
    ) -> None:
        ws = ctx["ws"]
        header_map = {
            service_col: clean_text(ws.cell(3, service_col).value),
            mileage_col: clean_text(ws.cell(3, mileage_col).value),
        }
        for column_index in (diesel_col, diesel_promo_col, gasoline_col, gasoline_promo_col):
            if column_index is not None:
                header_map[column_index] = clean_text(ws.cell(3, column_index).value)
        for row_index in range(row_start, row_end + 1):
            service_label = clean_text(ws.cell(row_index, service_col).value)
            if not service_label:
                continue
            mileage = first_number(clean_text(ws.cell(row_index, mileage_col).value))
            service_index = int(first_number(service_label)) if first_number(service_label) is not None else None
            row = {
                "source_file": ctx["source_file"],
                "source_sheet": ctx["source_sheet"],
                "base_model": ctx["base_model"],
                "group_name": group_name,
                "service_label": service_label,
                "service_index": service_index,
                "mileage_km": mileage,
                "cost_diesel_rub": first_number(clean_text(ws.cell(row_index, diesel_col).value)) if diesel_col else None,
                "cost_diesel_promo_rub": first_number(clean_text(ws.cell(row_index, diesel_promo_col).value)) if diesel_promo_col else None,
                "cost_gasoline_rub": first_number(clean_text(ws.cell(row_index, gasoline_col).value)) if gasoline_col else None,
                "cost_gasoline_promo_rub": first_number(clean_text(ws.cell(row_index, gasoline_promo_col).value)) if gasoline_promo_col else None,
                "notes": None,
            }
            self.insert_service(row)
            for column_index in [service_col, mileage_col, diesel_col, diesel_promo_col, gasoline_col, gasoline_promo_col]:
                if column_index is None:
                    continue
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="service_group",
                        record_scope="service_group",
                        param_name_raw=header_map[column_index] or service_label,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=group_name,
                    )
                )

    def insert_reviews_text_only_block(
        self,
        ctx: dict[str, Any],
        *,
        text_col: int,
        header_row: int,
        body_row_start: int,
        body_row_end: int,
        category_rows: list[int],
    ) -> None:
        ws = ctx["ws"]
        group_name = clean_text(ws.cell(header_row, text_col).value)
        current_category: str | None = None
        category_row_set = set(category_rows)
        for row_index in range(body_row_start, body_row_end + 1):
            value = clean_text(ws.cell(row_index, text_col).value)
            if value is None:
                continue
            if row_index in category_row_set:
                current_category = value
                continue
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="qualitative",
                    record_scope="qualitative",
                    param_name_raw=current_category,
                    value_raw=value,
                    row_order=row_index,
                    column_order=text_col,
                    group_name=group_name,
                )
            )

    def insert_qualitative_column(
        self,
        ctx: dict[str, Any],
        *,
        text_col: int,
        header_row: int,
        row_start: int,
        row_end: int,
    ) -> None:
        ws = ctx["ws"]
        group_name = clean_text(ws.cell(header_row, text_col).value)
        for row_index in range(row_start, row_end + 1):
            value = clean_text(ws.cell(row_index, text_col).value)
            if value is None:
                continue
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="qualitative",
                    record_scope="qualitative",
                    param_name_raw=None,
                    value_raw=value,
                    row_order=row_index,
                    column_order=text_col,
                    group_name=group_name,
                )
            )

    def insert_negative_positive_orm_section(
        self,
        ctx: dict[str, Any],
        *,
        title_row: int,
        header_row: int,
        body_row_start: int,
        body_row_end: int,
        summary_row: int,
    ) -> None:
        ws = ctx["ws"]
        section_title = clean_text(ws.cell(title_row, 1).value)
        negative_header = clean_text(ws.cell(header_row, 2).value)
        positive_header = clean_text(ws.cell(header_row, 6).value)
        current_topic: str | None = None
        for row_index in range(body_row_start, body_row_end + 1):
            topic = clean_text(ws.cell(row_index, 1).value)
            if topic:
                current_topic = topic
            if current_topic is None:
                continue
            for column_index in (2, 3, 4):
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_topic,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=section_title,
                        parse_comment=negative_header,
                    )
                )
            positive_value = clean_text(ws.cell(row_index, 6).value)
            if positive_value is not None:
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_topic,
                        value_raw=positive_value,
                        row_order=row_index,
                        column_order=6,
                        group_name=section_title,
                        parse_comment=positive_header,
                    )
                )
        for column_index in (3, 4):
            value = clean_text(ws.cell(summary_row, column_index).value)
            if value is None:
                continue
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="qualitative",
                    record_scope="qualitative",
                    param_name_raw=None,
                    value_raw=value,
                    row_order=summary_row,
                    column_order=column_index,
                    group_name=section_title,
                )
            )

    def insert_quality_note_cell(self, ctx: dict[str, Any], *, group_name: str, row_index: int, column_index: int) -> None:
        ws = ctx["ws"]
        value = clean_text(ws.cell(row_index, column_index).value)
        if value is None:
            return
        self.insert_raw(
            self.raw_row(
                ctx,
                sheet_type="qualitative",
                record_scope="qualitative",
                param_name_raw=None,
                value_raw=value,
                row_order=row_index,
                column_order=column_index,
                group_name=group_name,
            )
        )

    def insert_quality_metric_table(
        self,
        ctx: dict[str, Any],
        *,
        group_name: str,
        header_row: int,
        row_start: int,
        row_end: int,
        label_col: int,
        value_cols: list[int],
    ) -> None:
        ws = ctx["ws"]
        headers = {column_index: clean_text(ws.cell(header_row, column_index).value) for column_index in value_cols}
        current_label: str | None = None
        for row_index in range(row_start, row_end + 1):
            label = clean_text(ws.cell(row_index, label_col).value)
            if label:
                current_label = label
            if current_label is None:
                continue
            for column_index in value_cols:
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_label,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=group_name,
                        parse_comment=headers[column_index],
                    )
                )

    def insert_quality_top_problems_table(
        self,
        ctx: dict[str, Any],
        *,
        section_title: str,
        count_header: str | None,
        row_start: int,
        row_end: int,
        label_col: int,
        count_col: int,
    ) -> None:
        ws = ctx["ws"]
        current_node: str | None = None
        for row_index in range(row_start, row_end + 1):
            label = clean_text(ws.cell(row_index, label_col).value)
            count_value = clean_text(ws.cell(row_index, count_col).value)
            if label and count_value is not None:
                current_node = label
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_node,
                        value_raw=count_value,
                        row_order=row_index,
                        column_order=count_col,
                        group_name=section_title,
                        parse_comment=count_header,
                    )
                )
                continue
            if label is None:
                continue
            if current_node is None:
                raise ValueError(f"{ws.title}: top problems row {row_index} has no current node")
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="qualitative",
                    record_scope="qualitative",
                    param_name_raw=None,
                    value_raw=label,
                    row_order=row_index,
                    column_order=label_col,
                    group_name=current_node,
                )
            )

    def parse_sheet_quality_gazelle(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Справка")
        self.require_exact(ws, "A5", "1. Рекламации от потребителей.")
        self.require_exact(ws, "A7", "Показатель")
        self.require_exact(ws, "B7", "Единица измерения")
        self.require_exact(ws, "A16", "ТОП проблем по качеству")
        self.require_exact(ws, "A17", "Узел")
        self.require_exact(ws, "B17", "количество дефектов за период")

        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=2, column_index=1)
        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=3, column_index=1)
        self.insert_quality_metric_table(
            ctx,
            group_name=clean_text(ws["A2"].value) or "качество",
            header_row=7,
            row_start=8,
            row_end=14,
            label_col=1,
            value_cols=[2, 3, 4, 5, 6],
        )
        self.insert_quality_top_problems_table(
            ctx,
            section_title="ТОП проблем по качеству",
            count_header=clean_text(ws["B17"].value),
            row_start=18,
            row_end=72,
            label_col=1,
            count_col=2,
        )

    def parse_sheet_quality_sadko(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Общее - 10 т")
        self.require_exact(ws, "J1", "Общее - 8,7 т")
        self.require_exact(ws, "A2", "Узел")
        self.require_exact(ws, "J2", "Узел")
        self.require_exact(ws, "A28", "Справка")
        self.require_exact(ws, "J27", "Справка")
        self.require_exact(ws, "A34", "Показатель")
        self.require_exact(ws, "J34", "Показатель")
        self.require_exact(ws, "A43", "ТОП проблем по качеству - 10 т")
        self.require_exact(ws, "J43", "ТОП проблем по качеству - 8,7 т")

        self.insert_quality_metric_table(
            ctx,
            group_name=clean_text(ws["A1"].value) or "качество",
            header_row=2,
            row_start=3,
            row_end=26,
            label_col=1,
            value_cols=[2, 3, 4, 5, 6, 7, 8],
        )
        self.insert_quality_metric_table(
            ctx,
            group_name=clean_text(ws["J1"].value) or "качество",
            header_row=2,
            row_start=3,
            row_end=25,
            label_col=10,
            value_cols=[11, 12, 13, 14, 15, 16],
        )

        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=28, column_index=1)
        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=29, column_index=1)
        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=30, column_index=1)
        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=27, column_index=10)
        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=28, column_index=10)
        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=29, column_index=10)
        self.insert_quality_note_cell(ctx, group_name="Справка", row_index=30, column_index=10)

        self.insert_quality_metric_table(
            ctx,
            group_name=clean_text(ws["A29"].value) or "качество",
            header_row=34,
            row_start=35,
            row_end=41,
            label_col=1,
            value_cols=[2, 3, 4, 5],
        )
        self.insert_quality_metric_table(
            ctx,
            group_name=clean_text(ws["J28"].value) or "качество",
            header_row=34,
            row_start=35,
            row_end=41,
            label_col=10,
            value_cols=[11, 12, 13, 14],
        )

        self.insert_quality_top_problems_table(
            ctx,
            section_title=clean_text(ws["A43"].value) or "качество",
            count_header=clean_text(ws["B44"].value),
            row_start=45,
            row_end=64,
            label_col=1,
            count_col=2,
        )
        self.insert_quality_top_problems_table(
            ctx,
            section_title=clean_text(ws["J43"].value) or "качество",
            count_header=clean_text(ws["B44"].value),
            row_start=45,
            row_end=82,
            label_col=10,
            count_col=11,
        )

    def parse_sheet_svesy_bazy(self, ctx: dict[str, Any]) -> None:
        return

    def parse_sheet_list2(self, ctx: dict[str, Any]) -> None:
        return

    def parse_sheet_paz(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_labels = {
            4: "Старт производства",
            5: "Класс",
            6: "Колесная формула",
            7: "Полная / снаряженная массы, кг",
            8: "Кол-во служебных дверей",
            9: "Пассажировместимость общая",
            10: "Мест для сидения",
            11: "Уровень пола",
            12: "Габариты ДxШxВ, мм",
            13: "Колесная база, мм",
            14: "Дорожный просвет, мм",
            15: "Двигатель",
            16: "Мощность двигателя, л.с.",
            20: "КПП",
            21: "Емкость топливного бака, л",
            26: "Гарантия",
            27: "Цена по прайсу, тыс. руб. с НДС",
            29: "Комплектации и опции",
        }
        for row_index, expected_label in expected_labels.items():
            actual_label = clean_text(ws.cell(row_index, 2).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected B{row_index} = {expected_label!r}, got {actual_label!r}")
        records = self.build_paz_records(ctx)
        for row_index in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 20, 21, 26, 27]:
            label = clean_text(ws.cell(row_index, 2).value)
            if not label:
                raise ValueError(f"{ws.title}: expected label in B{row_index}")
            for record in records:
                column_index = record["column_index"]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment, value_num = self.apply_paz_exact_label(record, label, raw_value)
                raw_row = self.raw_row(
                    ctx,
                    sheet_type="technical",
                    record_scope="model",
                    param_name_raw=label,
                    value_raw=raw_value,
                    row_order=row_index,
                    column_order=column_index,
                    comp_record=record,
                    parse_status=parse_status,
                    parse_comment=parse_comment,
                )
                raw_row["value_num"] = value_num
                self.insert_raw(raw_row)
        self.insert_paz_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_paz_expanded(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_labels = {
            4: "Старт производства",
            5: "Класс",
            6: "Колесная формула",
            7: "Полная / снаряженная массы, кг",
            8: "Нагрузка на переднюю/заднюю ось",
            9: "Кол-во служебных дверей",
            10: "Пассажировместимость общая",
            11: "Мест для сидения",
            12: "Уровень пола",
            13: "Объём багажника, м3",
            14: "Габариты ДxШxВ, мм",
            15: "Высота салона",
            16: "Колесная база, мм",
            17: "Радиус разворота",
            18: "Дорожный просвет, мм",
            19: "Двигатель",
            20: "Объем двигателя",
            21: "Мощность двигателя, л.с.",
            22: "Крутящий момент, Н*м",
            23: "Максимальная скорость, км/ч",
            24: "Экология",
            25: "Расход топлива на л/м3/100 км",
            26: "КПП",
            27: "Емкость топливного бака, л",
            28: "Емкость газового баллона, м³",
            29: "Шины",
            30: "Передняя подвеска",
            31: "Задняя подвеска",
            32: "Тормозная системв",
            33: "Тормозные механизмы передние / задние",
            34: "Вентиляция",
            35: "Система отопления",
            36: "Гарантия",
            37: "Межсервисный интервал",
            38: "Гарантия на кузов",
            39: "Цена по прайсу, тыс. руб. с НДС",
            41: "Комплектации и опции",
        }
        for row_index, expected_label in expected_labels.items():
            actual_label = clean_text(ws.cell(row_index, 2).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected B{row_index} = {expected_label!r}, got {actual_label!r}")
        records = self.build_paz_expanded_records(ctx)
        for row_index in range(4, 40):
            label = clean_text(ws.cell(row_index, 2).value)
            if not label:
                raise ValueError(f"{ws.title}: expected label in B{row_index}")
            for record in records:
                column_index = record["column_index"]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment, value_num = self.apply_paz_expanded_exact_label(record, label, raw_value)
                raw_row = self.raw_row(
                    ctx,
                    sheet_type="technical",
                    record_scope="model",
                    param_name_raw=label,
                    value_raw=raw_value,
                    row_order=row_index,
                    column_order=column_index,
                    comp_record=record,
                    parse_status=parse_status,
                    parse_comment=parse_comment,
                )
                raw_row["value_num"] = value_num
                self.insert_raw(raw_row)
        self.insert_paz_expanded_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_bort_ttx(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "A2", "Марка")
        self.require_contains(ws, "A3", "Модель")
        layout = self.get_bort_ttx_layout(ws)
        records = self.models_from_brand_model_rows(
            ctx,
            2,
            3,
            allowed_cols=layout["allowed_cols"],
            prefix_rows=[1],
            sheet_type="technical",
            vehicle_type="truck",
            body_type="flatbed",
        )
        self.ingest_model_matrix(
            ctx,
            sheet_type="technical",
            records=records,
            label_col=1,
            row_start=4,
            row_end=ws.max_row,
            apply_label=self.apply_bort_ttx_label,
        )
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_cmf_ttx(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "A2", "Марка")
        self.require_contains(ws, "A3", "Модель")
        layout = self.get_cmf_ttx_layout(ws)
        records = self.models_from_brand_model_rows(
            ctx,
            2,
            3,
            allowed_cols=layout["allowed_cols"],
            prefix_rows=[1],
            sheet_type="technical",
            vehicle_type="van",
            body_type="panel_van",
        )
        self.ingest_model_matrix(
            ctx,
            sheet_type="technical",
            records=records,
            label_col=1,
            row_start=4,
            row_end=ws.max_row,
            apply_label=self.apply_cmf_ttx_label,
        )
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_ldt_ttx(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_labels = {
            2: "Модель",
            3: "Колесная база, мм",
            4: "Кабина, мест",
            5: 'Модификация "спальник" или "двухрядка"',
            6: "Полная масса, кг",
            7: "Снаряженная масса шасси, кг",
            8: "Грузоподъемность шасси, кг",
            9: "Габаритные размеры шасси, мм (ДхШхВ)",
            13: "Дорожный просвет, мм",
            14: "Двигатель",
            15: "Экологический класс",
            16: "Объем двигателя, л",
            17: "Max мощность, кВт (л.с.)",
            25: "Коробка передач",
            31: "Топливный бак, л",
            32: "Комплектация и опции",
            72: "Сервис и гарантия",
            73: "Гарантия, месяцы / пробег тыс. км",
            75: "Межсервисный интервал, км",
            77: "Дополнительно",
            80: "Средний расход топлива фактический, л/100 км",
            81: "Стоимость владения, руб",
            82: "Цена а/м",
            86: "Затраты на 1 км пробега",
        }
        for row_index, expected_label in expected_labels.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
        records = self.build_ldt_ttx_records(ctx)
        for row_index in list(range(3, 32)) + list(range(73, 87)):
            label = clean_text(ws.cell(row_index, 1).value)
            if not label:
                raise ValueError(f"{ws.title}: expected label in A{row_index}")
            for record in records:
                column_index = record["column_index"]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment, value_num = self.apply_ldt_ttx_exact_label(record, label, raw_value)
                raw_row = self.raw_row(
                    ctx,
                    sheet_type="technical",
                    record_scope="model",
                    param_name_raw=label,
                    value_raw=raw_value,
                    row_order=row_index,
                    column_order=column_index,
                    comp_record=record,
                    parse_status=parse_status,
                    parse_comment=parse_comment,
                )
                raw_row["value_num"] = value_num
                self.insert_raw(raw_row)
        self.insert_ldt_ttx_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_miniven(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "A2", "Марка")
        self.require_contains(ws, "A3", "Модель")
        self.require_contains(ws, "A56", "Комплектации и опции")
        self.require_contains(ws, "A136", "Гарантия общая")
        records = self.build_minivan_records(ctx)
        self.ingest_model_matrix(ctx, sheet_type="technical", records=records, label_col=1, row_start=4, row_end=55, apply_label=self.apply_miniven_ttx_label)
        self.ingest_model_matrix(ctx, sheet_type="technical", records=records, label_col=1, row_start=136, row_end=150, apply_label=self.apply_miniven_ttx_label)
        self.insert_minivan_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_miniven_2(self, ctx: dict[str, Any]) -> None:
        self.parse_sheet_miniven(ctx)

    def parse_sheet_bus_ttx(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_cells = {
            "A2": "Производитель",
            "A3": "модификация",
            "A39": "Тормозное управление",
            "A45": "Колеса, шины",
            "A51": "Комплектации и опции",
            "A52": "Функциональность",
            "A65": "Безопасность",
            "A73": "Экстерьер",
            "A79": "Кабина",
            "A98": "Пассажирский салон",
            "A116": "Сервис и гарантия",
        }
        for coord, expected_value in expected_cells.items():
            actual_value = clean_text(ws[coord].value)
            if actual_value != expected_value:
                raise ValueError(f"{ws.title}: expected {coord} = {expected_value!r}, got {actual_value!r}")

        layout = self.get_bus_ttx_layout(ws)
        model_columns = layout["model_columns"]
        tail_mode = layout["tail_mode"]
        records = self.build_bus_ttx_records(ctx, model_columns)
        records_by_col = {record["column_index"]: record for record in records}

        model_rows = {
            4: "Класс автобуса",
            5: "Кол-во мест для сидения (включая водителя), чел.",
            6: "Пассажировместимость (без учета водителя), чел.",
            7: "Полная масса",
            8: "Снаряженная масса min, кг",
            9: "Снаряженная масса max, кг",
            10: "Грузоподьемность min, кг",
            11: "Грузоподьемность max, кг",
            12: "Длина, мм",
            13: "без задн.подножки, мм",
            14: "колесная база, мм",
            15: "передний свес, мм",
            16: "задний свес, мм",
            17: "Ширина, мм",
            18: "Высота (по люку), мм",
            19: "Высота с доп. кондиционером, мм",
            20: "Ширина проема боковой двери, мм",
            21: "Высота проема боковой двери, мм",
            22: "Высота проема боковой двери от ступеньки, мм",
            23: "Высота проема задних дверей, мм",
            24: "Ширина проема задних дверей, мм",
            25: "Высота потолка в пассажирском салоне, мм",
            26: "Ширина между входными поручнями",
            27: "Мин. радиус разворота, м",
            28: "Дорожный просвет, мм",
            29: "Двигатель",
            30: "Топливо",
            31: "Объем двигателя, куб.см",
            32: "Мощность двигателя, кВт (л.с.)",
            33: "при об. мин.",
            34: "Крутящий момент",
            35: "при об. мин.",
            36: "Экологический класс",
            37: "КПП",
            38: "Задний блокируемый дифференциал",
            40: "Тормозные механизмы",
            41: "Электронная тормозная система",
            42: "Стояночный тормоз",
            43: "Блокируемый дифференциал",
            44: "Топливный бак, л",
            46: "Колесная формула",
            47: "Ошиновка",
            48: "Размер шин передних",
            49: "задних",
            50: "Диски",
            117: "Гарантия общая, мес./пробег тыс. км",
            118: "Гарантия от сквозной коррозии",
            119: "Межсервисный интервал, км",
        }
        for row_index, expected_label in model_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
            for column_index in model_columns:
                record = records_by_col[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment, _ = self.apply_bus_ttx_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        if clean_text(ws["A120"].value) is not None:
            raise ValueError(f"{ws.title}: expected A120 to be empty, got {clean_text(ws['A120'].value)!r}")
        for column_index in model_columns:
            record = records_by_col[column_index]
            raw_price = clean_text(ws.cell(120, column_index).value)
            if raw_price is None:
                continue
            parse_status, parse_comment, _ = self.apply_bus_ttx_exact_label(record, "Цена", raw_price)
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="technical",
                    record_scope="model",
                    param_name_raw="Цена",
                    value_raw=raw_price,
                    row_order=120,
                    column_order=column_index,
                    comp_record=record,
                    parse_status=parse_status,
                    parse_comment=parse_comment,
                )
            )

        self.insert_bus_ttx_main_option_rows(ctx, records)
        self.insert_bus_ttx_tail_option_rows(ctx, records_by_col, tail_mode)
        for record in records:
            self.insert_normalized(record)

    def build_9_10t_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in range(2, 8):
            full_name = clean_text(ws.cell(1, column_index).value)
            if not full_name:
                raise ValueError(f"{ws.title}: expected model name in {get_column_letter(column_index)}1")
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    full_name,
                    comp_brand=full_name,
                    vehicle_type="truck",
                    body_type="chassis",
                )
            )
        return records

    def apply_9_10t_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        label_text = clean_text(label)
        mapped = False

        if label_text == "Колесная формула":
            normalized = raw_value.replace("х", "x").replace("Х", "x")
            formula_match = re.fullmatch(r"\s*(\d)\s*[x*]\s*(\d)\s*", normalized)
            if formula_match:
                record["wheel_formula"] = f"{formula_match.group(1)}x{formula_match.group(2)}"
                mapped = True
        elif label_text == "Полная масса, кг":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса шасси, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподъемность шасси, кг":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text in {
            "Коэффициент PL/GVW",
            "Допустимая нагрузка на переднюю / заднюю оси, кг",
            "Габариты ДxШxВ, м",
            "Радиус поворота, м",
            "Угол преодолеваемого подъема, %",
            "Двигатель",
            "Крутящий момент, Н*м",
            "Экология",
            "Шины",
            "Задняя подвеска",
            "Тормозная система",
            "Тормозные механизмы передние / задние",
            "Максимальная скорость, км/ч",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Дорожный просвет, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Мощность двигателя, л.с.":
            value = strict_single_number(raw_value)
            if value is not None:
                record["engine_power_hp"] = value
                mapped = True
        elif label_text == "КПП":
            transmission_type = clean_text(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
                if transmission_type[0].isdigit():
                    record["transmission_gears"] = int(transmission_type[0])
        elif label_text == "Емкость топливного бака, л":
            value = strict_single_number(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Гарантия":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огранич" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал":
            value = parse_thousand_km_value(raw_value)
            if value is None:
                value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def parse_sheet_9_10t(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A1"].value) != "Технические характеристики":
            raise ValueError(f"{ws.title}: expected A1 = 'Технические характеристики', got {clean_text(ws['A1'].value)!r}")
        expected_rows = {
            2: "Колесная формула",
            3: "Полная масса, кг",
            4: "Снаряженная масса шасси, кг",
            5: "Грузоподъемность шасси, кг",
            6: "Коэффициент PL/GVW",
            7: "Допустимая нагрузка на переднюю / заднюю оси, кг",
            8: "Габариты ДxШxВ, м",
            9: "Колесная база, мм",
            10: "Дорожный просвет, мм",
            11: "Радиус поворота, м",
            12: "Угол преодолеваемого подъема, %",
            13: "Двигатель",
            15: "Мощность двигателя, л.с.",
            16: "Крутящий момент, Н*м",
            17: "Экология",
            18: "КПП",
            19: "Емкость топливного бака, л",
            20: "Шины",
            21: "Задняя подвеска",
            22: "Тормозная система",
            23: "Тормозные механизмы передние / задние",
            24: "Максимальная скорость, км/ч",
            25: "Гарантия",
            26: "Межсервисный интервал",
        }
        for row_index, expected_label in expected_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
        for row_index in range(27, ws.max_row + 1):
            if any(clean_text(ws.cell(row_index, column_index).value) is not None for column_index in range(1, ws.max_column + 1)):
                raise ValueError(f"{ws.title}: expected rows 27+ to be empty, found data in row {row_index}")

        records = self.build_9_10t_records(ctx)
        records_by_column = {record["column_index"]: record for record in records}
        for row_index, expected_label in expected_rows.items():
            for column_index in range(2, 8):
                record = records_by_column[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = self.apply_9_10t_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        for column_index in range(2, 8):
            continuation_raw = clean_text(ws.cell(14, column_index).value)
            if continuation_raw is None:
                continue
            record = records_by_column[column_index]
            self.append_note(record, f"Двигатель: {continuation_raw}")
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="technical",
                    record_scope="model",
                    param_name_raw="Двигатель",
                    value_raw=continuation_raw,
                    row_order=14,
                    column_order=column_index,
                    comp_record=record,
                    parse_status="text",
                    parse_comment="continuation_row",
                )
            )
        for record in records:
            self.insert_normalized(record)

    def build_bort_short_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in range(2, 11):
            brand = clean_text(ws.cell(1, column_index).value)
            model = clean_text(ws.cell(2, column_index).value)
            if not brand:
                raise ValueError(f"{ws.title}: expected brand in {get_column_letter(column_index)}1")
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    join_parts(brand, model) or brand,
                    comp_brand=brand,
                    vehicle_type="truck",
                    body_type="flatbed",
                )
            )
        return records

    def parse_bort_short_dimensions(self, raw_value: str) -> tuple[float | None, float | None, float | None]:
        matches = re.findall(r"(\d{3,5})\s*[хxХX×]\s*(\d{3,5})\s*[хxХX×]\s*(\d{3,5})(?!\d)", raw_value)
        if len(matches) != 1:
            return None, None, None
        return (
            parse_number_literal(matches[0][0]),
            parse_number_literal(matches[0][1]),
            parse_number_literal(matches[0][2]),
        )

    def parse_bort_short_engine(self, raw_value: str) -> tuple[str | None, float | None, float | None]:
        matches = re.findall(r"([А-Яа-яA-Za-z0-9 ()/+.-]+):\s*([\d.,]+)\s*л,\s*([\d.,]+)\s*л\.с\.", raw_value)
        if len(matches) != 1:
            return None, None, None
        return (
            clean_text(matches[0][0]),
            parse_number_literal(matches[0][1]),
            parse_number_literal(matches[0][2]),
        )

    def parse_bort_short_transmission(self, raw_value: str) -> tuple[str | None, int | None]:
        text = clean_text(raw_value)
        if not text:
            return None, None
        norm = normalize_label(text)
        if "/" in text or "опция" in norm or "база" in norm or "диз" in norm or "бенз" in norm:
            return None, None
        gears = int(text[0]) if text[0].isdigit() else None
        return text, gears

    def is_bort_short_ambiguous_option_status(self, raw_value: str) -> bool:
        text = clean_text(raw_value)
        if not text:
            return False
        norm = normalize_label(text)
        has_base = "база" in norm
        has_optional = "опция" in norm
        has_not_present = "-" in text or "нет" in norm
        if has_base and has_optional:
            return True
        if (has_base or has_optional) and has_not_present and "/" in text:
            return True
        return False

    def apply_bort_short_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        label_text = clean_text(label)
        mapped = False

        if label_text == "колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
        elif label_text == "Полная масса а/м":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса а/м, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподьемность а/м, кг":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text == "Габариты а/м (ДхШхВ), мм":
            length_mm, width_mm, height_mm = self.parse_bort_short_dimensions(raw_value)
            if length_mm is not None and width_mm is not None and height_mm is not None:
                record["length_mm"] = length_mm
                record["width_mm"] = width_mm
                record["height_mm"] = height_mm
                mapped = True
        elif label_text in {"вместимость палет, шт", "Мин. радиус разворота, м", "Погруз-я высота, мм"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Дорожный просвет (клиренс), мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
            fuel_type, volume_l, power_hp = self.parse_bort_short_engine(raw_value)
            if fuel_type is not None:
                record["engine_fuel_type"] = fuel_type
                mapped = True
            if volume_l is not None:
                record["engine_volume_l"] = volume_l
                mapped = True
            if power_hp is not None:
                record["engine_power_hp"] = power_hp
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = self.parse_bort_short_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text in {
            "Тормозная система",
            "Тормозные механизмы",
            "Электронная тормозная система",
            "Моторный (горный) тормоз",
            "Стояночный тормоз",
            "Блокируемый дифференциал",
            "Ошиновка",
            "от сквозной коррозии, лет",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливный бак, л":
            if "/" not in raw_value:
                value = strict_single_number(raw_value)
                if value is not None:
                    record["fuel_tank_l"] = value
                    mapped = True
        elif label_text == "Колесная формула (привод)":
            normalized = raw_value.replace("х", "x").replace("Х", "x")
            formula_match = re.match(r"(\d)x(\d)\s*\(([^)]+)\)", normalized)
            if formula_match:
                record["wheel_formula"] = f"{formula_match.group(1)}x{formula_match.group(2)}"
                record["drive_type"] = clean_text(formula_match.group(3))
                mapped = True
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огранич" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def insert_bort_short_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A22"].value) != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected A22 = 'Комплектации и опции'")
        records_by_column = {record["column_index"]: record for record in records}
        for row_index in range(23, 44):
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                if self.is_bort_short_ambiguous_option_status(raw_value):
                    self.insert_option(
                        {
                            "source_file": ctx["source_file"],
                            "source_sheet": ctx["source_sheet"],
                            "base_model": ctx["base_model"],
                            "comp_full_name": record.get("comp_full_name"),
                            "comp_brand": record.get("comp_brand"),
                            "option_group": "Комплектации и опции",
                            "option_name": option_name,
                            "option_status_raw": raw_value,
                            "option_status_norm": None,
                            "option_price_raw": None,
                            "option_price_rub": None,
                            "notes": None,
                            "row_order": row_index,
                            "column_order": record["column_index"],
                        }
                    )
                    self.insert_raw(
                        self.raw_row(
                            ctx,
                            sheet_type="options",
                            record_scope="option",
                            param_name_raw=option_name,
                            value_raw=raw_value,
                            row_order=row_index,
                            column_order=record["column_index"],
                            comp_record=record,
                            parse_status="text",
                            parse_comment="ambiguous_option_status",
                        )
                    )
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group="Комплектации и опции",
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def parse_sheet_bort_short(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_cells = {
            "A1": "Марка",
            "A2": "Модель",
            "A22": "Комплектации и опции",
            "A44": "Колесная формула (привод)",
            "A45": "Ошиновка",
            "A46": "Гарантия общая, мес./пробег тыс. км",
            "A47": "от сквозной коррозии, лет",
            "A48": "Межсервисный интервал, км",
        }
        for coord, expected_value in expected_cells.items():
            actual_value = clean_text(ws[coord].value)
            if actual_value != expected_value:
                raise ValueError(f"{ws.title}: expected {coord} = {expected_value!r}, got {actual_value!r}")

        records = self.build_bort_short_records(ctx)
        records_by_column = {record["column_index"]: record for record in records}
        model_rows = {
            3: "колесная база, мм",
            4: "Кабина, мест",
            5: "Полная масса а/м",
            6: "Снаряженная масса а/м, кг",
            7: "Грузоподьемность а/м, кг",
            8: "Габариты а/м (ДхШхВ), мм",
            9: "вместимость палет, шт",
            10: "Мин. радиус разворота, м",
            11: "Погруз-я высота, мм",
            12: "Дорожный просвет (клиренс), мм",
            13: "Двигатель",
            14: "КПП",
            15: "Тормозная система",
            16: "Тормозные механизмы",
            17: "Электронная тормозная система",
            18: "Моторный (горный) тормоз",
            19: "Стояночный тормоз",
            20: "Блокируемый дифференциал",
            21: "Топливный бак, л",
            44: "Колесная формула (привод)",
            45: "Ошиновка",
            46: "Гарантия общая, мес./пробег тыс. км",
            47: "от сквозной коррозии, лет",
            48: "Межсервисный интервал, км",
        }
        for row_index, expected_label in model_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
            for column_index in range(2, 11):
                record = records_by_column[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = self.apply_bort_short_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        self.insert_bort_short_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def build_cmf_short_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in range(2, 11):
            brand = clean_text(ws.cell(1, column_index).value)
            model = clean_text(ws.cell(2, column_index).value)
            if not brand:
                raise ValueError(f"{ws.title}: expected brand in {get_column_letter(column_index)}1")
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    join_parts(brand, model) or brand,
                    comp_brand=brand,
                    vehicle_type="van",
                    body_type="panel_van",
                )
            )
        return records

    def parse_cmf_short_dimensions(self, raw_value: str) -> tuple[float | None, float | None, float | None]:
        matches = re.findall(r"(\d{3,5})\s*[хxХX×]\s*(\d{3,5})\s*[хxХX×]\s*(\d{3,5})(?!\d)", raw_value)
        if len(matches) != 1:
            return None, None, None
        return (
            parse_number_literal(matches[0][0]),
            parse_number_literal(matches[0][1]),
            parse_number_literal(matches[0][2]),
        )

    def parse_cmf_short_engine(self, raw_value: str) -> tuple[str | None, float | None, float | None]:
        matches = re.findall(r"([А-Яа-яA-Za-z0-9 ()/+.-]+):\s*([\d.,]+)\s*л,\s*([\d.,]+)\s*л\.с\.", raw_value)
        if len(matches) != 1:
            return None, None, None
        return (
            clean_text(matches[0][0]),
            parse_number_literal(matches[0][1]),
            parse_number_literal(matches[0][2]),
        )

    def parse_cmf_short_transmission(self, raw_value: str) -> tuple[str | None, int | None]:
        text = clean_text(raw_value)
        if not text or "/" in text:
            return None, None
        gears = int(text[0]) if text[0].isdigit() else None
        return text, gears

    def parse_cmf_short_drive(self, raw_value: str) -> tuple[str | None, str | None]:
        text = clean_text(raw_value)
        if not text:
            return None, None
        normalized = text.replace("х", "x").replace("Х", "x")
        formula_match = re.match(r"(\d)x(\d)\s*\(([^)]+)\)", normalized)
        if not formula_match:
            return None, None
        wheel_formula = f"{formula_match.group(1)}x{formula_match.group(2)}"
        drive_type = clean_text(formula_match.group(3))
        if drive_type not in {"задний", "передний", "полный"}:
            drive_type = None
        return wheel_formula, drive_type

    def parse_cmf_short_service_interval(self, raw_value: str) -> float | None:
        text = clean_text(raw_value)
        if not text:
            return None
        km_matches = re.findall(r"(\d[\d\s.,]*)\s*км", text, flags=re.IGNORECASE)
        if len(km_matches) == 1:
            return parse_number_literal(km_matches[0])
        return strict_single_number(text)

    def is_cmf_short_ambiguous_option_status(self, raw_value: str) -> bool:
        text = clean_text(raw_value)
        if not text:
            return False
        norm = normalize_label(text)
        has_base = "база" in norm
        has_optional = "опция" in norm
        has_not_present = "-" in text or "нет" in norm
        if has_base and has_optional:
            return True
        if (has_base or has_optional) and has_not_present and "/" in text:
            return True
        return False

    def apply_cmf_short_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        label_text = clean_text(label)
        mapped = False

        if label_text == "колесная база, мм":
            value = None if "/" in raw_value else strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
        elif label_text == "Полная масса":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподьемность, кг":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text == "Габариты а/м (ДхШхВ), мм":
            length_mm, width_mm, height_mm = self.parse_cmf_short_dimensions(raw_value)
            if length_mm is not None and width_mm is not None and height_mm is not None:
                record["length_mm"] = length_mm
                record["width_mm"] = width_mm
                record["height_mm"] = height_mm
                mapped = True
        elif label_text in {
            "Ширина проема боковой двери, мм",
            "Высота проема боковой двери, мм",
            "Высота проема задних дверей, мм",
            "Ширина проема задних дверей, мм",
            "Длина грузового отсека, мм",
            "Ширина грузового отсека, мм",
            "Высота грузового отсека, мм",
            "Объем грузового отсека, м³",
            "Количество палет",
            "Погруз-я высота, мм",
            "Мин. радиус разворота, м",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Дорожный просвет, мм":
            value = None if "/" in raw_value else strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
            fuel_type, volume_l, power_hp = self.parse_cmf_short_engine(raw_value)
            if fuel_type is not None:
                record["engine_fuel_type"] = fuel_type
                mapped = True
            if volume_l is not None:
                record["engine_volume_l"] = volume_l
                mapped = True
            if power_hp is not None:
                record["engine_power_hp"] = power_hp
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = self.parse_cmf_short_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text in {"Тормозная система", "Тормозные механизмы", "Электронная тормозная система", "Стояночный тормоз"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливный бак, л":
            value = None if "/" in raw_value else strict_single_number(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула (привод)":
            wheel_formula, drive_type = self.parse_cmf_short_drive(raw_value)
            if wheel_formula is not None:
                record["wheel_formula"] = wheel_formula
                mapped = True
            if drive_type is not None:
                record["drive_type"] = drive_type
                mapped = True
            if drive_type is None:
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text in {"Ошиновка", "Диски", "от сквозной коррозии"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огр" in normalize_label(raw_value) or "год" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = self.parse_cmf_short_service_interval(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
            if "год" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def insert_cmf_short_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A28"].value) != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected A28 = 'Комплектации и опции'")
        if clean_text(ws["A49"].value) != "Грузовой фургон":
            raise ValueError(f"{ws.title}: expected A49 = 'Грузовой фургон'")
        records_by_column = {record["column_index"]: record for record in records}
        option_rows = {**{row_index: "Комплектации и опции" for row_index in range(29, 49)}, **{row_index: "Грузовой фургон" for row_index in range(50, 53)}}
        for row_index, option_group in option_rows.items():
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                if self.is_cmf_short_ambiguous_option_status(raw_value):
                    self.insert_option(
                        {
                            "source_file": ctx["source_file"],
                            "source_sheet": ctx["source_sheet"],
                            "base_model": ctx["base_model"],
                            "comp_full_name": record.get("comp_full_name"),
                            "comp_brand": record.get("comp_brand"),
                            "option_group": option_group,
                            "option_name": option_name,
                            "option_status_raw": raw_value,
                            "option_status_norm": None,
                            "option_price_raw": None,
                            "option_price_rub": None,
                            "notes": None,
                            "row_order": row_index,
                            "column_order": record["column_index"],
                        }
                    )
                    self.insert_raw(
                        self.raw_row(
                            ctx,
                            sheet_type="options",
                            record_scope="option",
                            param_name_raw=option_name,
                            value_raw=raw_value,
                            row_order=row_index,
                            column_order=record["column_index"],
                            comp_record=record,
                            parse_status="text",
                            parse_comment="ambiguous_option_status",
                        )
                    )
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group=option_group,
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def parse_sheet_cmf_short(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_cells = {
            "A2": "Модель",
            "A28": "Комплектации и опции",
            "A49": "Грузовой фургон",
            "A53": "Колесная формула (привод)",
            "A54": "Ошиновка",
            "A55": "Диски",
            "A56": "Гарантия общая, мес./пробег тыс. км",
            "A57": "от сквозной коррозии",
            "A58": "Межсервисный интервал, км",
        }
        for coord, expected_value in expected_cells.items():
            actual_value = clean_text(ws[coord].value)
            if actual_value != expected_value:
                raise ValueError(f"{ws.title}: expected {coord} = {expected_value!r}, got {actual_value!r}")

        records = self.build_cmf_short_records(ctx)
        records_by_column = {record["column_index"]: record for record in records}
        model_rows = {
            3: "колесная база, мм",
            4: "Кабина, мест",
            5: "Полная масса",
            6: "Снаряженная масса, кг",
            7: "Грузоподьемность, кг",
            8: "Габариты а/м (ДхШхВ), мм",
            9: "Ширина проема боковой двери, мм",
            10: "Высота проема боковой двери, мм",
            11: "Высота проема задних дверей, мм",
            12: "Ширина проема задних дверей, мм",
            13: "Длина грузового отсека, мм",
            14: "Ширина грузового отсека, мм",
            15: "Высота грузового отсека, мм",
            16: "Объем грузового отсека, м³",
            17: "Количество палет",
            18: "Погруз-я высота, мм",
            19: "Дорожный просвет, мм",
            20: "Мин. радиус разворота, м",
            21: "Двигатель",
            22: "КПП",
            23: "Тормозная система",
            24: "Тормозные механизмы",
            25: "Электронная тормозная система",
            26: "Стояночный тормоз",
            27: "Топливный бак, л",
            53: "Колесная формула (привод)",
            54: "Ошиновка",
            55: "Диски",
            56: "Гарантия общая, мес./пробег тыс. км",
            57: "от сквозной коррозии",
            58: "Межсервисный интервал, км",
        }
        for row_index, expected_label in model_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
            for column_index in range(2, 11):
                record = records_by_column[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = self.apply_cmf_short_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        self.insert_cmf_short_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def build_bus_short_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in (2, 3, 4):
            brand = clean_text(ws.cell(1, column_index).value)
            model = clean_text(ws.cell(2, column_index).value)
            if not brand or not model:
                raise ValueError(
                    f"{ws.title}: expected complete header in {get_column_letter(column_index)}1:{get_column_letter(column_index)}2"
                )
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    join_parts(brand, model) or brand,
                    comp_brand=brand,
                    vehicle_type="bus",
                    body_type="bus",
                )
            )
        return records

    def parse_bus_short_dimensions(self, raw_value: str) -> tuple[float | None, float | None, float | None]:
        matches = re.findall(r"(\d[\d\s]*)\s*[хxХX]\s*(\d[\d\s]*)\s*[хxХX]\s*(\d[\d\s]*)", raw_value)
        if len(matches) != 1:
            return None, None, None
        length = parse_number_literal(matches[0][0])
        width = parse_number_literal(matches[0][1])
        height = parse_number_literal(matches[0][2])
        return length, width, height

    def parse_bus_short_engine(self, raw_value: str) -> tuple[str | None, float | None, float | None]:
        matches = re.findall(r"([А-Яа-яA-Za-z0-9 ()/+.-]+):\s*([\d.,]+)\s*л,\s*([\d.,]+)\s*л\.с\.", raw_value)
        if len(matches) != 1:
            return None, None, None
        fuel = clean_text(matches[0][0])
        volume_l = parse_number_literal(matches[0][1])
        power_hp = parse_number_literal(matches[0][2])
        return fuel, volume_l, power_hp

    def apply_bus_short_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        label_text = clean_text(label)
        mapped = False

        if label_text == "Класс автобуса":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Кол-во мест для сидения (включая водителя), чел.":
            value = strict_single_number(raw_value)
            if value is not None:
                record["seat_count"] = int(value)
                mapped = True
        elif label_text == "Пассажировместимость (без учета водителя), чел.":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["passenger_capacity"] = value
                mapped = True
        elif label_text == "Полная масса":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Габариты а/м (ДхШхВ), мм":
            length_mm, width_mm, height_mm = self.parse_bus_short_dimensions(raw_value)
            if length_mm is not None and width_mm is not None and height_mm is not None:
                record["length_mm"] = length_mm
                record["width_mm"] = width_mm
                record["height_mm"] = height_mm
                mapped = True
        elif label_text in {
            "Ширина проема боковой двери, мм",
            "Высота проема боковой двери, мм",
            "Высота потолка в пассажирском салоне, мм",
            "Ширина между входными поручнями",
            "Мин. радиус разворота, м",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Дорожный просвет, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
            fuel_type, volume_l, power_hp = self.parse_bus_short_engine(raw_value)
            if fuel_type is not None:
                record["engine_fuel_type"] = fuel_type
                mapped = True
            if volume_l is not None:
                record["engine_volume_l"] = volume_l
                mapped = True
            if power_hp is not None:
                record["engine_power_hp"] = power_hp
                mapped = True
        elif label_text == "КПП":
            transmission_type = clean_text(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
                gears = {"5МКПП": 5, "6МКПП": 6}.get(transmission_type)
                if gears is not None:
                    record["transmission_gears"] = gears
        elif label_text in {"Тормозные механизмы", "Электронная тормозная система", "Стояночный тормоз"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливный бак, л":
            value = strict_single_number(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула":
            normalized = raw_value.replace("х", "x").replace("Х", "x")
            formula_match = re.match(r"(\d)x(\d)\s*\(([^)]+)\)", normalized)
            if formula_match:
                record["wheel_formula"] = f"{formula_match.group(1)}x{formula_match.group(2)}"
                record["drive_type"] = clean_text(formula_match.group(3))
                mapped = True
        elif label_text in {"Ошиновка", "Диски"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огранич" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Гарантия от сквозной коррозии":
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def insert_bus_short_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A25"].value) != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected A25 = 'Комплектации и опции'")
        records_by_column = {record["column_index"]: record for record in records}
        for row_index in range(26, 59):
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group="Комплектации и опции",
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def parse_sheet_bus_short(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_cells = {
            "A1": "Производитель",
            "A2": "модификация",
            "A16": "Тормозное управление",
            "A21": "Колеса, шины",
            "A25": "Комплектации и опции",
            "A59": "Сервис и гарантия",
            "A60": "Гарантия общая, мес./пробег тыс. км",
            "A61": "Гарантия от сквозной коррозии",
            "A62": "Межсервисный интервал, км",
        }
        for coord, expected_value in expected_cells.items():
            actual_value = clean_text(ws[coord].value)
            if actual_value != expected_value:
                raise ValueError(f"{ws.title}: expected {coord} = {expected_value!r}, got {actual_value!r}")

        records = self.build_bus_short_records(ctx)
        records_by_column = {record["column_index"]: record for record in records}
        model_rows = {
            3: "Класс автобуса",
            4: "Кол-во мест для сидения (включая водителя), чел.",
            5: "Пассажировместимость (без учета водителя), чел.",
            6: "Полная масса",
            7: "Габариты а/м (ДхШхВ), мм",
            8: "Ширина проема боковой двери, мм",
            9: "Высота проема боковой двери, мм",
            10: "Высота потолка в пассажирском салоне, мм",
            11: "Ширина между входными поручнями",
            12: "Мин. радиус разворота, м",
            13: "Дорожный просвет, мм",
            14: "Двигатель",
            15: "КПП",
            17: "Тормозные механизмы",
            18: "Электронная тормозная система",
            19: "Стояночный тормоз",
            20: "Топливный бак, л",
            22: "Колесная формула",
            23: "Ошиновка",
            24: "Диски",
            60: "Гарантия общая, мес./пробег тыс. км",
            61: "Гарантия от сквозной коррозии",
            62: "Межсервисный интервал, км",
        }
        for row_index, expected_label in model_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
            for column_index in (2, 3, 4):
                record = records_by_column[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = self.apply_bus_short_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        self.insert_bus_short_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def build_ldt_short_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in range(2, 11):
            brand = clean_text(ws.cell(1, column_index).value)
            model = clean_text(ws.cell(2, column_index).value)
            if not brand:
                raise ValueError(f"{ws.title}: expected model brand in {get_column_letter(column_index)}1")
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    join_parts(brand, model) or brand,
                    comp_brand=brand,
                    vehicle_type="truck",
                    body_type="chassis",
                )
            )
        return records

    def parse_ldt_short_engine(self, raw_value: str) -> tuple[str | None, float | None, float | None]:
        matches = re.findall(r"([А-Яа-яA-Za-z0-9 ()/+.-]+):\s*([\d.,]+)\s*л,\s*([\d.,]+)\s*л\.с\.", raw_value)
        if len(matches) != 1:
            return None, None, None
        fuel = clean_text(matches[0][0])
        volume_l = parse_number_literal(matches[0][1])
        power_hp = parse_number_literal(matches[0][2])
        return fuel, volume_l, power_hp

    def parse_ldt_short_transmission(self, raw_value: str) -> tuple[str | None, int | None]:
        text = clean_text(raw_value)
        if not text:
            return None, None
        if "/" in text or "," in text or "опция" in normalize_label(text) or "база" in normalize_label(text):
            return None, None
        if text[0].isdigit():
            return text, int(text[0])
        return text, None

    def apply_ldt_short_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        label_text = clean_text(label)
        mapped = False

        if label_text == "Колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
        elif label_text == "Полная масса, кг":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса шасси, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподъемность шасси, кг":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text in {
            "Габаритные размеры шасси, мм (ДхШхВ)",
            "Мин. радиус разворота, м",
            "Погрузочная высота шасси, мм",
            "Система нейтрализации",
            "Аккумулятор",
            "Напряжение бортовой сети",
            "Тормозная система",
            "Тормозные механизмы",
            "Электронная тормозная система",
            "Стояночный тормоз",
            "Горный тормоз",
            "от сквозной коррозии, лет",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Дорожный просвет, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
            fuel_type, volume_l, power_hp = self.parse_ldt_short_engine(raw_value)
            if fuel_type is not None:
                record["engine_fuel_type"] = fuel_type
                mapped = True
            if volume_l is not None:
                record["engine_volume_l"] = volume_l
                mapped = True
            if power_hp is not None:
                record["engine_power_hp"] = power_hp
                mapped = True
        elif label_text == "Коробка передач":
            transmission_type, gears = self.parse_ldt_short_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text == "Топливный бак, л":
            value = primary_numeric_value(raw_value)
            if value is not None and strict_single_number(raw_value.split(",", 1)[-1] if "," in raw_value else raw_value) is not None:
                record["fuel_tank_l"] = value
                mapped = True
            elif value is not None and "/" not in raw_value:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огр" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def insert_ldt_short_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A23"].value) != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected A23 = 'Комплектации и опции'")
        records_by_column = {record["column_index"]: record for record in records}
        for row_index in range(24, 47):
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group="Комплектации и опции",
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def parse_sheet_ldt_short(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_cells = {
            "A2": "Модель",
            "A23": "Комплектации и опции",
            "A47": "Гарантия общая, мес./пробег тыс. км",
            "A48": "от сквозной коррозии, лет",
            "A49": "Межсервисный интервал, км",
        }
        for coord, expected_value in expected_cells.items():
            actual_value = clean_text(ws[coord].value)
            if actual_value != expected_value:
                raise ValueError(f"{ws.title}: expected {coord} = {expected_value!r}, got {actual_value!r}")

        records = self.build_ldt_short_records(ctx)
        records_by_column = {record["column_index"]: record for record in records}
        model_rows = {
            3: "Колесная база, мм",
            4: "Кабина, мест",
            5: "Полная масса, кг",
            6: "Снаряженная масса шасси, кг",
            7: "Грузоподъемность шасси, кг",
            8: "Габаритные размеры шасси, мм (ДхШхВ)",
            9: "Мин. радиус разворота, м",
            10: "Погрузочная высота шасси, мм",
            11: "Дорожный просвет, мм",
            12: "Двигатель",
            13: "Система нейтрализации",
            14: "Аккумулятор",
            15: "Напряжение бортовой сети",
            16: "Коробка передач",
            17: "Тормозная система",
            18: "Тормозные механизмы",
            19: "Электронная тормозная система",
            20: "Стояночный тормоз",
            21: "Горный тормоз",
            22: "Топливный бак, л",
            47: "Гарантия общая, мес./пробег тыс. км",
            48: "от сквозной коррозии, лет",
            49: "Межсервисный интервал, км",
        }
        for row_index, expected_label in model_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
            for column_index in range(2, 11):
                record = records_by_column[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = self.apply_ldt_short_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        self.insert_ldt_short_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_miniven_short(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "A2", "Марка, модель")
        self.require_contains(ws, "A3", "модификация")
        self.require_contains(ws, "A56", "Комплектации и опции")
        self.require_contains(ws, "A135", "Гарантия общая")
        records = self.build_minivan_records(ctx)
        self.ingest_model_matrix(ctx, sheet_type="technical", records=records, label_col=1, row_start=4, row_end=55, apply_label=self.apply_miniven_short_label)
        self.insert_minivan_short_option_rows(ctx, records)
        self.ingest_model_matrix(ctx, sheet_type="technical", records=records, label_col=1, row_start=135, row_end=137, apply_label=self.apply_miniven_short_label)
        for record in records:
            self.insert_normalized(record)

    def build_compass_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in (2, 4, 6, 8):
            header_raw = ws.cell(1, column_index).value
            full_name = clean_text(header_raw)
            brand = clean_text(str(header_raw).splitlines()[0]) if header_raw is not None else None
            if not full_name or not brand:
                raise ValueError(f"{ws.title}: expected model header in {get_column_letter(column_index)}1")
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    full_name,
                    comp_brand=brand,
                    vehicle_type="truck",
                    body_type="flatbed",
                )
            )
        return records

    def parse_compass_transmission(self, raw_value: str) -> tuple[str | None, int | None]:
        text = clean_text(raw_value)
        if not text or "/" in text or "," in text or "опция" in normalize_label(text) or "база" in normalize_label(text):
            return None, None
        gears = int(text[0]) if text[0].isdigit() else None
        return text, gears

    def is_compass_ambiguous_option_status(self, raw_value: str) -> bool:
        text = clean_text(raw_value)
        if not text:
            return False
        norm = normalize_label(text)
        has_base = "база" in norm
        has_optional = "опция" in norm
        has_not_present = "-" in text or "нет" in norm
        if has_base and has_optional:
            return True
        if (has_base or has_optional) and has_not_present and "/" in text:
            return True
        return False

    def apply_compass_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        label_text = clean_text(label)
        mapped = False

        if label_text == "колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
        elif label_text == "Полная масса а/м":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса а/м, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподьемность а/м, кг":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text == "Габариты а/м (ДхШхВ), мм":
            length_mm, width_mm, height_mm = parse_dimensions(raw_value)
            if length_mm is not None:
                record["length_mm"] = length_mm
                record["width_mm"] = width_mm
                record["height_mm"] = height_mm
                mapped = True
        elif label_text in {
            "нагрузка на переднюю ось, кг",
            "на заднюю, кг",
            "снаряженная масса шасси, кг",
            "грузоподьемность шасси, кг",
            "передний свет, мм",
            "задний свес, мм",
            "Колея передняя/задняя, мм",
            "Габариты шасси (ДхШхВ), мм",
            "Грузовая платформа (ДхШ), мм",
            "вместимость палет, шт",
            "Борт, материал",
            "откидывание боковых бортов",
            "пол",
            "пол ",
            "тент",
            "Мин. радиус разворота, м",
            "Погруз-я высота, мм",
            "Двигатель",
            "при об. мин.",
            "Крутящий момент",
            "Экологический класс",
            "Система нейтрализации",
            "Тормозная система",
            "Тормозные механизмы",
            "Электронная тормозная система",
            "Стояночный тормоз",
            "Блокируемый дифференциал",
            "Моторный тормоз",
            "Подвеска передняя",
            "задняя",
            "Аккумулятор, АКБ",
            "Ошиновка",
            "Размер шин передних",
            "задних",
        }:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Дорожный просвет (клиренс), мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "топливо":
            value = clean_text(raw_value)
            if value is not None:
                record["engine_fuel_type"] = value
                mapped = True
        elif label_text == "Объем двигателя, куб.см":
            value = parse_engine_volume(raw_value, label_text)
            if value is not None:
                record["engine_volume_l"] = value
                mapped = True
        elif label_text == "Мощность двигателя, л.с.":
            hp, kw = parse_power(raw_value, label_text)
            if hp is not None:
                record["engine_power_hp"] = hp
                mapped = True
            if kw is not None:
                record["engine_power_kw"] = kw
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = self.parse_compass_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text == "Топливный бак, л":
            value = primary_numeric_value(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула (привод)":
            normalized = raw_value.replace("х", "x").replace("Х", "x")
            formula_match = re.match(r"(\d)x(\d)\s*\(([^)]+)\)", normalized)
            if formula_match:
                record["wheel_formula"] = f"{formula_match.group(1)}x{formula_match.group(2)}"
                record["drive_type"] = clean_text(formula_match.group(3))
                mapped = True
        elif label_text == "гарантия, мес":
            value = strict_single_number(raw_value)
            if value is not None:
                record["warranty_months"] = int(value)
                mapped = True
        elif label_text == "пробег тыс. км":
            value = parse_thousand_km_value(raw_value)
            if value is not None:
                record["warranty_km"] = value
                mapped = True
        elif label_text == "Межсервисный интервал, км":
            value = strict_single_number(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def insert_compass_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A45"].value) != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected A45 = 'Комплектации и опции'")
        records_by_column = {record["column_index"]: record for record in records}
        for row_index in range(46, 76):
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                if self.is_compass_ambiguous_option_status(raw_value):
                    self.insert_option(
                        {
                            "source_file": ctx["source_file"],
                            "source_sheet": ctx["source_sheet"],
                            "base_model": ctx["base_model"],
                            "comp_full_name": record.get("comp_full_name"),
                            "comp_brand": record.get("comp_brand"),
                            "option_group": "Комплектации и опции",
                            "option_name": option_name,
                            "option_status_raw": raw_value,
                            "option_status_norm": None,
                            "option_price_raw": None,
                            "option_price_rub": None,
                            "notes": None,
                            "row_order": row_index,
                            "column_order": record["column_index"],
                        }
                    )
                    self.insert_raw(
                        self.raw_row(
                            ctx,
                            sheet_type="options",
                            record_scope="option",
                            param_name_raw=option_name,
                            value_raw=raw_value,
                            row_order=row_index,
                            column_order=record["column_index"],
                            comp_record=record,
                            parse_status="text",
                            parse_comment="ambiguous_option_status",
                        )
                    )
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group="Комплектации и опции",
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def parse_sheet_compass(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_cells = {
            "A45": "Комплектации и опции",
            "A76": "Колеса, шины",
            "A77": "Колесная формула (привод)",
            "A81": "Сервис и гарантия",
            "A82": "гарантия, мес",
            "A83": "пробег тыс. км",
            "A84": "Межсервисный интервал, км",
        }
        for coord, expected_value in expected_cells.items():
            actual_value = clean_text(ws[coord].value)
            if actual_value != expected_value:
                raise ValueError(f"{ws.title}: expected {coord} = {expected_value!r}, got {actual_value!r}")

        records = self.build_compass_records(ctx)
        records_by_column = {record["column_index"]: record for record in records}
        model_rows = {
            2: "колесная база, мм",
            3: "Кабина, мест",
            4: "Полная масса а/м",
            5: "нагрузка на переднюю ось, кг",
            6: "на заднюю, кг",
            7: "Снаряженная масса а/м, кг",
            8: "снаряженная масса шасси, кг",
            9: "Грузоподьемность а/м, кг",
            10: "грузоподьемность шасси, кг",
            11: "Габариты а/м (ДхШхВ), мм",
            12: "передний свет, мм",
            13: "задний свес, мм",
            14: "Колея передняя/задняя, мм",
            15: "Габариты шасси (ДхШхВ), мм",
            16: "Грузовая платформа (ДхШ), мм",
            17: "вместимость палет, шт",
            18: "Борт, материал",
            19: "откидывание боковых бортов",
            20: "пол",
            21: "тент",
            22: "Мин. радиус разворота, м",
            23: "Погруз-я высота, мм",
            24: "Дорожный просвет (клиренс), мм",
            25: "Двигатель",
            26: "топливо",
            27: "Объем двигателя, куб.см",
            28: "Мощность двигателя, л.с.",
            29: "при об. мин.",
            30: "Крутящий момент",
            31: "при об. мин.",
            32: "Экологический класс",
            33: "Система нейтрализации",
            34: "КПП",
            35: "Тормозная система",
            36: "Тормозные механизмы",
            37: "Электронная тормозная система",
            38: "Стояночный тормоз",
            39: "Блокируемый дифференциал",
            40: "Моторный тормоз",
            41: "Подвеска передняя",
            42: "задняя",
            43: "Топливный бак, л",
            44: "Аккумулятор, АКБ",
            77: "Колесная формула (привод)",
            78: "Ошиновка",
            79: "Размер шин передних",
            80: "задних",
            82: "гарантия, мес",
            83: "пробег тыс. км",
            84: "Межсервисный интервал, км",
        }
        for row_index, expected_label in model_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
            for column_index in (2, 4, 6, 8):
                record = records_by_column[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = self.apply_compass_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        self.insert_compass_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def build_sf5_records(self, ctx: dict[str, Any]) -> list[dict[str, Any]]:
        ws = ctx["ws"]
        records: list[dict[str, Any]] = []
        for column_index in range(2, 7):
            brand = clean_text(ws.cell(1, column_index).value)
            model = clean_text(ws.cell(2, column_index).value)
            if not brand:
                raise ValueError(f"{ws.title}: expected brand in {get_column_letter(column_index)}1")
            records.append(
                self.new_model_record(
                    ctx,
                    "technical",
                    column_index,
                    join_parts(brand, model) or brand,
                    comp_brand=brand,
                    vehicle_type="van",
                    body_type="panel_van",
                )
            )
        return records

    def is_sf5_ambiguous_option_status(self, raw_value: str) -> bool:
        text = clean_text(raw_value)
        if not text:
            return False
        norm = normalize_label(text)
        has_base = "база" in norm
        has_optional = "опция" in norm
        has_not_present = "-" in text or "нет" in norm
        if has_base and has_optional:
            return True
        if (has_base or has_optional) and has_not_present and "/" in text:
            return True
        return False

    def apply_sf5_exact_label(self, record: dict[str, Any], label: str, raw_value: str) -> tuple[str, str | None]:
        label_text = clean_text(label)
        mapped = False

        if label_text == "колесная база, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["wheelbase_mm"] = value
                mapped = True
        elif label_text == "Кабина, мест":
            value = parse_additive_count(raw_value)
            if value is not None:
                record["cab_seat_count"] = value
                mapped = True
        elif label_text == "Полная масса":
            mapped = assign_numeric_range_from_raw(record, "gross_weight_kg", raw_value) or mapped
        elif label_text == "Снаряженная масса, кг":
            mapped = assign_numeric_range_from_raw(record, "curb_weight_kg", raw_value) or mapped
        elif label_text == "Грузоподьемность, кг":
            mapped = assign_numeric_range_from_raw(record, "payload_kg", raw_value) or mapped
        elif label_text == "Габариты а/м (ДхШхВ), мм":
            length_mm, width_mm, height_mm = self.parse_cmf_short_dimensions(raw_value)
            if length_mm is not None and width_mm is not None and height_mm is not None:
                record["length_mm"] = length_mm
                record["width_mm"] = width_mm
                record["height_mm"] = height_mm
                mapped = True
        elif label_text in {"Объем грузового отсека, м³", "Количество палет", "Погруз-я высота, мм", "Мин. радиус разворота, м"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Дорожный просвет, мм":
            value = strict_single_number(raw_value)
            if value is not None:
                record["ground_clearance_mm"] = value
                mapped = True
        elif label_text == "Двигатель":
            self.append_note(record, f"{label_text}: {raw_value}")
            fuel_type, volume_l, power_hp = self.parse_cmf_short_engine(raw_value)
            if fuel_type is not None:
                record["engine_fuel_type"] = fuel_type
                mapped = True
            if volume_l is not None:
                record["engine_volume_l"] = volume_l
                mapped = True
            if power_hp is not None:
                record["engine_power_hp"] = power_hp
                mapped = True
        elif label_text == "КПП":
            transmission_type, gears = self.parse_cmf_short_transmission(raw_value)
            if transmission_type is not None:
                record["transmission_type"] = transmission_type
                mapped = True
            if gears is not None:
                record["transmission_gears"] = gears
                mapped = True
        elif label_text in {"Тормозная система", "Тормозные механизмы", "Электронная тормозная система", "Стояночный тормоз"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Топливный бак, л":
            value = strict_single_number(raw_value)
            if value is not None:
                record["fuel_tank_l"] = value
                mapped = True
        elif label_text == "Колесная формула (привод)":
            wheel_formula, drive_type = self.parse_cmf_short_drive(raw_value)
            if wheel_formula is not None:
                record["wheel_formula"] = wheel_formula
                mapped = True
            if drive_type is not None:
                record["drive_type"] = drive_type
                mapped = True
        elif label_text in {"Ошиновка", "Диски", "от сквозной коррозии"}:
            self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Гарантия общая, мес./пробег тыс. км":
            months, km = parse_combined_warranty_value(raw_value)
            if months is not None:
                record["warranty_months"] = months
                mapped = True
            if km is not None:
                record["warranty_km"] = km
                mapped = True
            if "без огр" in normalize_label(raw_value):
                self.append_note(record, f"{label_text}: {raw_value}")
        elif label_text == "Межсервисный интервал, км":
            value = self.parse_cmf_short_service_interval(raw_value)
            if value is not None:
                record["service_interval_km"] = value
                mapped = True
        else:
            raise ValueError(f"{record['source_sheet']}: unexpected label {label_text!r}")

        status = "parsed" if mapped else ("partial" if number_tokens(raw_value) else "text")
        return status, None

    def insert_sf5_option_rows(self, ctx: dict[str, Any], records: list[dict[str, Any]]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A21"].value) != "Комплектации и опции":
            raise ValueError(f"{ws.title}: expected A21 = 'Комплектации и опции'")
        if clean_text(ws["A43"].value) != "Грузовой фургон":
            raise ValueError(f"{ws.title}: expected A43 = 'Грузовой фургон'")
        records_by_column = {record["column_index"]: record for record in records}
        option_rows = {**{row_index: "Комплектации и опции" for row_index in range(22, 43)}, **{row_index: "Грузовой фургон" for row_index in range(44, 48)}}
        for row_index, option_group in option_rows.items():
            option_name = clean_text(ws.cell(row_index, 1).value)
            if not option_name:
                raise ValueError(f"{ws.title}: expected option name in A{row_index}")
            for column_index, record in records_by_column.items():
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                if self.is_sf5_ambiguous_option_status(raw_value):
                    self.insert_option(
                        {
                            "source_file": ctx["source_file"],
                            "source_sheet": ctx["source_sheet"],
                            "base_model": ctx["base_model"],
                            "comp_full_name": record.get("comp_full_name"),
                            "comp_brand": record.get("comp_brand"),
                            "option_group": option_group,
                            "option_name": option_name,
                            "option_status_raw": raw_value,
                            "option_status_norm": None,
                            "option_price_raw": None,
                            "option_price_rub": None,
                            "notes": None,
                            "row_order": row_index,
                            "column_order": record["column_index"],
                        }
                    )
                    self.insert_raw(
                        self.raw_row(
                            ctx,
                            sheet_type="options",
                            record_scope="option",
                            param_name_raw=option_name,
                            value_raw=raw_value,
                            row_order=row_index,
                            column_order=record["column_index"],
                            comp_record=record,
                            parse_status="text",
                            parse_comment="ambiguous_option_status",
                        )
                    )
                    continue
                self.insert_option_row(
                    ctx,
                    record=record,
                    option_group=option_group,
                    option_name=option_name,
                    row_order=row_index,
                    status_override=raw_value,
                    status_only=True,
                )

    def parse_sheet_sf5(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        expected_cells = {
            "A2": "Модель",
            "A21": "Комплектации и опции",
            "A43": "Грузовой фургон",
            "A48": "Колесная формула (привод)",
            "A49": "Ошиновка",
            "A50": "Диски",
            "A51": "Гарантия общая, мес./пробег тыс. км",
            "A52": "от сквозной коррозии",
            "A53": "Межсервисный интервал, км",
        }
        for coord, expected_value in expected_cells.items():
            actual_value = clean_text(ws[coord].value)
            if actual_value != expected_value:
                raise ValueError(f"{ws.title}: expected {coord} = {expected_value!r}, got {actual_value!r}")

        records = self.build_sf5_records(ctx)
        records_by_column = {record["column_index"]: record for record in records}
        model_rows = {
            3: "колесная база, мм",
            4: "Кабина, мест",
            5: "Полная масса",
            6: "Снаряженная масса, кг",
            7: "Грузоподьемность, кг",
            8: "Габариты а/м (ДхШхВ), мм",
            9: "Объем грузового отсека, м³",
            10: "Количество палет",
            11: "Погруз-я высота, мм",
            12: "Дорожный просвет, мм",
            13: "Мин. радиус разворота, м",
            14: "Двигатель",
            15: "КПП",
            16: "Тормозная система",
            17: "Тормозные механизмы",
            18: "Электронная тормозная система",
            19: "Стояночный тормоз",
            20: "Топливный бак, л",
            48: "Колесная формула (привод)",
            49: "Ошиновка",
            50: "Диски",
            51: "Гарантия общая, мес./пробег тыс. км",
            52: "от сквозной коррозии",
            53: "Межсервисный интервал, км",
        }
        for row_index, expected_label in model_rows.items():
            actual_label = clean_text(ws.cell(row_index, 1).value)
            if actual_label != expected_label:
                raise ValueError(f"{ws.title}: expected A{row_index} = {expected_label!r}, got {actual_label!r}")
            for column_index in range(2, 7):
                record = records_by_column[column_index]
                raw_value = clean_text(ws.cell(row_index, column_index).value)
                if raw_value is None:
                    continue
                parse_status, parse_comment = self.apply_sf5_exact_label(record, expected_label, raw_value)
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="technical",
                        record_scope="model",
                        param_name_raw=expected_label,
                        value_raw=raw_value,
                        row_order=row_index,
                        column_order=column_index,
                        comp_record=record,
                        parse_status=parse_status,
                        parse_comment=parse_comment,
                    )
                )

        self.insert_sf5_option_rows(ctx, records)
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_bort_stvlad(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "A1", "Производитель")
        records = self.models_from_full_name_row(
            ctx,
            1,
            start_col=2,
            allowed_cols=[2, 3, 4, 5, 6, 7, 8, 9, 10],
            sheet_type="ownership_summary",
            vehicle_type="truck",
            body_type="flatbed",
        )
        self.ingest_model_matrix(ctx, sheet_type="ownership_summary", records=records, label_col=1, row_start=2, row_end=ws.max_row, apply_label=self.apply_bort_stvlad_label)
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_cmf_stvlad(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "A1", "Производитель")
        records = self.models_from_full_name_row(
            ctx,
            1,
            start_col=2,
            allowed_cols=[2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14],
            sheet_type="ownership_summary",
            vehicle_type="van",
            body_type="panel_van",
        )
        self.ingest_model_matrix(
            ctx,
            sheet_type="ownership_summary",
            records=records,
            label_col=1,
            row_start=2,
            row_end=ws.max_row,
            apply_label=self.apply_cmf_stvlad_label,
        )
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_ldt_stvlad(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "A2", "Производитель")
        records = self.models_from_full_name_row(
            ctx,
            2,
            start_col=2,
            allowed_cols=[2, 3, 4, 5, 6, 7, 8, 9, 10],
            sheet_type="ownership_summary",
            vehicle_type="truck",
            body_type="chassis",
        )
        self.ingest_model_matrix(
            ctx,
            sheet_type="ownership_summary",
            records=records,
            label_col=1,
            row_start=3,
            row_end=ws.max_row,
            apply_label=self.apply_ldt_stvlad_label,
        )
        for record in records:
            self.insert_normalized(record)

    def parse_sheet_to_bort(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "B3", "ТО")
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["B2"].value) or "group_1",
            row_start=4,
            row_end=16,
            service_col=2,
            mileage_col=3,
            diesel_col=4,
            gasoline_col=5,
        )
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["F2"].value) or "group_2",
            row_start=4,
            row_end=14,
            service_col=6,
            mileage_col=7,
            diesel_col=8,
            diesel_promo_col=9,
            gasoline_col=10,
            gasoline_promo_col=11,
        )
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["L2"].value) or "group_3",
            row_start=4,
            row_end=16,
            service_col=12,
            mileage_col=13,
            diesel_col=14,
        )

    def parse_sheet_to_cmf(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "B3", "ТО")
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["D2"].value) or "group_1",
            row_start=5,
            row_end=14,
            service_col=2,
            mileage_col=3,
            diesel_col=4,
        )
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["E2"].value) or "group_2",
            row_start=5,
            row_end=14,
            service_col=2,
            mileage_col=3,
            diesel_col=5,
            gasoline_col=6,
        )
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["I2"].value) or "group_3",
            row_start=4,
            row_end=14,
            service_col=7,
            mileage_col=8,
            diesel_col=9,
            diesel_promo_col=10,
        )
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["K2"].value) or "group_4",
            row_start=4,
            row_end=14,
            service_col=7,
            mileage_col=8,
            diesel_col=11,
            diesel_promo_col=12,
        )
        self.parse_service_group(
            ctx,
            group_name=clean_text(ws["M2"].value) or "group_5",
            row_start=4,
            row_end=16,
            service_col=13,
            mileage_col=14,
            diesel_col=15,
        )

    def parse_sheet_complectation(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_contains(ws, "B5", "Соболь")
        self.require_contains(ws, "D5", "ATLANT")
        left_name = clean_text(ws["B5"].value)
        right_name = clean_text(ws["D5"].value)
        if not left_name or not right_name:
            raise ValueError(f"{ws.title}: expected model headers in B5 and D5")
        for column_index in [5, 6]:
            extra_values = [clean_text(ws.cell(row_index, column_index).value) for row_index in range(6, ws.max_row + 1)]
            if any(value is not None for value in extra_values):
                raise ValueError(f"{ws.title}: unexpected option values in {get_column_letter(column_index)}6:{get_column_letter(column_index)}{ws.max_row}")
        left_record = self.new_model_record(
            ctx,
            "options",
            2,
            left_name,
            comp_brand=left_name.split()[0],
            vehicle_type="minivan",
            body_type="minivan",
        )
        right_record = self.new_model_record(
            ctx,
            "options",
            4,
            right_name,
            comp_brand=right_name.split()[0],
            vehicle_type="minivan",
            body_type="minivan",
        )
        current_group = None
        for row_index in range(6, ws.max_row + 1):
            group_value = clean_text(ws.cell(row_index, 3).value)
            if group_value:
                current_group = group_value
            left_option = clean_text(ws.cell(row_index, 2).value)
            right_option = clean_text(ws.cell(row_index, 4).value)
            if current_group is None and (left_option or right_option):
                raise ValueError(f"{ws.title}: option row {row_index} has no group in column C")
            if left_option:
                self.insert_option_row(
                    ctx,
                    record=left_record,
                    option_group=current_group,
                    option_name=left_option,
                    row_order=row_index,
                    status_override="listed",
                )
            if right_option:
                self.insert_option_row(
                    ctx,
                    record=right_record,
                    option_group=current_group,
                    option_name=right_option,
                    row_order=row_index,
                    status_override="listed",
                )

    def parse_sheet_advantages(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Преимущество")
        self.require_exact(ws, "B1", "Газель ЦМФ")
        self.require_exact(ws, "C1", "Sollers ЦМФ")
        self.require_exact(ws, "E1", "Преимущество")
        self.require_exact(ws, "F1", "Газель Автобус")
        self.require_exact(ws, "G1", "Sollers Автобус")
        self.require_exact(ws, "I1", "Преимущество")
        self.require_exact(ws, "J1", "Борт")
        self.require_exact(ws, "K1", "ЦМФ")
        self.require_exact(ws, "L1", "Автобус")
        self.require_exact(ws, "A27", "Отличная маневренность")
        self.require_exact(ws, "E27", "Отличная маневренность")
        self.require_exact(ws, "I39", "Конкурентоспособная стоимость владения автомобилем")

        self.insert_advantages_text_block(
            ctx,
            label_col=1,
            left_value_col=2,
            right_value_col=3,
            row_start=4,
            row_end=40,
        )
        self.insert_advantages_text_block(
            ctx,
            label_col=5,
            left_value_col=6,
            right_value_col=7,
            row_start=3,
            row_end=40,
        )
        self.insert_advantages_summary_block(
            ctx,
            label_col=9,
            value_cols=[10, 11, 12],
            row_start=3,
            row_end=40,
        )

    def parse_sheet_advantages_minivan(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Преимущество")
        self.require_exact(ws, "B1", "Соболь НН Минивэн")
        self.require_exact(ws, "C1", "Соболь НН Минивэн 4х4")
        self.require_exact(ws, "D1", "Sollers Atlant 7+1 Тандем")
        self.require_exact(ws, "E1", "Largus Универсал 6+1 мест")
        self.require_exact(ws, "F1", "Largus Фургон")
        self.require_exact(ws, "G1", "УАЗ Patriot Base ИКАР Limited")
        self.require_exact(ws, "A20", "Отличная маневренность")
        self.require_exact(ws, "A29", "Конкурентоспособная стоимость владения автомобилем")
        self.require_contains(ws, "A32", "Демонтаж сидений")

        group_names = {column_index: clean_text(ws.cell(1, column_index).value) for column_index in range(2, 8)}
        current_label: str | None = None
        for row_index in range(2, 31):
            label = clean_text(ws.cell(row_index, 1).value)
            if label:
                current_label = label
            if not current_label:
                continue
            for column_index in range(2, 8):
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_label,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=group_names[column_index],
                    )
                )

        footnote = clean_text(ws["A32"].value)
        if footnote is not None:
            self.insert_raw(
                self.raw_row(
                    ctx,
                    sheet_type="qualitative",
                    record_scope="qualitative",
                    param_name_raw="Трансформируемый салон и пассажировместимость",
                    value_raw=footnote,
                    row_order=32,
                    column_order=1,
                    group_name=None,
                )
            )

    def parse_sheet_reviews(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "ОТЗЫВЫ за 2025 год")
        self.require_exact(ws, "A2", "✗ Отрицательные отзывы Газель NEXT")
        self.require_exact(ws, "D2", "✗ Отрицательные отзывы Sollers Argo")
        self.require_exact(ws, "G2", "✗ Отрицательные отзывы Sollers Atlant")
        self.require_exact(ws, "J2", "✗ Отрицательные отзывы Dongfeng K33/К39")
        self.require_exact(ws, "A52", "✓ Положительные отзывы Газель NEXT")
        self.require_exact(ws, "D52", "✓ Положительные отзывы Sollers Argo")
        self.require_exact(ws, "G52", "✓ Положительные отзывы Sollers Atlant")
        self.require_exact(ws, "J52", "✓ Положительные отзывы Dongfeng K33/К39")
        self.require_exact(ws, "A54", "Общее")
        self.require_exact(ws, "D54", "Общее")
        self.require_exact(ws, "G54", "Общее")

        self.insert_reviews_block(
            ctx,
            text_col=1,
            count_col=2,
            header_row=2,
            body_row_start=3,
            body_row_end=50,
            category_rows=[3, 11, 19, 28, 37, 44],
        )
        self.insert_reviews_block(
            ctx,
            text_col=4,
            count_col=5,
            header_row=2,
            body_row_start=3,
            body_row_end=50,
            category_rows=[3, 13, 17, 22, 26],
        )
        self.insert_reviews_block(
            ctx,
            text_col=7,
            count_col=8,
            header_row=2,
            body_row_start=3,
            body_row_end=50,
            category_rows=[3, 14, 19, 27, 32, 39, 44, 49],
        )
        self.insert_reviews_block(
            ctx,
            text_col=10,
            count_col=11,
            header_row=2,
            body_row_start=3,
            body_row_end=8,
            category_rows=[3],
        )
        self.insert_reviews_block(
            ctx,
            text_col=1,
            count_col=2,
            header_row=52,
            body_row_start=54,
            body_row_end=62,
            category_rows=[54],
        )
        self.insert_reviews_block(
            ctx,
            text_col=4,
            count_col=5,
            header_row=52,
            body_row_start=54,
            body_row_end=59,
            category_rows=[54],
        )
        self.insert_reviews_block(
            ctx,
            text_col=7,
            count_col=8,
            header_row=52,
            body_row_start=54,
            body_row_end=59,
            category_rows=[54],
        )
        self.insert_reviews_block(
            ctx,
            text_col=10,
            count_col=11,
            header_row=52,
            body_row_start=53,
            body_row_end=53,
            category_rows=[],
        )

    def parse_sheet_reviews_lower(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "ОТЗЫВЫ за 2025 год")
        self.require_exact(ws, "A2", "✗ Отрицательные отзывы Газон NEXT - 154 отзыва всего")
        self.require_exact(ws, "C2", "✗ Отрицательные отзывы Isuzu ELF")
        self.require_exact(ws, "E2", "✗ Отрицательные отзывы КАМАЗ Компас 9")
        self.require_exact(ws, "G2", "✗ Отрицательные отзывы КАМАЗ Компас 12")
        self.require_exact(ws, "I2", "✗ Отрицательные отзывы JAC N90")
        self.require_exact(ws, "K2", "✗ Отрицательные отзывы JAC N120")
        self.require_exact(ws, "M2", "✗ Отрицательные отзывы DonFeng 80")
        self.require_exact(ws, "O2", "✗ Отрицательные отзывы DonFeng 120")
        self.require_exact(ws, "A73", "✓ Положительные отзывы Газон NEXT")
        self.require_exact(ws, "C73", "✓ Положительные отзывы Газон NEXT")
        self.require_exact(ws, "E73", "✓ Положительные отзывы КАМАЗ Компас 9")
        self.require_exact(ws, "K73", "✓ Положительные отзывы JAC N120")

        self.insert_reviews_text_only_block(
            ctx,
            text_col=1,
            header_row=2,
            body_row_start=3,
            body_row_end=71,
            category_rows=[3, 12, 19, 25, 37, 43, 48, 55, 59, 63, 68],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=3,
            header_row=2,
            body_row_start=3,
            body_row_end=24,
            category_rows=[3, 11, 14, 18],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=5,
            header_row=2,
            body_row_start=3,
            body_row_end=17,
            category_rows=[3, 8, 12],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=7,
            header_row=2,
            body_row_start=3,
            body_row_end=28,
            category_rows=[3, 8, 12, 17],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=9,
            header_row=2,
            body_row_start=3,
            body_row_end=26,
            category_rows=[3, 7, 13, 20, 24],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=11,
            header_row=2,
            body_row_start=3,
            body_row_end=20,
            category_rows=[3, 8, 19],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=13,
            header_row=2,
            body_row_start=4,
            body_row_end=5,
            category_rows=[],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=15,
            header_row=2,
            body_row_start=4,
            body_row_end=4,
            category_rows=[],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=1,
            header_row=73,
            body_row_start=74,
            body_row_end=78,
            category_rows=[],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=5,
            header_row=73,
            body_row_start=74,
            body_row_end=74,
            category_rows=[],
        )
        self.insert_reviews_text_only_block(
            ctx,
            text_col=11,
            header_row=73,
            body_row_start=74,
            body_row_end=87,
            category_rows=[],
        )

    def parse_sheet_client_reviews_nn(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Что нравится в Соболь NN, Газель NN")
        self.require_exact(ws, "B1", "Что не нравится в Соболь NN, Газель NN")
        self.insert_qualitative_column(ctx, text_col=1, header_row=1, row_start=2, row_end=49)
        self.insert_qualitative_column(ctx, text_col=2, header_row=1, row_start=2, row_end=49)

    def parse_sheet_client_reviews_atlant(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Что нравится в Атлант Соллерс")
        self.require_exact(ws, "B1", "Что не нравится в Атлант Соллерс")
        self.insert_qualitative_column(ctx, text_col=1, header_row=1, row_start=2, row_end=27)
        self.insert_qualitative_column(ctx, text_col=2, header_row=1, row_start=2, row_end=27)

    def parse_sheet_negative_positive_orm(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Соболь НН")
        self.require_exact(ws, "B3", "Негатив")
        self.require_exact(ws, "F3", "Позитив")
        self.require_exact(ws, "A16", "Соболь NN 4х4")
        self.require_exact(ws, "B17", "Негатив")
        self.require_exact(ws, "F17", "Позитив")
        self.require_exact(ws, "A26", "Sollers SF1")
        self.require_exact(ws, "B27", "Негатив")
        self.require_exact(ws, "F27", "Позитив")
        self.require_exact(ws, "A34", "УАЗ патриот")
        self.require_exact(ws, "B35", "Негатив")
        self.require_exact(ws, "F35", "Позитив")
        self.require_exact(ws, "A42", "УАЗ Профи")
        self.require_exact(ws, "B43", "Негатив")
        self.require_exact(ws, "F43", "Позитив")

        self.insert_negative_positive_orm_section(
            ctx,
            title_row=1,
            header_row=3,
            body_row_start=4,
            body_row_end=11,
            summary_row=14,
        )
        self.insert_negative_positive_orm_section(
            ctx,
            title_row=16,
            header_row=17,
            body_row_start=18,
            body_row_end=23,
            summary_row=24,
        )
        self.insert_negative_positive_orm_section(
            ctx,
            title_row=26,
            header_row=27,
            body_row_start=28,
            body_row_end=31,
            summary_row=32,
        )
        self.insert_negative_positive_orm_section(
            ctx,
            title_row=34,
            header_row=35,
            body_row_start=36,
            body_row_end=40,
            summary_row=41,
        )
        self.insert_negative_positive_orm_section(
            ctx,
            title_row=42,
            header_row=43,
            body_row_start=44,
            body_row_end=48,
            summary_row=49,
        )

    def parse_sheet_quality(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        if clean_text(ws["A1"].value) == "Справка" and clean_text(ws["A7"].value) == "Показатель":
            self.parse_sheet_quality_gazelle(ctx)
            return
        if clean_text(ws["A1"].value) == "Общее - 10 т" and clean_text(ws["J1"].value) == "Общее - 8,7 т":
            self.parse_sheet_quality_sadko(ctx)
            return
        raise ValueError(f"{ws.title}: unsupported качество layout")

    def parse_sheet_dpo_stats(self, ctx: dict[str, Any]) -> None:
        ws = ctx["ws"]
        self.require_exact(ws, "A1", "Справка")
        self.require_exact(ws, "A7", "Показатель")
        self.require_exact(ws, "A18", "Подразделения")
        self.require_exact(ws, "A29", "Узел")
        self.require_exact(ws, "C64", "Основные дефекты")
        self.require_exact(ws, "A66", "№ п/п")
        self.require_exact(ws, "C66", "Наименование дефекта")

        for row_index, column_index in (
            (2, 1),
            (3, 1),
            (5, 2),
            (5, 3),
            (16, 2),
            (16, 3),
            (27, 2),
            (27, 3),
            (62, 2),
            (64, 2),
            (64, 3),
        ):
            self.insert_quality_note_cell(ctx, group_name="Справка", row_index=row_index, column_index=column_index)

        metric_headers_main = {
            14: clean_text(ws["N7"].value),
            19: clean_text(ws["S7"].value),
            24: clean_text(ws["X7"].value),
            29: clean_text(ws["AC7"].value),
            35: clean_text(ws["AI7"].value),
        }
        current_metric: str | None = None
        for row_index in range(8, 15):
            label = clean_text(ws.cell(row_index, 1).value)
            if label:
                current_metric = label
            if current_metric is None:
                continue
            if row_index == 8:
                special_values = {
                    14: clean_text(ws["N8"].value),
                    19: clean_text(ws["S8"].value),
                    24: clean_text(ws["X8"].value),
                    34: clean_text(ws["AH8"].value),
                    40: clean_text(ws["AN8"].value),
                }
                special_comments = {
                    14: clean_text(ws["N7"].value),
                    19: clean_text(ws["S7"].value),
                    24: clean_text(ws["X7"].value),
                    34: clean_text(ws["AC7"].value),
                    40: clean_text(ws["AI7"].value),
                }
                for column_index, value in special_values.items():
                    if value is None:
                        continue
                    self.insert_raw(
                        self.raw_row(
                            ctx,
                            sheet_type="qualitative",
                            record_scope="qualitative",
                            param_name_raw=current_metric,
                            value_raw=value,
                            row_order=row_index,
                            column_order=column_index,
                            group_name=clean_text(ws["A2"].value),
                            parse_comment=special_comments[column_index],
                        )
                    )
                continue
            for column_index, header in metric_headers_main.items():
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=current_metric,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=clean_text(ws["A2"].value),
                        parse_comment=header,
                    )
                )

        self.insert_quality_metric_table(
            ctx,
            group_name=clean_text(ws["C16"].value) or "Статистика ДПО",
            header_row=19,
            row_start=20,
            row_end=25,
            label_col=1,
            value_cols=[11, 16, 21, 26, 31, 36],
        )
        self.insert_quality_metric_table(
            ctx,
            group_name=clean_text(ws["C27"].value) or "Статистика ДПО",
            header_row=30,
            row_start=31,
            row_end=60,
            label_col=1,
            value_cols=[11, 16, 21, 26, 31, 36],
        )

        defect_headers = {
            1: clean_text(ws["A66"].value),
            23: clean_text(ws["W66"].value),
            25: clean_text(ws["Y66"].value),
            27: clean_text(ws["AA66"].value),
            29: clean_text(ws["AC66"].value),
            31: clean_text(ws["AE66"].value),
            33: clean_text(ws["AG66"].value),
            35: clean_text(ws["AI66"].value),
        }
        defect_index = 1
        for row_index in range(67, ws.max_row + 1):
            raw_rank = clean_text(ws.cell(row_index, 1).value)
            defect_name = clean_text(ws.cell(row_index, 3).value)
            if raw_rank is None and defect_name is None:
                continue
            if defect_name == "Всего:":
                for column_index, header in defect_headers.items():
                    if column_index == 1:
                        continue
                    value = clean_text(ws.cell(row_index, column_index).value)
                    if value is None:
                        continue
                    self.insert_raw(
                        self.raw_row(
                            ctx,
                            sheet_type="qualitative",
                            record_scope="qualitative",
                            param_name_raw="Всего:",
                            value_raw=value,
                            row_order=row_index,
                            column_order=column_index,
                            group_name=clean_text(ws["C64"].value),
                            parse_comment=header,
                        )
                    )
                break
            if raw_rank is None or defect_name is None:
                raise ValueError(f"{ws.title}: incomplete defect row {row_index}")
            try:
                rank_value = int(float(raw_rank))
            except ValueError as exc:
                raise ValueError(f"{ws.title}: invalid defect index at row {row_index}: {raw_rank!r}") from exc
            if rank_value != defect_index:
                raise ValueError(
                    f"{ws.title}: expected defect index {defect_index} at row {row_index}, got {rank_value}"
                )
            for column_index, header in defect_headers.items():
                value = clean_text(ws.cell(row_index, column_index).value)
                if value is None:
                    continue
                self.insert_raw(
                    self.raw_row(
                        ctx,
                        sheet_type="qualitative",
                        record_scope="qualitative",
                        param_name_raw=defect_name,
                        value_raw=value,
                        row_order=row_index,
                        column_order=column_index,
                        group_name=clean_text(ws["C64"].value),
                        parse_comment=header,
                    )
                )
            defect_index += 1


def main() -> int:
    importer = GazPricingImporter(SOURCE_DIR, DB_PATH)
    importer.run()
    sys.stdout.write(
        "Imported workbooks={workbooks} selected_sheets={selected_sheets} staged={staged_rows} "
        "merged={merged_rows} warnings={warnings}\n".format(**importer.stats)
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
